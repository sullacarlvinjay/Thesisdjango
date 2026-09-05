"""The TES applicants report — a plain table — and the school-year picker.

Two things are being guarded here. That the report is generated rather than
being CHED's Annex 1 workbook with rows poured into it: that file is the guide
this takes its columns and its lookup lists from, and none of its sheets,
dropdowns or macros belong in the output. And the school year: that picking one
scopes the applicants page, this download and the Annex 2 report alike, so what
an officer looks at is what they generate.
"""
import openpyxl
from unittest import mock
from datetime import date
from io import BytesIO

from django.test import TestCase, Client

from api import annex1_report, tes_report
from api.models import (
    User, StudentProfile, Scholarship, TESApplication, SystemSettings,
)


class Annex1ReportTest(TestCase):
    def setUp(self):
        SystemSettings.objects.create(pk=1, academic_year='26-1',
                                      active_semester='1st Semester')
        Scholarship.objects.create(
            name='TES', type='TES', category='application',
            description='x', eligibility='x', requirements=[],
        )
        self.unifast = User.objects.create_user(
            username='unifast@bipsu.edu.ph', email='unifast@bipsu.edu.ph',
            password='pw', role='unifast',
        )
        self.c = Client()
        self.assertTrue(self.c.login(email='unifast@bipsu.edu.ph', password='pw'))

    def _applicant(self, last, sid, school_year='2026-2027', **kwargs):
        u = User.objects.create_user(
            username=f'{sid}@bipsu.edu.ph', email=f'{sid}@bipsu.edu.ph',
            password='pw', first_name='Juan', last_name=last, role='student',
        )
        p = StudentProfile.objects.create(
            user=u, student_id=sid, course='BSCS', year_level=2,
            gender='Female', middle_name='Santos', suffix='Jr.',
            contact_number='09171234567',
        )
        p.father_last_name = last
        p.father_first_name = 'Pedro'
        p.mother_last_name = 'Reyes'
        p.mother_first_name = 'Maria'
        p.save()
        fields = {
            'lrn': '123456789012',
            'birthdate': date(2005, 3, 14),
            'complete_program': 'BACHELOR OF SCIENCE IN COMPUTER SCIENCE',
            'street_barangay': 'Purok 1, Poblacion',
            'city_municipality': 'Naval',
            'province': 'Biliran',
            'region': 'Region VIII',
            'zip_code': '6560',
            'contact_number': '09171234567',
            'email_address': f'{sid}@bipsu.edu.ph',
        }
        fields.update(kwargs)
        return TESApplication.objects.create(
            student=p, school_year=school_year, semester='1st Semester', **fields)

    # ── The rows ────────────────────────────────────────────────────────────

    def test_every_applicant_is_listed_whatever_the_decision(self):
        """Annex 1 is the applicant list, not the grantee list."""
        self._applicant('Cruz', 'S1', status='Pending')
        self._applicant('Bautista', 'S2', status='Approved')
        self._applicant('Aquino', 'S3', status='Rejected')

        rows = annex1_report.applicant_rows(school_year='2026-2027')

        self.assertEqual([r['last_name'] for r in rows],
                         ['Aquino', 'Bautista', 'Cruz'])
        self.assertEqual([r['seq'] for r in rows], [1, 2, 3])

    def test_rows_are_scoped_to_the_school_year(self):
        self._applicant('Cruz', 'S1', school_year='2026-2027')
        self._applicant('Reyes', 'S2', school_year='2025-2026')

        self.assertEqual(
            [r['last_name'] for r in annex1_report.applicant_rows(school_year='2026-2027')],
            ['Cruz'])
        self.assertEqual(
            [r['last_name'] for r in annex1_report.applicant_rows(school_year='2025-2026')],
            ['Reyes'])
        # No year means every year, which is how the review queue reads.
        self.assertEqual(len(annex1_report.applicant_rows()), 2)

    def test_sex_is_the_code_the_form_validates_against(self):
        """0 = Male, 1 = Female — the Sex_Code sheet, not the word."""
        self._applicant('Cruz', 'S1')
        male = self._applicant('Reyes', 'S2')
        male.student.gender = 'Male'
        male.student.save()

        rows = {r['last_name']: r['sex'] for r in annex1_report.applicant_rows()}
        self.assertEqual(rows['Cruz'], 1)
        self.assertEqual(rows['Reyes'], 0)

    def test_mobile_number_loses_its_leading_zero(self):
        """CHED asks for 10 digits starting with 9; the profile holds 11 with a 0."""
        self._applicant('Cruz', 'S1', contact_number='0917 123 4567')
        self.assertEqual(annex1_report.applicant_rows()[0]['contact'], '9171234567')

    def test_a_landline_is_passed_through_unchanged(self):
        self._applicant('Cruz', 'S1', contact_number='53500123')
        self.assertEqual(annex1_report.applicant_rows()[0]['contact'], '53500123')

    def test_blank_disability_and_ip_group_export_as_no(self):
        """Both columns spell 'not applicable' as NO on this form."""
        self._applicant('Cruz', 'S1', disability_type='', indigenous_people_group='')
        row = annex1_report.applicant_rows()[0]
        self.assertEqual(row['disability'], 'NO')
        self.assertEqual(row['ip_group'], 'NO')

    def test_a_real_disability_is_kept(self):
        self._applicant('Cruz', 'S1', disability_type='Visual Disability')
        self.assertEqual(annex1_report.applicant_rows()[0]['disability'],
                         'Visual Disability')

    def test_priority_flags_export_as_one_and_zero(self):
        self._applicant('Cruz', 'S1', is_solo_parent_dependent=True,
                        is_first_gen_college=False)
        row = annex1_report.applicant_rows()[0]
        self.assertEqual(row['solo_parent'], 1)
        self.assertEqual(row['first_gen'], 0)

    def test_uncollected_id_numbers_export_blank(self):
        """PhilSys and 4Ps IDs are optional and this system never asks for them."""
        self._applicant('Cruz', 'S1')
        row = annex1_report.applicant_rows()[0]
        self.assertEqual(row['philsys_id'], '')
        self.assertEqual(row['four_ps_id'], '')

    def test_row_values_match_the_header_count(self):
        self._applicant('Cruz', 'S1')
        values = annex1_report.row_values(annex1_report.applicant_rows()[0])
        self.assertEqual(len(values), len(annex1_report.ANNEX1_HEADERS))

    # ── The sheet ───────────────────────────────────────────────────────────

    def _sheet(self, school_year='2026-2027', **kwargs):
        buf, written = annex1_report.build_workbook(school_year, **kwargs)
        wb = openpyxl.load_workbook(BytesIO(buf.getvalue()))
        return wb, wb[annex1_report.SHEET], written

    def test_it_is_a_plain_table_not_cheds_workbook(self):
        """The Annex 1 file is the guide for the columns, not the output."""
        self._applicant('Cruz', 'S1')
        wb, ws, written = self._sheet()
        self.assertEqual(written, 1)

        # One sheet of our own — none of the form's tabs come along.
        self.assertEqual(wb.sheetnames, [annex1_report.SHEET])
        for tab in ('Annex 1', 'General Instructions', 'Registry_Courses'):
            self.assertNotIn(tab, wb.sheetnames)
        # Nothing inherited from the form: no dropdowns, no macros.
        self.assertEqual(len(ws.data_validations.dataValidation), 0)
        self.assertIsNone(wb.vba_archive)

    def test_the_header_row_is_the_columns_the_guide_names(self):
        self._applicant('Cruz', 'S1')
        _wb, ws, _ = self._sheet()

        row = annex1_report.HEADER_ROW
        headings = [ws.cell(row=row, column=c).value
                    for c in range(1, len(annex1_report.ANNEX1_HEADERS) + 1)]
        self.assertEqual(headings, annex1_report.ANNEX1_HEADERS)

    def test_one_row_per_applicant_under_the_headings(self):
        self._applicant('Cruz', 'S1')
        self._applicant('Bautista', 'S2')
        _wb, ws, written = self._sheet()
        self.assertEqual(written, 2)

        first = annex1_report.FIRST_ROW
        self.assertEqual(ws.cell(row=first, column=1).value, 1)          # SEQ
        self.assertEqual(ws.cell(row=first, column=2).value, 'S2')       # alphabetical
        self.assertEqual(ws.cell(row=first, column=6).value, 'Bautista')
        self.assertEqual(ws.cell(row=first + 1, column=6).value, 'Cruz')
        # And nothing past the last applicant.
        self.assertIsNone(ws.cell(row=first + 2, column=2).value)

    def test_the_computed_columns_land_where_the_guide_puts_them(self):
        self._applicant('Cruz', 'S1')
        _wb, ws, _ = self._sheet()
        first = annex1_report.FIRST_ROW
        self.assertEqual(ws.cell(row=first, column=10).value, 1)          # SEX, female
        self.assertEqual(ws.cell(row=first, column=25).value, '9171234567')
        self.assertEqual(ws.cell(row=first, column=27).value, 'NO')       # disability

    def test_the_title_says_what_the_table_is_and_which_year(self):
        self._applicant('Cruz', 'S1')
        _wb, ws, _ = self._sheet()
        title = ws['A1'].value
        self.assertIn('LIST OF TES APPLICANTS', title)
        self.assertIn('Biliran Province State University', title)
        self.assertIn('2026-2027', title)

    def test_the_birthdate_column_is_formatted_as_a_date(self):
        """Otherwise Excel shows it in whatever the reader's locale prefers."""
        self._applicant('Cruz', 'S1')
        _wb, ws, _ = self._sheet()
        self.assertEqual(ws.cell(row=annex1_report.FIRST_ROW, column=11).number_format,
                         'yyyy-mm-dd')

    def test_the_headings_freeze_and_the_table_filters(self):
        """A thirty-column list is unreadable without both."""
        self._applicant('Cruz', 'S1')
        _wb, ws, _ = self._sheet()
        self.assertEqual(ws.freeze_panes, f'A{annex1_report.FIRST_ROW}')
        self.assertIsNotNone(ws.auto_filter.ref)

    def test_it_scopes_itself_to_the_year_it_is_titled_with(self):
        self._applicant('Cruz', 'S1', school_year='2026-2027')
        self._applicant('Reyes', 'S2', school_year='2025-2026')

        _wb, ws, written = self._sheet('2025-2026')
        self.assertEqual(written, 1)
        self.assertIn('2025-2026', ws['A1'].value)
        self.assertEqual(ws.cell(row=annex1_report.FIRST_ROW, column=6).value, 'Reyes')

    def test_a_year_with_no_applicants_still_builds_the_table(self):
        _wb, ws, written = self._sheet('2024-2025')
        self.assertEqual(written, 0)
        # The headings are there to be read even with nothing under them.
        self.assertEqual(ws.cell(row=annex1_report.HEADER_ROW, column=1).value, 'SEQ')
        self.assertIsNone(ws.cell(row=annex1_report.FIRST_ROW, column=1).value)

    def test_it_builds_without_the_guide_file(self):
        """The guide feeds the apply form's dropdowns; the report does not need it."""
        with mock.patch.object(annex1_report, 'TEMPLATE_PATH', '/nowhere/missing.xlsm'):
            _buf, written = annex1_report.build_workbook('2026-2027')
        self.assertEqual(written, 0)

    # ── The school-year options ─────────────────────────────────────────────

    def test_year_options_cover_the_data_and_the_active_term(self):
        self._applicant('Cruz', 'S1', school_year='2025-2026')
        # 2026-2027 is the active term and has no applications yet; it is still
        # offered, or the office could not report on the year it is working in.
        self.assertEqual(tes_report.school_year_options(),
                         ['2026-2027', '2025-2026'])

    def test_year_options_are_not_repeated(self):
        self._applicant('Cruz', 'S1', school_year='2026-2027')
        self._applicant('Reyes', 'S2', school_year='2026-2027')
        self.assertEqual(tes_report.school_year_options(), ['2026-2027'])

    # ── The pages ───────────────────────────────────────────────────────────

    def test_applications_page_offers_the_years_and_the_generator(self):
        self._applicant('Cruz', 'S1')
        r = self.c.get('/unifast/tes-applications/')
        self.assertContains(r, '/unifast/tes-applications/report/')
        self.assertContains(r, 'All school years')
        self.assertContains(r, '2026-2027')

    def test_the_controls_sit_on_the_applications_card(self):
        """No generator panel of its own — they are on the list they act on."""
        r = self.c.get('/unifast/tes-applications/')
        body = r.content.decode()
        self.assertNotIn('Generate report', body)
        # Head of the TES Applications card, above its own table.
        self.assertIn('report-gen-head', body)
        self.assertIn('data-preview-open="annex1Preview"', body)

    def test_applications_page_narrows_to_the_chosen_year(self):
        self._applicant('Cruz', 'S1', school_year='2026-2027')
        self._applicant('Reyes', 'S2', school_year='2025-2026')

        r = self.c.get('/unifast/tes-applications/?sy=2025-2026')
        self.assertEqual(r.context['total'], 1)
        self.assertEqual([a['last_name'] for a in r.context['annex1_rows']], ['Reyes'])
        self.assertContains(r, 'Reyes')
        self.assertNotContains(r, 'S1')

    def test_download_is_an_xlsx_named_for_the_year(self):
        self._applicant('Cruz', 'S1')
        r = self.c.get('/unifast/tes-applications/report/?sy=2026-2027')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        self.assertIn('TES_Applicants_2026_2027.xlsx', r['Content-Disposition'])

        ws = openpyxl.load_workbook(BytesIO(r.content))[annex1_report.SHEET]
        self.assertEqual(ws.cell(row=annex1_report.FIRST_ROW, column=6).value, 'Cruz')

    def test_download_without_a_year_says_so_in_the_filename(self):
        self._applicant('Cruz', 'S1')
        r = self.c.get('/unifast/tes-applications/report/')
        self.assertIn('TES_Applicants_all_years.xlsx', r['Content-Disposition'])

    def test_report_is_the_unifast_office_only(self):
        student = User.objects.create_user(
            username='s@bipsu.edu.ph', email='s@bipsu.edu.ph', password='pw',
            role='student')
        StudentProfile.objects.create(user=student, student_id='X1')
        c = Client()
        self.assertTrue(c.login(email='s@bipsu.edu.ph', password='pw'))
        r = c.get('/unifast/tes-applications/report/')
        self.assertNotEqual(r.status_code, 200)

    # ── The year picker on the Reports page ─────────────────────────────────

    def test_reports_page_offers_the_year_picker(self):
        r = self.c.get('/unifast/reports/')
        self.assertContains(r, 'All school years')
        self.assertContains(r, 'name="sy"')
        self.assertEqual(r.context['school_years'], ['2026-2027'])

    def test_reports_page_scopes_the_grantee_list_to_the_year(self):
        self._applicant('Cruz', 'S1', school_year='2026-2027', status='Approved')
        self._applicant('Reyes', 'S2', school_year='2025-2026', status='Approved')

        r = self.c.get('/unifast/reports/?sy=2025-2026')
        self.assertEqual(r.context['tes_total'], 1)
        self.assertEqual(r.context['ay'], '2025-2026')
        self.assertEqual([row['last_name'] for row in r.context['rows']], ['Reyes'])

    def test_reports_page_defaults_to_every_year(self):
        """Unchanged behaviour until a year is chosen."""
        self._applicant('Cruz', 'S1', school_year='2026-2027', status='Approved')
        self._applicant('Reyes', 'S2', school_year='2025-2026', status='Approved')

        r = self.c.get('/unifast/reports/')
        self.assertEqual(r.context['tes_total'], 2)

    def test_annex2_download_is_named_and_titled_for_the_chosen_year(self):
        self._applicant('Reyes', 'S2', school_year='2025-2026', status='Approved')
        r = self.c.get('/unifast/reports/download/tes/?sy=2025-2026')
        self.assertEqual(r.status_code, 200)
        self.assertIn('TES_Validation_Billing_2025_2026.xlsx', r['Content-Disposition'])

        ws = openpyxl.load_workbook(BytesIO(r.content))['Official List']
        self.assertIn('AY 2025-2026', ws['A3'].value)
        self.assertEqual(ws.cell(row=tes_report.OFFICIAL_LIST['first_row'],
                                 column=4).value, 'Reyes')

    def test_annex2_grantee_rows_still_default_to_every_year(self):
        self._applicant('Cruz', 'S1', school_year='2026-2027', status='Approved')
        self._applicant('Reyes', 'S2', school_year='2025-2026', status='Approved')
        self.assertEqual(len(tes_report.grantee_rows()), 2)
