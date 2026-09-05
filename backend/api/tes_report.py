"""TES validation & billing report — the CHED 'Annex 2 Continuing' workbook.

The office's own workbook is bundled at ``templates/xlsx/`` and used as the
template: it is opened, grantee rows are written into the pre-formatted data
ranges, and unused rows are cleared and hidden. No part of the layout,
instructions, formulas or signatory blocks is recreated in code, so the download
matches the form CHED expects exactly apart from the data.

Because the template's own formulas are preserved, filling the Form 2 amount
columns makes the totals, the 1% management fee and the whole Form 1 billing
statement compute themselves when the file is opened.
"""
import os

TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'templates', 'xlsx', 'tes_validation_billing_template.xlsx',
)

# Pre-formatted data ranges. Past these rows the workbook holds its totals,
# undertaking text and signatory blocks, which must survive untouched.
OFFICIAL_LIST = {'sheet': 'Official List', 'first_row': 20, 'last_row': 1110}
BILLING_FORM2 = {'sheet': 'Annex 2-Form 2', 'first_row': 31, 'last_row': 1100}

# The amount columns are filled only from what the office has entered on the
# Billing tab for that school year — see models.TESBilling. With no row for the
# year they still export blank, which is what this did for every year before the
# tab existed.
#
# The distinction that matters and has not changed: CHED sets what a grantee is
# paid. The SUC rate quoted in the Form 2 header is a reference, not this
# office's figure, and TES-3A top-ups are decided per case. A rate this system
# chose for itself would be a guess an officer then signed; a rate an officer
# typed after being told it by CHED is theirs.

# Column order of the 'Official List' sheet. Annex 2-Form 2 repeats these in
# columns B..P, keeping a 5-digit control number in column A and a per-student
# total in column Q.
OFFICIAL_LIST_HEADERS = [
    'SEQ', 'STUDENT ID NUMBER', 'AWARD NUMBER', 'LAST NAME', 'FIRST NAME',
    'EXT NAME', 'MIDDLE NAME', 'SEX (F/M)', 'BIRTHDATE',
    'COURSE/ PROGRAM ENROLLED', 'YEAR LEVEL', 'CONTACT NUMBER', 'BATCH',
    'TES AMOUNT', 'TES-3A (PWD) AMOUNT', 'VALIDATION REMARKS',
]
_NCOLS = len(OFFICIAL_LIST_HEADERS)


def _is_pwd(tes):
    value = (tes.disability_type or '').strip().lower()
    return bool(value) and value not in ('n/a', 'na', 'none', 'not applicable')


def school_year_options():
    """The school years a TES report can be generated for, newest first.

    Every year TES applications were actually submitted in, plus the active
    term from SystemSettings — so the year the office is working in is always
    on the list even before the first application of it comes in. Shared by
    both TES reports; :mod:`api.annex1_report` reads the same list.
    """
    from .models import SystemSettings, TESApplication

    years = set(
        TESApplication.objects
        .exclude(school_year='')
        .values_list('school_year', flat=True)
    )
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active = SystemSettings.parse_label(settings_obj.academic_year)['sy']
    if active:
        years.add(active)
    return sorted(years, reverse=True)


def grantee_rows(batch='', school_year=''):
    """Approved TES grantees in the column order the CHED form expects.

    Sourced from the TES applications this portal reviews. The award number is
    read directly from TESApplication.award_number (set by UniFAST staff during
    the review decision). ``school_year`` is the expanded '2026-2027' form;
    blank means every year, which is what this returned before the office could
    pick one.
    """
    from .models import TESApplication, TESBilling

    rows = []
    grantees = (
        TESApplication.objects.filter(status='Approved')
        .select_related('student__user')
        .order_by('student__user__last_name', 'student__user__first_name')
    )
    if school_year:
        grantees = grantees.filter(school_year=school_year)

    # None until the office enters one on the Billing tab, and None means the
    # amount columns go out blank — the behaviour every year had before.
    billing = TESBilling.for_year(school_year)
    for seq, tes in enumerate(grantees, 1):
        p = tes.student
        u = p.user
        sex = (p.gender or '').strip().upper()[:1]
        pwd = _is_pwd(tes)
        # The top-up is only ever on the row it belongs to. Billing every
        # grantee for a disability allowance is the error worth being careful
        # about here — it is money, and CHEDRO reconciles it.
        tes_3a = billing.tes_3a_amount if (billing and pwd) else None
        rows.append({
            'seq': seq,
            # Which application this row is. Not a report column — the
            # liquidation page keys its per-grantee payments on it, and
            # re-deriving that from a second query ordered the same way is a
            # match waiting to slip. row_values() names its 16 cells, so this
            # never reaches the workbook.
            'app_id': tes.id,
            'student_no': p.student_id or '',
            'award_no': tes.award_number or '',
            'last_name': u.last_name or '',
            'first_name': u.first_name or '',
            'ext_name': '',
            'middle_name': tes.student.middle_name or '',
            'sex': sex if sex in ('F', 'M') else '',
            'birthdate': tes.birthdate or p.date_of_birth or None,
            'course': tes.complete_program or p.course or '',
            'year_level': p.year_level or '',
            'contact': tes.contact_number or p.contact_number or '',
            'batch': batch,
            # From the Billing tab, or blank when nothing is set for the year.
            'tes_amount': billing.tes_amount if billing else None,
            'tes_3a_amount': tes_3a,
            'is_pwd': pwd,
            'remarks': 'Enrolled',
        })
    return rows


def row_values(row):
    """The 16 'Official List' cell values, left to right."""
    return [
        row['seq'], row['student_no'], row['award_no'], row['last_name'],
        row['first_name'], row['ext_name'], row['middle_name'], row['sex'],
        row['birthdate'], row['course'], row['year_level'], row['contact'],
        row['batch'], row['tes_amount'], row['tes_3a_amount'], row['remarks'],
    ]


def _fill(ws, spec, rows, layout):
    """Write rows into a pre-formatted range; clear and hide what is left over.

    Unused rows are hidden rather than deleted so the merged cells, totals
    formulas and signatory blocks below the table keep their row positions —
    and so the totals keep summing the range they were written against.
    """
    first, last = spec['first_row'], spec['last_row']
    capacity = last - first + 1
    written = 0

    for i, row in enumerate(rows[:capacity]):
        r = first + i
        values = row_values(row)
        if layout == 'official':
            for j, value in enumerate(values):
                ws.cell(row=r, column=1 + j).value = value
        else:
            ws.cell(row=r, column=1).value = f'{i + 1:05d}'
            for j, value in enumerate(values[1:]):      # column A holds the control no.
                ws.cell(row=r, column=2 + j).value = value
            ws.cell(row=r, column=17).value = (
                (row['tes_amount'] or 0) + (row['tes_3a_amount'] or 0)
            )
        ws.row_dimensions[r].hidden = False
        written += 1

    ncols = _NCOLS if layout == 'official' else _NCOLS + 1
    for r in range(first + written, last + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).value = None
        ws.row_dimensions[r].hidden = True

    return written


def liquidation_rows(school_year):
    """Every approved grantee for a year, beside whatever was paid to them.

    Built from the same :func:`grantee_rows` the billing and the CHED workbook
    use, so the two pages cannot come to disagree about who is a grantee — the
    liquidation accounts for exactly the people the office billed for.

    A grantee with no :class:`~api.models.TESDisbursement` row yet is returned
    with ``disbursement`` None. That is deliberately not the same as zero: it
    means nobody has said anything about this person's share, and the page
    counts them as unaccounted for rather than quietly as unpaid.
    """
    from .models import TESDisbursement, TESLiquidation

    rows = grantee_rows(school_year=school_year)
    liquidation = TESLiquidation.for_year(school_year)

    paid = {}
    if liquidation:
        paid = {
            d.tes_application_id: d
            for d in TESDisbursement.objects.filter(liquidation=liquidation)
        }

    out = []
    for row in rows:
        entitled = (row['tes_amount'] or 0) + (row['tes_3a_amount'] or 0)
        out.append({
            **row,
            # What the office billed CHED for this grantee, shown beside what
            # was paid so the two can be compared without arithmetic. None when
            # no rate is set for the year — see TESBilling.
            'entitled': entitled or None,
            'disbursement': paid.get(row['app_id']),
        })
    return out


def liquidation_summary(school_year):
    """What was received, what went out, and what is still to be accounted for.

    The reconciliation a liquidation report exists to state:

        received   what CHEDRO remitted, per the credit advice
        released   what the cashier actually paid grantees
        returned   what has already gone back to CHED
        balance    received - released - returned, still held by the office

    ``balance`` is only meaningful once a remittance has been recorded, so it is
    None until then rather than a figure equal to the whole payout — a page that
    reports 'nothing received, everything owed back' on a term nobody has
    touched is worse than one that says the remittance is not in yet.

    Unclaimed money is not subtracted anywhere. It is still the office's to
    hold, so it sits inside ``balance`` by construction, and is counted
    separately because it is the usual reason a balance is not zero.
    """
    from decimal import Decimal

    from .constants import TES_RELEASED
    from .models import TESLiquidation

    rows = liquidation_rows(school_year)
    liquidation = TESLiquidation.for_year(school_year)
    billing = billing_summary(school_year)

    released = Decimal('0')
    returned = Decimal('0')
    released_count = unclaimed_count = returned_count = unaccounted = 0

    for row in rows:
        d = row['disbursement']
        if d is None:
            unaccounted += 1
            continue
        amount = d.amount_released or Decimal('0')
        if d.status == TES_RELEASED:
            released += amount
            released_count += 1
        elif d.status == 'Returned':
            returned += amount
            returned_count += 1
        else:
            unclaimed_count += 1

    received = liquidation.funds_received if liquidation else None
    balance = None if received is None else received - released - returned

    return {
        'liquidation': liquidation,
        'rows': rows,
        'grantees': len(rows),
        'billed': billing['total'] if billing['ready'] else None,
        'entitlement': billing['rate'],
        'received': received,
        'released': released,
        'returned': returned,
        'balance': balance,
        'released_count': released_count,
        'unclaimed_count': unclaimed_count,
        'returned_count': returned_count,
        'unaccounted': unaccounted,
        # Every grantee has been spoken for and nothing is left in hand. The
        # only state where the report can be signed off without a note
        # explaining the difference.
        'settled': bool(rows) and not unaccounted and balance == 0,
        # Paying out more than arrived is not an unusual balance, it is a
        # mistake somewhere — a mistyped amount, or a remittance recorded short.
        # Worth saying out loud rather than printing as a negative number.
        'over_released': balance is not None and balance < 0,
    }


def build_workbook(academic_year, semester, batch='', school_year=''):
    """Return ``(BytesIO, written, overflow)`` for the filled CHED workbook.

    ``academic_year`` is what gets printed in the sheet headers; ``school_year``
    is what the grantee list is filtered on. They are the same value whenever
    the office picks a year, and are kept apart only so a caller can still print
    a header over an unfiltered list, as this did before the picker existed.
    """
    import openpyxl
    from io import BytesIO

    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f'The TES workbook template is missing at {TEMPLATE_PATH}. '
            'Restore it from the office copy before generating this report.'
        )

    all_rows = grantee_rows(batch=batch, school_year=school_year)
    # Both sheets must list the same grantees, so the smaller pre-formatted
    # range is the real capacity. Anything past it is reported, not silently
    # dropped — the office would need extra lines inserted in the template.
    capacity = min(
        OFFICIAL_LIST['last_row'] - OFFICIAL_LIST['first_row'] + 1,
        BILLING_FORM2['last_row'] - BILLING_FORM2['first_row'] + 1,
    )
    rows = all_rows[:capacity]
    overflow = len(all_rows) - len(rows)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)

    official = wb[OFFICIAL_LIST['sheet']]
    official['A3'] = f'TES Batch {batch or "On-going"}, {semester}, AY {academic_year}'
    written = _fill(official, OFFICIAL_LIST, rows, 'official')

    form2 = wb[BILLING_FORM2['sheet']]
    form2['J28'] = f'{semester}, AY {academic_year}'
    _fill(form2, BILLING_FORM2, rows, 'form2')

    # Grantee headcount on the billing statement; its money figures already
    # pull straight from the Form 2 totals.
    form1 = wb['Annex 2-Form 1']
    form1['Q24'] = written

    # The two things on Form 1 that are the office's to state rather than
    # anything the sheet can compute. Left as the template's own placeholders
    # when unset, so a half-filled statement reads as one.
    from .models import TESBilling
    billing = TESBilling.for_year(school_year)
    if billing:
        if billing.reference_no:
            form1['T10'] = billing.reference_no
        if billing.statement_date:
            form1['T11'] = billing.statement_date.strftime('%B %d, %Y')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, written, overflow


def billing_summary(school_year, batch=''):
    """The billing statement's own arithmetic, for the page that sets the rate.

    Computed here rather than read back from the workbook because the workbook's
    totals are Excel formulas — openpyxl would hand back the formula string, not
    a number. These mirror Form 2's totals row and Form 1's statement:

        N1101  TES benefits      = rate x grantees
        O1101  TES-3A            = top-up x grantees with a disability
        Q1102  management fee    = 1% of the TES benefits, not of the total
        Q1103  TOTAL billed      = benefits + top-ups + fee

    The 1% is on ``N1101`` alone — the template's own formula, and the reason
    this is not simply one percent of the grand total.
    """
    from decimal import Decimal

    from .models import TESBilling

    rows = grantee_rows(batch=batch, school_year=school_year)
    billing = TESBilling.for_year(school_year)
    grantees = len(rows)
    pwd = sum(1 for r in rows if r['is_pwd'])

    rate = billing.tes_amount if billing else None
    top_up = billing.tes_3a_amount if billing else None

    benefits = (rate or Decimal('0')) * grantees
    tes_3a = (top_up or Decimal('0')) * pwd
    fee = (benefits * Decimal('0.01')).quantize(Decimal('0.01'))

    return {
        'billing': billing,
        'grantees': grantees,
        'pwd': pwd,
        'rate': rate,
        'top_up': top_up,
        'benefits': benefits,
        'tes_3a': tes_3a,
        'subtotal': benefits + tes_3a,
        'fee': fee,
        'total': benefits + tes_3a + fee,
        # Nothing is billable until somebody has entered a rate.
        'ready': bool(billing and billing.is_set and grantees),
    }
