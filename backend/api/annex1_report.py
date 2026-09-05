"""TES applicants report — a plain table of everyone who applied.

**The bundled Annex 1 workbook is a guide, not the output.** It is the office's
copy of CHED's form, and what this module takes from it is two things: the list
of columns a TES applicant record has to carry, and the lookup sheets that fill
the apply form's Programme and Disability dropdowns. The report itself is
generated here — one header row and one row per applicant — rather than being
that workbook with rows written into it.

That is the whole difference from :mod:`api.tes_report`, which really does fill
CHED's Annex 2 in place, because Annex 2 is submitted to CHED as-is and carries
formulas and signatory blocks that must survive. Nothing here is submitted on a
CHED form, so nothing here needs the form.

Where Annex 2 lists the grantees the office bills for, this one lists everybody
who *applied*: a decision has not necessarily been made yet, so the default is
every application in the chosen school year rather than the approved ones only.
"""
import os

# The guide. Still read — see registry_programs() and disability_types() — but
# never written to and never handed to anybody as the report.
TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'templates', 'xlsx', 'tes_annex1_applicants_template.xlsm',
)

SHEET = 'TES Applicants'

# The columns, in the order CHED's form lists them. Taken from the guide, kept
# flat rather than nested the way its merged header draws them, because a simple
# table is what this produces and what the on-screen preview shows.
ANNEX1_HEADERS = [
    'SEQ', 'STUDENT ID', 'LRN', 'PHILSYS ID', '4PS ID',
    'LAST NAME', 'GIVEN NAME', 'EXT. NAME', 'MIDDLE NAME',
    'SEX', 'BIRTHDATE', 'COMPLETE PROGRAM NAME', 'YEAR LEVEL',
    "FATHER'S LAST NAME", "FATHER'S GIVEN NAME", "FATHER'S MIDDLE NAME",
    "MOTHER'S LAST NAME", "MOTHER'S GIVEN NAME", "MOTHER'S MIDDLE NAME",
    'STREET & BARANGAY', 'CITY/MUNICIPALITY', 'PROVINCE', 'REGION', 'ZIPCODE',
    'CONTACT NUMBER', 'EMAIL ADDRESS', 'DISABILITY TYPE',
    'SOLO PARENT DEPENDENT', 'FIRST-GEN COLLEGE', 'INDIGENOUS PEOPLE GROUP',
]
_NCOLS = len(ANNEX1_HEADERS)

# What the form wants in the two columns that are never left blank. The
# Disability_List and the IP instruction both spell 'not applicable' as NO.
NOT_APPLICABLE = 'NO'

# PhilSys and 4Ps ID numbers are optional on CHED's form and are asked for the
# same way on the apply form. A student who has neither leaves them blank and
# the columns export empty. They are never derived from the 4Ps flag on
# TESEligibility: that is a yes/no, and this column wants a household's ID.


# The workbook's own lookup sheets, read once and kept. The apply form offers
# exactly these, so a student cannot enter a programme or a disability the sheet
# would then reject — and when the office drops in a newer template, the form
# follows it without anybody editing a list in code.
LOOKUP_SHEETS = {
    'programs': 'Registry_Courses',
    'disabilities': 'Disability_List',
}

# What a student picks when their case is not on CHED's list. Kept out of the
# lookup sheets because it is this system's word, not CHED's — the value that
# reaches the workbook is whatever they then type.
OTHER = 'Other'

_LOOKUP_CACHE = {}


def _lookup(key):
    """The values of one lookup sheet, minus its heading row.

    An empty list when the template is missing rather than an exception: the
    apply form has to render for a student either way, and the office is
    already told the template is gone on every screen that needs it.
    """
    if key in _LOOKUP_CACHE:
        return _LOOKUP_CACHE[key]

    values = []
    if os.path.exists(TEMPLATE_PATH):
        import openpyxl
        wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
        try:
            sheet = wb[LOOKUP_SHEETS[key]]
            values = [str(cell.value).strip()
                      for (cell,) in sheet.iter_rows(min_row=2, max_col=1)
                      if cell.value is not None and str(cell.value).strip()]
        finally:
            wb.close()

    _LOOKUP_CACHE[key] = values
    return values


def registry_programs():
    """The programme names CHED's registry holds, for the apply form's dropdown.

    The reason this is not ``BIPSU_COURSES``: that list is the abbreviations the
    university uses internally — 'BSCS', 'BSEd - English' — and CHED asks for
    the registered name, 'BACHELOR OF SCIENCE IN COMPUTER SCIENCE'. Typing the
    abbreviation into this form is what put 'BSHM' in an Annex 1 that CHED reads
    against its own registry.
    """
    return _lookup('programs')


def disability_types():
    """The disability values the form offers, in the sheet's own order.

    'NO' leads it, which is how this form spells 'not applicable' — see
    :func:`_disability`. ``OTHER`` is appended for a condition CHED's list does
    not name; the student types that one out.
    """
    return _lookup('disabilities') + [OTHER]


def _sex_code(gender):
    """0 = Male, 1 = Female — the codes the Sex_Code sheet validates against.

    Anything else defaults to 0, which is what the General Instructions tab
    says a blank does anyway.
    """
    return 1 if (gender or '').strip().lower().startswith('f') else 0


def _contact(number):
    """A contact number in the shape the form asks for.

    CHED wants a 10-digit mobile starting with 9 (or an 8-digit landline), but
    Philippine mobiles are written and stored as 11 digits beginning '09'. Only
    that leading zero is dropped; a number of any other length is passed through
    as typed rather than guessed at.
    """
    digits = ''.join(c for c in (number or '') if c.isdigit())
    if len(digits) == 11 and digits.startswith('0'):
        return digits[1:]
    return digits or (number or '').strip()


def _disability(value):
    value = (value or '').strip()
    if not value or value.lower() in ('n/a', 'na', 'none', 'not applicable', 'no'):
        return NOT_APPLICABLE
    return value


def _ip_group(value):
    value = (value or '').strip()
    if not value or value.lower() in ('n/a', 'na', 'none', 'not applicable', 'no'):
        return NOT_APPLICABLE
    return value


def applicant_rows(school_year='', status=''):
    """TES applicants in the column order the Annex 1 form expects.

    ``school_year`` is the expanded '2026-2027' form; blank means every year.
    ``status`` filters to one review status; blank means all of them, which is
    the point of this list — an applicant belongs on it before the decision.
    """
    from .models import STUDENT_DETAILS, TESApplication

    # Every detail row is joined in: this report reads personal, enrolment and
    # family fields on each applicant, and they live on separate tables.
    applications = (
        TESApplication.objects
        .select_related('student__user', *STUDENT_DETAILS)
        .order_by('student__user__last_name', 'student__user__first_name')
    )
    if school_year:
        applications = applications.filter(school_year=school_year)
    if status:
        applications = applications.filter(status=status)

    rows = []
    for seq, tes in enumerate(applications, 1):
        p = tes.student
        u = p.user
        rows.append({
            'seq': seq,
            'student_no': p.student_id or '',
            'lrn': tes.lrn or p.learner_ref_no or '',
            'philsys_id': tes.philsys_id or '',
            'four_ps_id': tes.four_ps_id or '',
            'last_name': u.last_name or '',
            'first_name': u.first_name or '',
            'ext_name': p.suffix or '',
            'middle_name': p.middle_name or '',
            'sex': _sex_code(p.gender),
            'birthdate': tes.birthdate or p.date_of_birth or None,
            'course': tes.complete_program or p.course or '',
            'year_level': p.year_level or '',
            'father_last_name': p.father_last_name or '',
            'father_first_name': p.father_first_name or '',
            'father_middle_name': p.father_middle_name or '',
            'mother_last_name': p.mother_last_name or '',
            'mother_first_name': p.mother_first_name or '',
            'mother_middle_name': p.mother_middle_name or '',
            'street_barangay': tes.street_barangay or '',
            'city_municipality': tes.city_municipality or '',
            'province': tes.province or '',
            'region': tes.region or '',
            'zip_code': tes.zip_code or '',
            'contact': _contact(tes.contact_number or p.contact_number),
            'email': tes.email_address or u.email or '',
            'disability': _disability(tes.disability_type),
            'solo_parent': 1 if tes.is_solo_parent_dependent else 0,
            'first_gen': 1 if tes.is_first_gen_college else 0,
            'ip_group': _ip_group(tes.indigenous_people_group),
            # Not a form column — the preview shows it so an officer can see at
            # a glance which of these applicants have been decided.
            'status': tes.status,
        })
    return rows


def row_values(row):
    """The 30 'Annex 1' cell values, left to right (columns A..AD)."""
    return [
        row['seq'], row['student_no'], row['lrn'], row['philsys_id'],
        row['four_ps_id'], row['last_name'], row['first_name'], row['ext_name'],
        row['middle_name'], row['sex'], row['birthdate'], row['course'],
        row['year_level'], row['father_last_name'], row['father_first_name'],
        row['father_middle_name'], row['mother_last_name'],
        row['mother_first_name'], row['mother_middle_name'],
        row['street_barangay'], row['city_municipality'], row['province'],
        row['region'], row['zip_code'], row['contact'], row['email'],
        row['disability'], row['solo_parent'], row['first_gen'], row['ip_group'],
    ]


# Roughly how wide each column needs to be, in the header order above. Set once
# here rather than measured per run, so two reports of the same list look the
# same however long one person's address happens to be.
COLUMN_WIDTHS = [
    5, 14, 16, 16, 14,          # seq, ids
    16, 16, 9, 16,              # student name
    5, 12, 40, 6,               # sex, birthdate, programme, year
    16, 16, 16,                 # father
    16, 16, 16,                 # mother
    26, 18, 14, 12, 8,          # address
    15, 26, 22,                 # contact, email, disability
    10, 10, 22,                 # flags, IP group
]

# The header row starts here. Two rows above it say what the table is and which
# year it covers — a sheet handed to somebody has to answer that on its face.
HEADER_ROW = 3
FIRST_ROW = HEADER_ROW + 1


def build_workbook(school_year, status=''):
    """Return ``(BytesIO, written)`` — the applicant list as a plain table.

    One title line, one header row, one row per applicant. Not CHED's workbook
    with data poured into it: that file is the guide this takes its columns
    from, and nothing here is submitted on a CHED form.
    """
    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows = applicant_rows(school_year=school_year, status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = ws.cell(row=1, column=1)
    title.value = (
        'LIST OF TES APPLICANTS — Biliran Province State University'
        + (f' — Academic Year {school_year}' if school_year else '')
        + (f' — {status}' if status else '')
    )
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOLS)

    for column, heading in enumerate(ANNEX1_HEADERS, 1):
        cell = ws.cell(row=HEADER_ROW, column=column, value=heading)
        cell.font = Font(bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = border
    ws.row_dimensions[HEADER_ROW].height = 30

    for offset, row in enumerate(rows):
        for column, value in enumerate(row_values(row), 1):
            cell = ws.cell(row=FIRST_ROW + offset, column=column, value=value)
            cell.font = Font(size=9)
            cell.border = border
            # The one column openpyxl would otherwise write as a datetime and
            # Excel would show in whatever the reader's locale prefers.
            if ANNEX1_HEADERS[column - 1] == 'BIRTHDATE':
                cell.number_format = 'yyyy-mm-dd'

    for column, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(column)].width = width

    # Headings stay put while the office scrolls a long list, and every column
    # gets a filter dropdown — the two things anyone does with a table this wide.
    ws.freeze_panes = ws.cell(row=FIRST_ROW, column=1)
    if rows:
        ws.auto_filter.ref = (
            f'A{HEADER_ROW}:'
            f'{get_column_letter(_NCOLS)}{FIRST_ROW + len(rows) - 1}'
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(rows)
