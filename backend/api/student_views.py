from django.shortcuts import render, redirect
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.views.decorators.clickjacking import xframe_options_exempt
from . import scholar_columns
from .models import STAFF_APPLICATION_DETAILS, STUDENT_DETAILS, StudentProfile, Scholarship, Application, Notification, Announcement, User, AffirmativeStaffApplication, AcademicRenewal, ScholarshipLinkRequest, TESApplication, BIPSU_SCHOOLS, BIPSU_COURSES, split_ched
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.core.files.base import ContentFile
from django.db import transaction
from . import notify
from django.http import HttpResponse
from io import BytesIO
from decimal import Decimal, InvalidOperation


# — Landing ——————————————————————————————————

def landing_view(request):
    qs = Scholarship.objects.filter(is_active=True).order_by('type')
    return render(request, 'landing.html', {
        'scholarships': qs,
        'internal': qs.filter(group='internal'),
        'external': qs.filter(group='external'),
        'institutional': qs.filter(group='institutional'),
    })


# — Academic auth ———————————————————————————————————

PORTAL_FOR_ROLE = {
    'student': '/student/applications/',
    'nsu_staff': '/nsu-staff/',
    'vpsea': '/vpsea/',
    'unifast': '/unifast/',
    'super': '/super/',
}


def _portal_for(user):
    """Where this account lands after signing in."""
    return PORTAL_FOR_ROLE.get(user.role, '/')


def login_view(request):
    # Already signed in — including someone the middleware just released, who
    # would otherwise be staring at a login form they no longer need.
    if request.user.is_authenticated:
        return redirect(_portal_for(request.user))

    if request.method == 'POST':
        email = (request.POST.get('email') or '').strip()
        password = request.POST.get('password') or ''
        user = authenticate(request, username=email, password=password)
        if user and not user.can_sign_in:
            # Only reachable with the right password, so this tells the person
            # about their own account and nobody about anyone else's.
            return render(request, 'login.html', {
                'verification_status': user.verification_status,
                'verification_note': user.verification_note,
            })
        if user:
            login(request, user)
            return redirect(_portal_for(user))
        # 'Invalid credentials' told nobody anything. Someone who mistyped their
        # address and someone who forgot their password got the same sentence,
        # and both retyped the same wrong thing. Naming which half is wrong is
        # what the office was fielding phone calls about.
        #
        # It does mean the form will confirm whether an address is registered.
        # The registration form already does — it refuses a duplicate email —
        # so the page is not giving away anything the site did not already say.
        return render(request, 'login.html', _sign_in_error(email, password))
    return render(request, 'login.html')


def _sign_in_error(email, password):
    """Work out which part of the sign-in the person got wrong, and say so.

    Returns the template context, keeping ``email`` filled in so a wrong
    password does not cost them the address they typed as well.
    """
    ctx = {'email': email}

    if not email:
        ctx['error'] = 'Enter the email address you registered with.'
        return ctx
    if not password:
        ctx['error'] = 'Enter your password.'
        return ctx

    account = User.objects.filter(email__iexact=email).first()
    if account is None:
        ctx['error'] = (
            f'No account is registered under {email}. Check the address for a '
            'typo, or register first if you have not yet.'
        )
        ctx['error_action'] = 'register'
        return ctx

    if not account.is_active:
        ctx['error'] = (
            'That account has been deactivated. Contact the SDSO office to have '
            'it reopened.'
        )
        return ctx

    ctx['error'] = (
        f'The password does not match the account for {email}. Check your '
        'capitals — passwords are case-sensitive.'
    )
    return ctx


def logout_view(request):
    logout(request)
    return redirect('/')


def _await_verification(request, user):
    """Hand a freshly registered account off to the SDSO, without signing it in.

    The email goes in the session rather than the URL so it is not left in
    browser history or a shared link.
    """
    from .models import ActivityLog
    ActivityLog.objects.create(
        user=user,
        action=f'Account registered — awaiting SDSO verification ({user.get_role_display()})',
    )
    from django.utils import timezone
    from .middleware import PENDING_EMAIL, PENDING_SINCE
    request.session[PENDING_EMAIL] = user.email
    request.session[PENDING_SINCE] = timezone.now().isoformat()
    return redirect('/register/received/')


def registration_received(request):
    """The waiting room. It refreshes itself, so nobody has to watch for the news.

    By the time this runs the middleware has already signed the visitor in if
    the SDSO released them, so an authenticated caller here means 'verified
    while you were waiting' — send them straight to their portal.
    """
    from .middleware import PENDING_EMAIL
    if request.user.is_authenticated:
        return redirect(_portal_for(request.user))

    email = request.session.get(PENDING_EMAIL, '')
    account = User.objects.filter(email=email).first() if email else None
    rejected = account is not None and account.verification_status == 'rejected'
    return render(request, 'registration_received.html', {
        'email': email,
        'rejected': rejected,
        'note': account.verification_note if rejected else '',
        # Only keep reloading while there is genuinely something to wait for.
        'waiting': bool(account) and account.awaiting_verification,
        # The other half of the wait: whether they have opened the link sent to
        # the address they typed. Until they do, nobody knows it reaches them.
        'confirm_email': bool(account) and not account.email_verified,
        'confirmed': request.GET.get('confirmed') == '1',
        'resent': request.GET.get('resent') == '1',
        'confirm_error': request.GET.get('confirm_error', ''),
    })


def verify_email(request, token):
    """Open the link mailed to a registrant, and mark the address confirmed.

    Nothing here signs anyone in or releases an account: the SDSO's decision is
    still the gate. All this records is that somebody could read mail at the
    address, which is what the office needs to know before writing to it.
    """
    from . import email_verify
    from .middleware import PENDING_EMAIL
    from urllib.parse import quote

    account, reason = email_verify.read_token(token)
    if account is None:
        message = {
            'expired': 'That confirmation link has expired. Registrations are '
                       'confirmed within three days — ask for a new link below.',
            'stale': 'That link was sent to a different address than the one on '
                     'the account now. Ask for a new link below.',
        }.get(reason, 'That confirmation link is not valid. Check that you '
                      'copied the whole of it, or ask for a new one below.')
        return redirect(f'/register/received/?confirm_error={quote(message)}')

    account.mark_email_verified()
    # Opening the link on the same browser they registered in puts them back in
    # the waiting room where the middleware can release them; on a different
    # one the session is empty and the page says to sign in instead.
    if not request.session.get(PENDING_EMAIL):
        request.session[PENDING_EMAIL] = account.email
    return redirect('/register/received/?confirmed=1')


def resend_confirmation(request):
    """Send the confirmation link again, to the address this browser registered.

    Only the address held in the session — this must not become a way to make
    the site email an arbitrary address on demand.
    """
    from . import email_verify
    from .middleware import PENDING_EMAIL
    from urllib.parse import quote

    # POST only: sending mail is not something a link preview or a prefetching
    # browser should be able to set off by following a URL.
    if request.method != 'POST':
        return redirect('/register/received/')

    email = request.session.get(PENDING_EMAIL, '')
    account = User.objects.filter(email=email).first() if email else None
    if account is None:
        return redirect('/register/received/?confirm_error=' + quote(
            'There is no registration on this browser to send a link for. '
            'Sign in with the email and password you registered with.'))
    if account.email_verified:
        return redirect('/register/received/?confirmed=1')

    email_verify.send_confirmation(account, request)
    return redirect('/register/received/?resent=1')


def _release_rejected_registration(email, student_id):
    """Clear a rejected registration out of the way of a fresh one.

    Only ever a rejected one. A pending registration still blocks — it is
    waiting on the office, not finished with — and an approved one is somebody's
    live account.

    The row goes rather than being rewritten in place: a re-registration can
    change the account type, which would leave a StaffProfile hanging off what is
    now a student. Deleting takes the profile with it and the new registration
    builds whatever it needs.

    What survives is the ActivityLog line. ``ActivityLog.user`` is SET_NULL, so
    the entry outlives the account it names and the office can still see that
    this address has been through here before — which is the part worth keeping
    when the rejection was for something other than a typo.
    """
    from django.db.models import Q

    from .models import ActivityLog

    claimed = Q(email=email)
    if student_id:
        claimed |= Q(profile__student_id=student_id)

    rejected = User.objects.filter(claimed, verification_status='rejected').distinct()

    for account in rejected:
        held = account.email
        profile = getattr(account, 'profile', None)
        if profile and profile.student_id:
            held += f' / {profile.student_id}'
        reason = account.verification_note or 'no reason recorded'
        ActivityLog.objects.create(
            user=None,
            action=(f'Rejected registration replaced by a new one — {held} '
                    f'— original reason: {reason}'),
        )
        account.delete()


def _register_context(post=None):
    """Everything the signup form needs to draw itself.

    The form asks for the whole student record now — the same groups My Profile
    shows — so it needs the same lists that page does. Students used to type
    their course free-hand, which is why courses on file read 'BSIT',
    'Batchelor of Science in Computer Science ' and so on, and why their school
    could not be worked out from them.
    """
    import json
    from .constants import CIVIL_STATUSES, GENDERS
    from .models import CHED_TIER_CHOICES, SCHOLARSHIP_TYPE_CHOICES, SystemSettings
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    return {
        'bipsu_schools': BIPSU_SCHOOLS,
        'bipsu_courses_json': json.dumps(BIPSU_COURSES),
        'civil_statuses': CIVIL_STATUSES,
        'genders': GENDERS,
        'scholarship_types': SCHOLARSHIP_TYPE_CHOICES,
        'ched_tiers': CHED_TIER_CHOICES,
        'max_upload_mb': settings_obj.max_file_size_mb or 5,
        **_disability_fields(
            (post or {}).get('disability_type') if post is not None else None,
            (post or {}).get('disability_type_other', ''),
            ''),
    }


# Everything the registration form collects onto the profile besides the
# identity columns handled by hand below. Grouped the way the form groups them,
# which is the way My Profile groups them, which is the way the detail tables
# are grouped — one shape all the way down.
def _registration_profile_fields(p, files, disability):
    """The StudentProfile kwargs a completed registration form describes."""
    from .constants import school_for_course
    course = p.get('course', '')
    fields = {
        # Enrolment
        'school': p.get('school', '').strip() or school_for_course(course),
        'course': course,
        'year_level': int(p.get('year_level', 1) or 1),
        # Personal. Collected here because nothing else does: the masterlist
        # exports carry a MIDDLE NAME and an M.I. column, and the office forms
        # only ever set the given and family names.
        'middle_name': p.get('middle_name', '').strip(),
        'suffix': p.get('suffix', '').strip(),
        'birth_place': p.get('birth_place', '').strip(),
        'civil_status': p.get('civil_status', '').strip(),
        'date_of_birth': p.get('date_of_birth') or None,
        'gender': p.get('gender', ''),
        'contact_number': p.get('contact_number', ''),
        'disability_type': disability,
        # Address
        'barangay': p.get('barangay', '').strip(),
        'municipality': p.get('municipality', '').strip(),
        'province': p.get('province', '').strip(),
        # Educational background
        'elementary': p.get('elementary', '').strip(),
        'highschool': p.get('highschool', '').strip(),
        'last_school': p.get('last_school', '').strip(),
        # Socio-economic
        'family_income': _decimal_or(p.get('family_income'), 0.0),
        'indigenous_group': p.get('indigenous_group', '').strip(),
        # Scholarship eligibility
        'shs_gpa': _decimal_or(p.get('shs_gpa'), None),
        'suc_exam_score': _decimal_or(p.get('suc_exam_score'), None),
        'suc_exam_total': _decimal_or(p.get('suc_exam_total'), None) or None,
        'is_tes_beneficiary': 'is_tes_beneficiary' in p,
        # TES eligibility. Unanswered stays unknown — see _tristate.
        'citizenship': p.get('citizenship', '').strip(),
        'household_size': _positive_int(p.get('household_size'), None),
        'year_first_enrolled': _positive_int(p.get('year_first_enrolled'), None),
        'is_listahanan_household': _tristate(p.get('is_listahanan_household'), None),
        'is_4ps_beneficiary': _tristate(p.get('is_4ps_beneficiary'), None),
        'has_previous_degree': _tristate(p.get('has_previous_degree'), None),
    }
    for name in ('shs_gpa_cert', 'suc_exam_cert'):
        if files.get(name):
            fields[name] = files[name]
    return fields


def _decimal_or(raw, fallback):
    """A number out of a form field, or `fallback` when it is blank or junk."""
    try:
        return float((raw or '').strip())
    except (TypeError, ValueError):
        return fallback


def _certificate_errors(files):
    """Type and size checks for the two certificates the form takes.

    Both are optional, so a missing one is not an error — but this is a public
    endpoint, and an upload nobody has looked at is not something to write to
    disk on the strength of the accept="..." attribute alone. Same rules as the
    proof document; the field validators on the model do not run on save().
    """
    from .models import SystemSettings
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    errors = []
    for field, label in (('shs_gpa_cert', 'SHS GPA Certificate'),
                         ('suc_exam_cert', 'SUC Exam Certificate')):
        upload = files.get(field)
        if upload:
            errors += [f'{label}: {problem}'
                       for problem in _validate_proof(upload, settings_obj)]
    return errors


def _declared_scholarship(p, files):
    """(ScholarshipLinkRequest kwargs, errors) for the Scholarship Data card.

    (None, []) when the box was not ticked: holding nothing yet is the ordinary
    case, not a mistake. Ticking it and then leaving the card empty is, because
    the office cannot verify a scholarship nobody named.
    """
    from .models import CHED_TIER_CHOICES, SCHOLARSHIP_TYPE_CHOICES, SystemSettings
    if 'has_scholarship' not in p:
        return None, []

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    stype = p.get('scholarship_type', '')
    # CHED is awarded at two tiers under a single programme and every masterlist
    # reports the two in separate blocks, so a CHED declaration has to say which
    # one. Other programmes have one tier; blank there.
    tier = p.get('award_tier', '') if stype == 'CHED' else ''
    proof = files.get('proof_document')

    errors = []
    if stype not in [t for t, _ in SCHOLARSHIP_TYPE_CHOICES]:
        errors.append('Say which scholarship you already hold, or clear the '
                      '"I already hold a scholarship" box.')
    elif stype == 'CHED' and tier not in [t for t, _ in CHED_TIER_CHOICES]:
        errors.append('Please choose whether your CHED award is Full Merit / Full Scholar '
                      'or Half Merit / Partial Scholar — your award letter says which.')
    errors += _validate_proof(proof, settings_obj)
    if errors:
        return None, errors

    return dict(
        scholarship_type=stype,
        proof_document=proof,
        award_number=p.get('award_number', '').strip(),
        award_tier=tier,
        notes=p.get('notes', ''),
        term_label=settings_obj.academic_year,
    ), []


def register_view(request):
    if request.method == 'POST':
        p = request.POST
        errors = []
        account_type = p.get('account_type', 'student')

        from . import email_verify

        if p.get('password') != p.get('confirm_password'):
            errors.append('Passwords do not match.')
        # Checked before uniqueness: 'already registered' is a confusing thing
        # to be told about an address that could never have been registered.
        address_problem = email_verify.address_error(p.get('email'))
        if address_problem:
            errors.append(address_problem)
        elif User.objects.filter(email=p.get('email')).exclude(
                verification_status='rejected').exists():
            errors.append('Email already registered.')

        # The scholarship a student already holds, declared here rather than on
        # a page of its own after the fact — see the Scholarship Data card. The
        # office reviews it as part of verifying the account, so the proof has
        # to arrive with the registration it belongs to.
        declared, disability = None, ''
        if account_type == 'student':
            if not p.get('student_id'):
                errors.append('Student ID is required.')
            elif StudentProfile.objects.filter(
                    student_id=p.get('student_id')).exclude(
                    user__verification_status='rejected').exists():
                errors.append('Student ID already registered.')

            disability, problem = _disability_answer(p)
            if problem:
                errors.append(problem)

            declared, link_errors = _declared_scholarship(p, request.FILES)
            errors.extend(link_errors)
            errors.extend(_certificate_errors(request.FILES))

        if errors:
            return render(request, 'register.html',
                          dict(_register_context(p), errors=errors, post=p))

        # A rejection is usually 'those details do not match our records', so the
        # answer to it is a corrected registration. Holding the address and the
        # student number hostage meant the one person who could fix the mistake
        # was the only one who could not: they could not re-register, and could
        # not edit the account they were locked out of either.
        _release_rejected_registration(p.get('email'), p.get('student_id'))

        # ── Create the Django user ──────────────────────────────────────────
        user = User.objects.create_user(
            username=p.get('email'),
            email=p.get('email'),
            password=p.get('password'),
            first_name=p.get('first_name', '').strip(),
            last_name=p.get('last_name', '').strip(),
            role=account_type,
            # Nobody signs in off the public form until the SDSO says so.
            verification_status='pending',
            # Nor is an address off the public form taken on trust — see
            # api/email_verify.py. The office's own accounts stay exempt.
            email_verified=False,
        )
        # Sent before the profile is built so a slow mail server delays the
        # confirmation, never the registration; it cannot fail the request.
        # The request is passed so the link is absolute even where SITE_URL
        # was never set — a relative path in an email is not a link.
        email_verify.send_confirmation(user, request)

        if account_type == 'student':
            # ── Student: create a StudentProfile ───────────────────────────
            profile = StudentProfile.objects.create(
                user=user,
                student_id=p.get('student_id'),
                **_registration_profile_fields(p, request.FILES, disability),
            )
            if declared:
                # Pending until the SDSO verifies it on the account queue.
                # Approving the account is what turns this into an award.
                ScholarshipLinkRequest.objects.create(student=profile, **declared)
            return _await_verification(request, user)

        else:
            # ── BiPSU Staff: the employment details go on the StaffProfile,
            #    which is the employee's own record. The application itself
            #    is created when they actually apply, not here.
            from .models import AffirmativeStaffApplication, StaffProfile
            # Picked from the BiPSU list, not typed — same dropdown the student
            # form and My Profile use. Under its own name because the student
            # block posts a 'school' of its own from the same form, and the last
            # value wins in request.POST.
            staff_school = p.get('staff_school', '').strip()
            StaffProfile.objects.create(
                user=user,
                middle_name=p.get('middle_name', '').strip(),
                suffix=p.get('suffix', '').strip(),
                contact_number=p.get('contact_number', '').strip(),
                date_of_birth=p.get('date_of_birth') or None,
                gender=p.get('gender', ''),
                employee_id=p.get('school_id', '').strip(),
                school=staff_school,
                department=p.get('department', '').strip(),
                position=p.get('position', '').strip(),
            )
            from .models import ActivityLog
            ActivityLog.objects.create(
                user=user,
                action=(
                    f"Staff account created — "
                    f"School ID: {p.get('school_id','—')} | "
                    f"School: {staff_school or '—'} | "
                    f"Department: {p.get('department','—')} | "
                    f"Position: {p.get('position','—')} | "
                    f"Contact: {p.get('contact_number','—')}"
                ),
            )
            return _await_verification(request, user)

    return render(request, 'register.html', dict(_register_context(), post={}))


# — Academic student pages ——————————————————————————

def _tristate(raw, current):
    """('yes' | 'no' | '') -> (True | False | None), keeping the current value on junk.

    A checkbox cannot express these fields. It only ever posts on or off, so an
    untouched box would record a confident "no" for a question nobody asked —
    and the TES recommender's whole design turns on telling 'confirmed no' apart
    from 'not yet collected'. Hence a three-option select, and hence this.
    """
    value = (raw or '').strip().casefold()
    if value == 'yes':
        return True
    if value == 'no':
        return False
    if value == 'unknown':
        return None
    return current


def _positive_int(raw, current):
    """A whole number above zero, or the value already on file."""
    text = (raw or '').strip()
    if text == '':
        return None
    if text.isdigit() and int(text) > 0:
        return int(text)
    return current


def _disability_answer(posted):
    """(value to store, error) for the Disability Type dropdown and its 'Other' box.

    The same pair the TES application uses: CHED's own Disability_List in a
    dropdown, plus one option this system adds for a condition their list does
    not name. 'Other' on its own says nothing, so it is refused rather than
    stored. 'NO' is how that list spells not applicable, and storing it is the
    student answering the question — not leaving it blank.
    """
    from . import annex1_report
    value = (posted.get('disability_type') or '').strip()
    if value != annex1_report.OTHER:
        return value, ''
    typed = (posted.get('disability_type_other') or '').strip()
    if not typed:
        return '', 'Name the disability you chose "Other" for.'
    return typed, ''


def _disability_fields(posted_value, posted_other, saved):
    """Everything a Disability Type dropdown needs to render.

    A stored value that is not on CHED's list is one somebody typed under
    'Other', so it comes back selected as 'Other' with the text beside it —
    otherwise re-opening the form would silently drop what they wrote. Passing
    `posted_value` (rather than None) redisplays what was just typed, which is
    what a form coming back with an error has to show.
    """
    from . import annex1_report
    options = annex1_report.disability_types()
    if posted_value is not None:
        value, other = posted_value, posted_other
    else:
        saved = (saved or '').strip()
        custom = saved if saved and saved not in options else ''
        value, other = (annex1_report.OTHER if custom else saved), custom
    return {
        'disabilities': options,
        'other_option': annex1_report.OTHER,
        'disability_value': value,
        'disability_other': other,
    }


def _scholarship_records(profile):
    """What the student holds, and what they declared that is still being checked.

    One list because a student thinks of it as one question — "what am I on?" —
    even though three tables answer it: the awards ledger, the TES application
    UniFAST decides on its own screen, and the scholarship declared at
    registration that nobody has verified yet. An approved declaration is left
    out: approving one writes the award, and it would otherwise appear twice.
    """
    if not profile:
        return []
    records = [{
        'name': app.scholarship.name,
        'type': app.scholarship.type,
        'status': 'Approved',
        'term': ' '.join(x for x in (app.school_year, app.semester) if x),
        'award_number': app.award_number,
        'note': '',
    } for app in Application.objects.filter(student=profile, status='Approved')
        .select_related('scholarship').order_by('-submitted_at')]

    for tes in TESApplication.objects.filter(
            student=profile, status='Approved').order_by('-submitted_at'):
        records.append({
            'name': 'Tertiary Education Subsidy (TES)', 'type': 'TES',
            'status': 'Approved',
            'term': ' '.join(x for x in (tes.school_year, tes.semester) if x),
            'award_number': '', 'note': '',
        })

    for req in ScholarshipLinkRequest.objects.filter(
            student=profile).exclude(status='Approved').order_by('-submitted_at'):
        records.append({
            'name': req.get_scholarship_type_display(),
            'type': req.scholarship_type,
            'status': req.status,
            'term': req.term_display,
            'award_number': req.award_number,
            'note': req.remarks if req.status == 'Rejected' else
                    'Declared at registration. The SDSO is still verifying your proof.',
        })
    return records


def _parse_gwa(raw):
    """A GWA out of a form field, or None when it is blank or not a number.

    Anything outside the 1.00–5.00 Philippine grading range is treated as not
    entered rather than saved — a stray keystroke should not overwrite a real
    grade on the profile.
    """
    try:
        value = float((raw or '').strip())
    except (TypeError, ValueError):
        return None
    return value if 1.0 <= value <= 5.0 else None


# Two awards at once are allowed for exactly one pair. TES is a UniFAST subsidy
# and an Academic scholarship is BiPSU's own recognition of a grade, so neither
# is the "other government assistance" that would disqualify the other. Every
# remaining programme is exclusive: a scholar on TDP, DOST, CHED or any of the
# office's imported lists holds that one and nothing else.
DUAL_SCHOLARSHIP_TYPES = frozenset({'Academic', 'TES'})


def held_scholarship_types(profile):
    """The programmes this student already holds, as canonical type keys.

    Three different records can make someone a scholar and they do not all write
    an Application: an approved award (applied for, linked or imported), an
    approved TES application — UniFAST decides those on their own screen and
    never creates an Application row — and an approved link request for the
    active term. An approved link only counts for the term it was granted for,
    so last semester's award does not keep blocking this semester.
    """
    if not profile:
        return set()
    held = set(
        Application.objects.filter(student=profile, status='Approved')
        .values_list('scholarship__type', flat=True))
    if TESApplication.objects.filter(student=profile, status='Approved').exists():
        held.add('TES')
    from .models import SystemSettings
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    held |= set(
        ScholarshipLinkRequest.objects.filter(
            student=profile, status='Approved', term_label=settings_obj.academic_year,
        ).values_list('scholarship_type', flat=True))
    return {t for t in held if t}


def can_hold_alongside(held, wanted):
    """May `wanted` be added to the programmes already `held`?"""
    held = set(held)
    if wanted in held:
        return False
    return not held or held | {wanted} <= DUAL_SCHOLARSHIP_TYPES


def scholarship_block_reason(profile, wanted, label):
    """Why `wanted` is closed to this student, in their words. '' when it is open."""
    held = held_scholarship_types(profile)
    if can_hold_alongside(held, wanted):
        return ''
    if wanted in held:
        return f'You already hold the {label}. There is nothing to apply for.'
    from .models import SCHOLARSHIP_TYPE_CHOICES
    display = dict(SCHOLARSHIP_TYPE_CHOICES)
    names = ', '.join(sorted(display.get(t, t) for t in held))
    return (f'You are already enrolled in {names}. Only TES and an Academic '
            'scholarship may be held at the same time — every other programme is '
            'held on its own.')


def _is_enrolled(profile):
    """True once the student holds any scholarship at all.

    What the nav's Renewal link and the profile's Scholarship Data card turn on.
    Whether a *particular* programme is still open to them is the different
    question `can_hold_alongside` answers, and the Apply pages ask that one.
    Pending or Needs Revision submissions do not count: a student waiting on a
    decision should still see their other options.
    """
    return bool(held_scholarship_types(profile))


def _validate_proof(uploaded, settings_obj):
    """Server-side check for an uploaded proof document.

    The template's accept="..." attribute is advisory only — anything can be
    POSTed — so the extension and size are enforced here.
    """
    if not uploaded:
        return ['Proof document is required.']
    import os
    errors = []
    ext = os.path.splitext(uploaded.name)[1].lower()
    if ext not in ('.pdf', '.jpg', '.jpeg', '.png'):
        errors.append(f'Unsupported file type "{ext or uploaded.name}". Upload a PDF, JPG or PNG.')
    max_mb = settings_obj.max_file_size_mb or 5
    if uploaded.size > max_mb * 1024 * 1024:
        errors.append(f'File is too large ({uploaded.size / 1048576:.1f} MB). Maximum is {max_mb} MB.')
    return errors


@login_required(login_url='/login/')
def student_dashboard(request):
    profile = StudentProfile.objects.filter(user=request.user).first()
    scholarships = Scholarship.objects.filter(is_active=True)
    matched = []
    for s in scholarships:
        score = s.match_score(profile)
        matched.append({'name': s.name, 'description': s.description, 'match': score})
    matched.sort(key=lambda x: x['match'], reverse=True)
    applications = Application.objects.filter(student=profile) if profile else Application.objects.none()
    announcements = Announcement.objects.order_by('-created_at')[:3]
    top_match = matched[0]['match'] if matched else 0
    top_match_offset = round(314 * (1 - top_match / 100))
    timeline = [
        {'date': a.submitted_at, 'title': f"{a.scholarship.name} — {a.status}", 'status': 'done' if a.status == 'Approved' else 'pending'}
        for a in applications.select_related('scholarship').order_by('-submitted_at')[:5]
    ]
    ctx = {
        'profile': profile,
        'enrolled': _is_enrolled(profile),
        'scholarships': matched,
        'announcements': announcements,
        'top_match': top_match,
        'top_match_offset': top_match_offset,
        'timeline': timeline,
        'dashboard': {
            'recommended_count': len([s for s in matched if s['match'] >= 50]),
            'pending_count': applications.filter(status='Pending Validation').count(),
            'approved_count': applications.filter(status='Approved').count(),
            'notification_count': Notification.objects.filter(student=profile, is_read=False).count() if profile else 0,
        },
    }
    return render(request, 'student/dashboard.html', ctx)


@login_required(login_url='/login/')
def student_apply_academic(request):
    from .models import ApplicationDocument
    from .constants import EDITABLE_APPLICATION_STATUSES
    profile = StudentProfile.objects.filter(user=request.user).first()
    # An application still waiting on a decision — or sent back for correction
    # — is the student's to edit, so the form below fills itself in from it and
    # the save updates it rather than filing a second one.
    editing = Application.objects.filter(
        student=profile, status__in=EDITABLE_APPLICATION_STATUSES
    ).select_related('scholarship').order_by('-submitted_at', '-pk').first() if profile else None
    # An Academic scholarship may be held alongside TES and nothing else, so
    # what closes this form is the set of programmes already held rather than
    # 'has any approved award' — see scholarship_block_reason.
    blocked_reason = scholarship_block_reason(
        profile, 'Academic', 'Academic Scholarship')
    if blocked_reason:
        return render(request, 'student/apply_academic.html', {
            'profile': profile, 'blocked': True,
            'blocked_reason': blocked_reason,
            'classification': '', 'eligible': False,
        })
    if request.method == 'POST':
        scholarship = Scholarship.objects.filter(type='Academic').first()
        if scholarship and profile:
            # The GWA the student declares here is the one the whole application
            # is judged on, so it belongs on the profile — the ranking page,
            # reports and the office's own screens all read it from there. It
            # used to be written only into form_data, which left the profile on
            # its 0.0 default and every self-registered student unrankable.
            # The uploaded Certificate of Grades is what the office checks it
            # against.
            declared = _parse_gwa(request.POST.get('gwa'))
            if declared is not None:
                profile.gwa = declared
                profile.save(update_fields=['gwa'])
            if editing:
                # A correction, not a second application. Sending it puts the
                # record back in the queue: 'Needs Revision' meant the office
                # was waiting on this, and the remark that asked for it has
                # been answered.
                app = editing
                app.form_data = request.POST.dict()
                app.status = 'Pending Validation'
                app.remarks = ''
                app.save()
            else:
                app = Application.objects.create(
                    student=profile, scholarship=scholarship,
                    status='Pending Validation',
                    form_data=request.POST.dict()
                )
            for field, label in [
                ('doc_certificate_of_grades', 'Certificate Of Grades'),
                ('doc_certificate_of_enrollment', 'Certificate Of Enrollment'),
                ('doc_prospectus', 'Prospectus'),
                ('doc_id_photo', 'Id Photo'),
                ('doc_application_form', 'Application Form'),
            ]:
                uploaded = request.FILES.get(field)
                if uploaded:
                    # One document per slot: re-uploading replaces what was
                    # there, which is the whole point of being sent back for a
                    # corrected copy.
                    app.documents.filter(name=label).delete()
                    ApplicationDocument.objects.create(application=app, name=label, file=uploaded)
        return redirect('/student/applications/')
    from .constants import (
        COLLEGE_SCHOLAR_MAX_GWA, UNIVERSITY_SCHOLAR_MAX_GWA, academic_classification,
    )
    from .models import SystemSettings
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    gwa = profile.gwa if profile else 0
    classification = academic_classification(gwa)
    # The term is the office's, not the applicant's. The form used to hard-code
    # '2024-2025' and '2nd Semester' into the markup, so every application ever
    # submitted claimed a term two years stale — and then failed to match the
    # office's own semester filter.
    term = SystemSettings.parse_label(settings_obj.academic_year)
    return render(request, 'student/apply_academic.html', {
        'profile': profile,
        'editing': editing,
        'submitted_documents': list(editing.documents.all()) if editing else [],
        'active_school_year': term['sy'],
        'active_semester': term['semester'],
        'classification': classification,
        'eligible': classification in ('University Scholar', 'College Scholar'),
        'enrolled': _is_enrolled(profile),
        # The page re-runs the same rule live as the applicant types, off these
        # numbers, so what it shows and what the view decides cannot disagree.
        'university_max_gwa': UNIVERSITY_SCHOLAR_MAX_GWA,
        'college_max_gwa': COLLEGE_SCHOLAR_MAX_GWA,
    })
    


@login_required(login_url='/login/')
def student_applications(request):
    profile = StudentProfile.objects.filter(user=request.user).first()
    applications = Application.objects.filter(student=profile).select_related('scholarship') if profile else Application.objects.none()
    tes_applications = TESApplication.objects.filter(student=profile) if profile else TESApplication.objects.none()
    return render(request, 'student/applications.html', {
        'applications': applications,
        'tes_applications': tes_applications,
        'enrolled': _is_enrolled(profile),
    })


@login_required(login_url='/login/')
def student_notifications(request):
    profile = StudentProfile.objects.filter(user=request.user).first()
    notifications = Notification.objects.filter(student=profile).order_by('-created_at') if profile else Notification.objects.none()
    return render(request, 'student/notifications.html', {'notifications': notifications, 'enrolled': _is_enrolled(profile)})


@login_required(login_url='/login/')
def student_renewal_academic(request):
    from .models import SystemSettings
    profile = StudentProfile.objects.filter(user=request.user).first()
    has_academic = Application.objects.filter(student=profile, scholarship__type='Academic', status='Approved').exists() if profile else False
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    if request.method == 'POST' and profile:
        cog = request.FILES.get('certificate_of_grades')
        coe = request.FILES.get('certificate_of_enrollment')
        errors = []
        if not cog:
            errors.append('Certificate of Grades is required.')
        if not coe:
            errors.append('Certificate of Enrollment is required.')
        if errors:
            renewals = AcademicRenewal.objects.filter(student=profile).order_by('-submitted_at')
            return render(request, 'student/renewal_academic.html', {
                'profile': profile, 'has_academic': has_academic,
                'renewals': renewals, 'errors': errors,
                'semester': settings_obj.active_semester, 'academic_year': settings_obj.academic_year,
                'enrolled': _is_enrolled(profile),
            })
        # A renewal still waiting on the office is the student's to correct —
        # the usual reason to come back is that the Certificate of Grades they
        # uploaded was the wrong semester's. Replacing it beats a second
        # submission, which would leave the office two to reconcile.
        pending = AcademicRenewal.objects.filter(
            student=profile, status='Pending', term_label=settings_obj.academic_year,
        ).order_by('-submitted_at').first()
        if pending:
            pending.certificate_of_grades = cog
            pending.certificate_of_enrollment = coe
            pending.save(update_fields=['certificate_of_grades', 'certificate_of_enrollment'])
        else:
            AcademicRenewal.objects.create(
                student=profile, certificate_of_grades=cog, certificate_of_enrollment=coe)
        return redirect('/student/renewal/academic/?submitted=1')
    renewals = AcademicRenewal.objects.filter(student=profile).order_by('-submitted_at') if profile else []
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    pending = AcademicRenewal.objects.filter(
        student=profile, status='Pending', term_label=settings_obj.academic_year,
    ).order_by('-submitted_at').first() if profile else None
    return render(request, 'student/renewal_academic.html', {
        'profile': profile, 'has_academic': has_academic, 'editing': pending,
        'renewals': renewals, 'submitted': request.GET.get('submitted'),
        'semester': parsed['semester'], 'academic_year': parsed['sy'],
        'enrolled': _is_enrolled(profile),
    })


@login_required(login_url='/login/')
@login_required(login_url='/login/')
def student_profile(request):
    profile = StudentProfile.objects.filter(user=request.user).first()
    errors = []
    saved = False
    if request.method == 'POST' and profile:
        p = request.POST
        u = profile.user
        u.save()
        # The given and family names are the office's to set, but nothing in the
        # system ever collected the middle name — which the masterlist exports
        # carry as their own MIDDLE NAME and M.I. columns — so the student
        # enters it here. Locked once filled, like the address below.
        if not profile.middle_name:
            profile.middle_name = p.get('middle_name', '').strip()
        profile.suffix = p.get('suffix', profile.suffix).strip()
        # Civil status, educational background and family background are all
        # locked once set, like the address below. They are what the masterlist
        # and the TES form are built from, so a later edit would silently change
        # a record the office has already reviewed. The archives edit screen is
        # the office's override when one of them was entered wrong.
        if not profile.civil_status:
            profile.civil_status = p.get('civil_status', profile.civil_status)
        if not profile.birth_place:
            profile.birth_place = p.get('birth_place', profile.birth_place).strip()
        profile.family_income = float(p.get('family_income', profile.family_income) or profile.family_income)
        profile.indigenous_group = p.get('indigenous_group', profile.indigenous_group)
        # PWD is asked the way CHED asks it — which disability, from their own
        # list — so the profile and the TES form cannot end up disagreeing about
        # the same student. 'NO' is how that list spells not applicable.
        disability, problem = _disability_answer(p)
        if problem:
            errors.append(problem)
        else:
            profile.disability_type = disability
        if not (profile.elementary and profile.highschool and profile.last_school):
            profile.elementary = p.get('elementary', profile.elementary)
            profile.highschool = p.get('highschool', profile.highschool)
            profile.last_school = p.get('last_school', profile.last_school)
        # Parent names in parts — the shape CHED's TES form needs, collected
        # once here so the TES application can fill itself in from the profile.
        if not (profile.father_last_name and profile.father_first_name
                and profile.mother_last_name and profile.mother_first_name):
            for parent in ('father', 'mother'):
                for part in ('last_name', 'first_name', 'middle_name'):
                    field = f'{parent}_{part}'
                    setattr(profile, field, p.get(field, getattr(profile, field)).strip())
            profile.father_occupation = p.get('father_occupation', profile.father_occupation)
            profile.mother_occupation = p.get('mother_occupation', profile.mother_occupation)
        # TES eligibility. Each of these answers one rule in api/tes_ranking.py;
        # left unanswered they stay unknown, and the recommender reports the
        # requirement as Needs Verification rather than failing the student.
        profile.citizenship = p.get('citizenship', profile.citizenship)
        profile.household_size = _positive_int(p.get('household_size'), profile.household_size)
        profile.year_first_enrolled = _positive_int(
            p.get('year_first_enrolled'), profile.year_first_enrolled)
        profile.is_listahanan_household = _tristate(
            p.get('is_listahanan_household'), profile.is_listahanan_household)
        profile.is_4ps_beneficiary = _tristate(
            p.get('is_4ps_beneficiary'), profile.is_4ps_beneficiary)
        profile.has_previous_degree = _tristate(
            p.get('has_previous_degree'), profile.has_previous_degree)
        # Affirmative eligibility
        raw_shs = p.get('shs_gpa', '').strip()
        if raw_shs:
            try: profile.shs_gpa = float(raw_shs)
            except ValueError: pass
        raw_total = p.get('suc_exam_total', '').strip()
        if raw_total:
            try:
                total = float(raw_total)
                # A total of zero would divide by nothing; treat it as unset.
                profile.suc_exam_total = total if total > 0 else None
            except ValueError:
                pass
        else:
            profile.suc_exam_total = None
        raw_suc = p.get('suc_exam_score', '').strip()
        if raw_suc:
            try: profile.suc_exam_score = float(raw_suc)
            except ValueError: pass
        profile.is_tes_beneficiary = 'is_tes_beneficiary' in p
        if request.FILES.get('shs_gpa_cert'):
            profile.shs_gpa_cert = request.FILES['shs_gpa_cert']
        if request.FILES.get('suc_exam_cert'):
            profile.suc_exam_cert = request.FILES['suc_exam_cert']
        # Address: only update if not yet locked (all three empty)
        if not (profile.barangay and profile.municipality and profile.province):
            profile.barangay = p.get('barangay', profile.barangay)
            profile.municipality = p.get('municipality', profile.municipality)
            profile.province = p.get('province', profile.province)
        if not errors:
            profile.save()
            saved = True
    import json
    from .constants import CIVIL_STATUSES
    address_locked = bool(profile and profile.barangay and profile.municipality and profile.province)
    # A group locks only once it is complete, so a half-filled one stays open.
    civil_status_locked = bool(profile and profile.civil_status)
    birth_place_locked = bool(profile and profile.birth_place)
    education_locked = bool(
        profile and profile.elementary and profile.highschool and profile.last_school)
    family_locked = bool(
        profile and profile.father_last_name and profile.father_first_name
        and profile.mother_last_name and profile.mother_first_name)
    return render(request, 'student/profile.html', {
        'profile': profile, 'errors': errors, 'saved': saved,
        'enrolled': _is_enrolled(profile),
        'bipsu_schools': BIPSU_SCHOOLS,
        'bipsu_courses_json': json.dumps(BIPSU_COURSES),
        'address_locked': address_locked,
        'middle_name_locked': bool(profile and profile.middle_name),
        'civil_status_locked': civil_status_locked,
        'birth_place_locked': birth_place_locked,
        'education_locked': education_locked,
        'family_locked': family_locked,
        'civil_statuses': CIVIL_STATUSES,
        'scholarships_held': _scholarship_records(profile),
        **_disability_fields(
            request.POST.get('disability_type') if errors else None,
            request.POST.get('disability_type_other', '') if errors else '',
            profile.disability_type if profile else ''),
    })


# — UniFAST portal pages ———————————————————————————

# Scholarship types the UniFAST office administers. Their portal is scoped to
# these everywhere — archives, analytics and reports alike.
UNIFAST_TYPES = ['TDP', 'TES']

# Programmes the VPSEA office must not review. TES applications are decided by
# UniFAST at /unifast/tes-applications/; VPSEA neither lists nor acts on them.
VPSEA_EXCLUDED_TYPES = ['TES']


def _unifast_required(view_fn):
    from functools import wraps
    @wraps(view_fn)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'unifast':
            return redirect('/login/')
        return view_fn(request, *args, **kwargs)
    return wrapper


@_unifast_required
def unifast_dashboard(request):
    """Everything the UniFAST office acts on, for the active semester.

    Every figure is derived from records in the system — the office reviews TES
    applications here and imports TDP/TES lists, so there is no need to invent
    billing percentages.
    """
    from .models import (Announcement, ImportedScholar, ActivityLog, SystemSettings,
                         TESApplication)
    from . import tes_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    label = settings_obj.academic_year
    parsed = SystemSettings.parse_label(label)

    tes = TESApplication.objects.all()
    tes_pending = tes.filter(status='Pending')
    tes_approved = tes.filter(status='Approved').count()
    tes_rejected = tes.filter(status='Rejected').count()
    tdp_scholars = Application.objects.filter(
        status='Approved', scholarship__type='TDP').count()

    reviewed = tes_approved + tes_rejected
    approval_rate = round(tes_approved * 100 / reviewed) if reviewed else 0

    # Imported rows still waiting to be claimed by a student account.
    imported = {
        stype: ImportedScholar.objects.filter(
            scholarship_type=stype, term_label=label, claimed_by__isnull=True,
        ).count()
        for stype in UNIFAST_TYPES
    }

    programs = [
        {'label': 'TES', 'count': tes_approved, 'color': '#3b5bdb',
         'href': '/unifast/tes-applications/'},
        {'label': 'TDP', 'count': tdp_scholars, 'color': '#1971c2',
         'href': '/unifast/archives/?type=TDP'},
    ]
    biggest = max((p['count'] for p in programs), default=0) or 1
    for p in programs:
        p['width'] = round(p['count'] * 100 / biggest)

    return render(request, 'unifast/dashboard.html', {
        'academic_year': parsed['sy'],
        'semester': parsed['semester'],
        'tes_total': tes.count(),
        'tes_pending_count': tes_pending.count(),
        'tes_beneficiaries': tes_approved,
        'tes_rejected': tes_rejected,
        'approval_rate': approval_rate,
        'tdp_scholars': tdp_scholars,
        'total_scholars': tes_approved + tdp_scholars,
        # No billing figure is derived here. It was grantees x a flat rate, which
        # is not something this office decides, and no template rendered it.
        'programs': programs,
        'imported': imported,
        'imported_total': sum(imported.values()),
        'pending_queue': list(
            tes_pending.select_related('student__user', *STUDENT_DETAILS).order_by('submitted_at')[:6]
        ),
        'announcements': list(
            Announcement.objects.filter(published_by=request.user)
            .order_by('-created_at')[:3]
        ),
        'recent_activity': list(
            ActivityLog.objects.filter(user=request.user)
            .order_by('-created_at')[:5]
        ),
    })



# — VPSEA portal pages ———————————————————————————————————

def _vpsea_required(view_fn):
    from functools import wraps
    @wraps(view_fn)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'vpsea':
            return redirect('/login/')
        return view_fn(request, *args, **kwargs)
    return wrapper


@_vpsea_required
def vpsea_affirmative_applications(request):
    from .constants import DECIDED_APPLICATION_STATUSES
    from .models import AffirmativeStaffApplication, Application
    from urllib.parse import quote
    if request.method == 'POST':
        app_id = request.POST.get('app_id')
        new_status = request.POST.get('status')
        remarks = request.POST.get('remarks', '')
        tab = request.POST.get('tab', 'affirmative')
        if tab == 'academic':
            try:
                # excluded types are UniFAST's to decide, so a posted id for one
                # finds nothing here rather than being approved by this office.
                acad_app = (
                    Application.objects.select_related('student')
                    .exclude(scholarship__type__in=VPSEA_EXCLUDED_TYPES)
                    .get(id=app_id)
                )
                # The office decides once. The buttons are already hidden for a
                # decided application, so reaching here means a stale page or a
                # posted id — either way the recorded decision stands.
                if acad_app.status in DECIDED_APPLICATION_STATUSES:
                    return redirect('/vpsea/affirmative/?tab=academic&error=' + quote(
                        f'APP-{acad_app.id:07d} was already decided '
                        f'({acad_app.status}). A decision is made once.'))
                acad_app.status = new_status
                acad_app.remarks = remarks
                acad_app.save()
                notify.decision(
                    acad_app.student, f'Your {acad_app.scholarship.name} application',
                    new_status, remarks, link='/student/applications/',
                )
            except Application.DoesNotExist:
                pass
            return redirect(f'/vpsea/affirmative/?tab=academic')
        else:
            try:
                aff_app = AffirmativeStaffApplication.objects.get(id=app_id)
                if aff_app.status in DECIDED_APPLICATION_STATUSES:
                    return redirect(f'/vpsea/affirmative/?tab={tab}&error=' + quote(
                        f'{aff_app.full_name} was already decided '
                        f'({aff_app.status}). A decision is made once.'))
                aff_app.status = new_status
                aff_app.remarks = remarks
                aff_app.save()
                notify.decision(
                    aff_app.email,
                    f'Your {aff_app.get_qualified_for_display()} application',
                    new_status, remarks,
                )
                # On approval: create Django User + StudentProfile + Application
                if new_status == 'Approved' and not User.objects.filter(email=aff_app.email).exists():
                    name_parts = aff_app.full_name.strip().split()
                    first_name = name_parts[0] if name_parts else ''
                    last_name = name_parts[-1] if len(name_parts) > 1 else ''
                    raw_password = aff_app.student_id or aff_app.email.split('@')[0]
                    new_user = User.objects.create_user(
                        username=aff_app.email,
                        email=aff_app.email,
                        password=raw_password,
                        first_name=first_name,
                        last_name=last_name,
                        role='student',
                    )
                    profile = StudentProfile.objects.create(
                        user=new_user,
                        student_id=aff_app.student_id or f'AFF-{aff_app.id}',
                        school=aff_app.school,
                        course=aff_app.course,
                        year_level=aff_app.year_level,
                        contact_number=aff_app.contact_number,
                        barangay=aff_app.barangay,
                        municipality=aff_app.municipality,
                        province=aff_app.province,
                        date_of_birth=aff_app.date_of_birth,
                        gender=aff_app.gender,
                    )
                    scholarship = Scholarship.objects.filter(type=aff_app.qualified_for).first()
                    if scholarship:
                        Application.objects.create(
                            student=profile,
                            scholarship=scholarship,
                            status='Approved',
                            remarks=remarks,
                            form_data={},
                        )
            except AffirmativeStaffApplication.DoesNotExist:
                pass
            return redirect(f'/vpsea/affirmative/?tab={tab}')

    # 'affirmative' was a tab until that programme moved to Student Ranking; an
    # old bookmark for it lands back on Academic rather than on the staff table.
    tab = request.GET.get('tab', 'academic')
    if tab not in ('academic', 'staff'):
        tab = 'academic'
    academic_apps = (
        Application.objects
        .select_related('student__user', 'scholarship', *STUDENT_DETAILS)
        .prefetch_related('documents')
        .exclude(scholarship__type__in=['Staff'])
        .exclude(scholarship__type__in=VPSEA_EXCLUDED_TYPES)
        .order_by('-submitted_at')
    )
    # No Affirmative queue: nobody applies for that programme. It is worked out
    # from the student's own profile and read on the Student Ranking page.
    staff_apps = AffirmativeStaffApplication.objects.filter(
        qualified_for='Staff').select_related(*STAFF_APPLICATION_DETAILS).order_by('-submitted_at')
    return render(request, 'vpsea/affirmative.html', {
        'academic_apps': academic_apps,
        'staff_apps': staff_apps,
        'active_tab': tab,
    })


@_vpsea_required
def vpsea_dashboard(request):
    from .models import Application, AcademicRenewal, AffirmativeStaffApplication, SystemSettings
    from django.db.models import Q
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    active_sy = parsed['sy']
    active_semester = parsed['semester']
    # One indexed column, set on every row. The filter this replaces read
    # form_data['academic_year'], a key the student apply form never wrote —
    # it wrote 'school_year' — so every student-submitted application was
    # silently missing from these counts.
    apps = Application.objects.filter(term_label=settings_obj.academic_year)
    ctx = {
        'total_applicants': apps.count(),
        'approved': apps.filter(status='Approved').count(),
        'rejected': apps.filter(status='Rejected').count(),
        'pending': apps.filter(status='Pending Validation').count(),
        'renewals': AcademicRenewal.objects.filter(status='Pending').count(),
        'pending_staff': AffirmativeStaffApplication.objects.filter(qualified_for='Staff', status='Pending Validation').count(),
        'pending_affirmative': AffirmativeStaffApplication.objects.filter(qualified_for='Affirmative', status='Pending Validation').count(),
        'active_sy_display': f"{active_sy} — {active_semester}",
    }
    return render(request, 'vpsea/dashboard.html', ctx)


@_vpsea_required
def vpsea_renewals(request):
    if request.method == 'POST':
        from .models import SystemSettings
        from django.utils import timezone
        renewal_id = request.POST.get('renewal_id')
        new_status = request.POST.get('status')
        remarks = request.POST.get('remarks', '')
        AcademicRenewal.objects.filter(id=renewal_id).update(status=new_status, remarks=remarks)
        reviewed = AcademicRenewal.objects.select_related('student__user', *STUDENT_DETAILS).filter(id=renewal_id).first()
        if reviewed:
            notify.decision(
                reviewed.student, 'Your scholarship renewal', new_status, remarks,
                link='/student/renewal/',
            )
        # On approval: create a new Application for the current semester so the
        # student appears in the current-SY archives.
        if new_status == 'Approved':
            try:
                renewal = AcademicRenewal.objects.select_related('student').get(id=renewal_id)
                scholarship = Scholarship.objects.filter(type='Academic').first()
                if scholarship:
                    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
                    parsed = SystemSettings.parse_label(settings_obj.academic_year)
                    already = Application.objects.filter(
                        student=renewal.student, scholarship=scholarship,
                        school_year=parsed['sy'], semester=parsed['semester'],
                    ).exists()
                    if not already:
                        Application.objects.create(
                            student=renewal.student,
                            scholarship=scholarship,
                            status='Approved',
                            remarks=remarks,
                            source='renewal',
                            school_year=parsed['sy'],
                            semester=parsed['semester'],
                            form_data={'renewal_id': renewal_id},
                        )
            except AcademicRenewal.DoesNotExist:
                pass
        return redirect('/vpsea/renewals/')
    renewals = AcademicRenewal.objects.select_related('student__user', *STUDENT_DETAILS).order_by('-submitted_at')
    return render(request, 'vpsea/renewals.html', {
        'renewals': renewals,
        'approved_count': renewals.filter(status='Approved').count(),
        'pending_count': renewals.filter(status='Pending').count(),
    })


def _archive_candidates(req, label=None):
    """Imported archive rows that could be the scholar behind a link request.

    Matched on student number, award number or exact first+last name. Rows that
    another student already claimed are never offered again.
    """
    from .models import ImportedScholar
    from django.db.models import Q

    profile = req.student
    qs = ImportedScholar.objects.filter(
        scholarship_type=req.scholarship_type, claimed_by__isnull=True,
    )
    if label is not None:
        qs = qs.filter(term_label=label)

    cond = Q()
    if profile.student_id:
        cond |= Q(student_id__iexact=profile.student_id)
    if req.award_number:
        cond |= Q(award_number__iexact=req.award_number)
    last = (profile.user.last_name or '').strip()
    first = (profile.user.first_name or '').strip()
    if last and first:
        cond |= Q(last_name__iexact=last, first_name__iexact=first)
    if not cond:
        return qs.none()
    return qs.filter(cond).order_by('last_name', 'first_name')


def declared_scholarship(profile):
    """The scholarship this student declared at registration, if it is undecided.

    One at a time: the registration form asks once, and a second declaration
    would be a second account. Returned so the account verification queue can
    show what was claimed and decide it in the same action.
    """
    if not profile:
        return None
    return (ScholarshipLinkRequest.objects
            .select_related('student__user', *STUDENT_DETAILS)
            .filter(student=profile, status='Pending')
            .order_by('-submitted_at').first())


def approve_declared_scholarship(req, reviewer, archive=None, remarks='', tier=''):
    """Turn a declared scholarship into the award it claims to be.

    This is what merges the office's imported data with the student's account:
    it writes the Approved Application for the active semester — so the scholar
    flows into archives, reports, ranking and renewals like any other — and
    marks the matched imported row as claimed so the same person is not counted
    twice. Returns (application, error); the error is a sentence for the officer.
    """
    from .models import (CHED_TIER_CHOICES, ImportedScholar, Notification,
                         ActivityLog, SystemSettings)
    from django.utils import timezone

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    profile = req.student
    label = req.get_scholarship_type_display()

    scholarship = Scholarship.objects.filter(type=req.scholarship_type).first()
    if not scholarship:
        return None, (f'No {req.scholarship_type} program is configured under '
                      'Scholarship Programs, so the award cannot be recorded.')

    # The reviewer can correct the tier the student picked — the proof document
    # is in front of them and a student is not always sure which one they were
    # awarded. It rides on the award as 'scholar_type', the key every CHED
    # masterlist splits its two blocks by.
    form_data = {}
    if req.scholarship_type == 'CHED':
        tier = tier or req.award_tier
        if tier not in [t for t, _ in CHED_TIER_CHOICES]:
            return None, ('Choose Full or Half Merit before verifying a CHED '
                          'scholar — the masterlists report the two separately.')
        req.award_tier = tier          # persisted by the req.save() below
        form_data['scholar_type'] = dict(CHED_TIER_CHOICES)[tier]

    if archive is not None:
        form_data['imported_from'] = archive.imported_from

    # The declaration behind this award is reachable in reverse through
    # ScholarshipLinkRequest.linked_application, so it is not copied here.
    award_fields = {
        'source': 'link',
        'school_year': parsed['sy'],
        'semester': parsed['semester'],
        'award_number': req.award_number or (archive.award_number if archive else ''),
        'congress_district': archive.congress_district if archive else '',
        'claimed_archive': archive,
    }

    # Reuse this semester's row if one already exists, so a re-approval after a
    # correction does not leave the scholar counted twice.
    app = Application.objects.filter(
        student=profile, scholarship=scholarship,
        school_year=parsed['sy'], semester=parsed['semester'],
    ).first()
    if app:
        app.status = 'Approved'
        app.remarks = remarks
        for field, value in award_fields.items():
            setattr(app, field, value)
        app.form_data = {**app.form_data, **form_data}
        app.save()
    else:
        app = Application.objects.create(
            student=profile, scholarship=scholarship,
            status='Approved', remarks=remarks, form_data=form_data,
            **award_fields,
        )

    if archive is not None:
        archive.claimed_by = profile
        archive.save(update_fields=['claimed_by'])
        # Carry over what the office already knows, without overwriting
        # anything the student filled in themselves.
        changed = []
        for field, value in (
            ('course', archive.course), ('barangay', archive.barangay),
            ('municipality', archive.municipality), ('province', archive.province),
            ('gender', archive.gender),
        ):
            if value and not getattr(profile, field):
                setattr(profile, field, value)
                changed.append(field)
        if archive.gwa and not profile.gwa:
            profile.gwa = archive.gwa
            changed.append('gwa')
        if archive.year_level and profile.year_level in (0, 1):
            profile.year_level = archive.year_level
            changed.append('year_level')
        if changed:
            profile.save(update_fields=changed)

    req.status = 'Approved'
    req.remarks = remarks
    req.reviewed_by = reviewer
    req.reviewed_at = timezone.now()
    req.matched_archive = archive
    req.linked_application = app
    req.save()

    Notification.objects.create(
        student=profile, type='success',
        title=f'{label} linked to your account',
        body=(f'Your {label} has been verified and linked for '
              f"{parsed['sy']} {parsed['semester']}. It now appears under My Applications."),
    )
    ActivityLog.objects.create(
        user=reviewer,
        action=(f'Verified the {label} declared by {profile.student_id}'
                + (f' (merged imported row #{archive.id})' if archive
                   else ' (no imported row matched)')),
    )
    return app, ''


def reject_declared_scholarship(req, reviewer, remarks):
    """Turn down a declared scholarship without writing an award."""
    from .models import Notification, ActivityLog
    from django.utils import timezone

    label = req.get_scholarship_type_display()
    req.status = 'Rejected'
    req.remarks = remarks
    req.reviewed_by = reviewer
    req.reviewed_at = timezone.now()
    req.save()
    Notification.objects.create(
        student=req.student, type='warning',
        title=f'{label} could not be verified',
        body=f'The SDSO could not verify the {label} you declared. Reason: {remarks}',
    )
    ActivityLog.objects.create(
        user=reviewer,
        action=f'Rejected the {label} declared by {req.student.student_id}',
    )


# The archives page answers "who holds scholarship X this term" one tab per
# programme, so a student holding nothing appears on no tab at all. This is the
# tab for the office's other question: who has the system not served.
UNAWARDED_TAB = 'No Scholarship'


def _unawarded_rows(term_label):
    """Students with no Approved award for ``term_label``, with why attached.

    "No scholarship" is not one state. A student who never applied needs an
    invitation; one waiting on review needs the queue cleared; one rejected
    needs a reason. The office can only act on the difference, so each row
    carries the student's most recent application — if they have one at all.

    An account the office rejected at registration is none of those. That
    person was turned away at the door and cannot sign in to apply, so they
    are not a student the system failed to serve — they were never admitted
    to it. They belong on the accounts screen, not in the archives.
    """
    from .constants import academic_classification

    awarded = Application.objects.filter(
        status='Approved', term_label=term_label,
    ).values_list('student_id', flat=True)

    students = (
        StudentProfile.objects
        .exclude(id__in=awarded)
        .exclude(user__verification_status='rejected')
        .select_related('user', *StudentProfile.DETAIL_RELATIONS)
        .prefetch_related('applications__scholarship')
        .order_by('user__last_name', 'user__first_name')
    )

    rows = []
    for profile in students:
        # Already prefetched, so this sorts in memory rather than re-querying.
        # submitted_at is a DateField, so two applications sent on the same day
        # tie on it; pk breaks the tie, otherwise a student who re-applied after
        # a rejection would still show as rejected.
        apps = sorted(profile.applications.all(),
                      key=lambda a: (a.submitted_at or date.min, a.pk),
                      reverse=True)
        latest = apps[0] if apps else None
        if latest is None:
            state = 'never'
        elif latest.status == 'Rejected':
            state = 'rejected'
        else:
            state = 'pending'
        rows.append({
            'profile': profile,
            'latest': latest,
            'state': state,
            'classification': academic_classification(profile.gwa),
        })
    return rows


def _scholar_groups(stype, groups, portal='vpsea'):
    """The archive tables for one programme, as the template renders them.

    ``groups`` is ``[(title, records, empty message)]`` — one entry for most
    programmes, two for CHED, which the office reports in a Full and a Half
    block. Returns the resolved columns alongside, because the heading row and
    the cells have to come from the same list or they drift apart.
    """
    from . import scholar_columns
    from .models import Scholarship

    programme = Scholarship.objects.filter(type=stype).first()
    # The type is passed as well: an archive tab can name a programme that has
    # no Scholarship row, and the default columns are chosen by type.
    columns = scholar_columns.resolve(programme, stype, portal)
    built = []
    for title, records, empty in groups:
        records = list(records)
        built.append({
            'title': title,
            'rows': scholar_columns.rows_for(records, columns),
            'empty': empty,
        })
    return {
        'columns': columns,
        'scholar_groups': built,
        'has_custom_columns': any(c['custom'] for c in columns),
    }


@_vpsea_required
def vpsea_archives(request):
    from .models import Application, AffirmativeStaffApplication, ScholarListImport, SystemSettings, AcademicRenewal, ActivityLog
    stype = request.GET.get('type', 'Academic')
    # Build archive_types dynamically from DB so newly added scholarship types appear
    base_types = ['Academic', 'TDP', 'DOST', 'CHED', 'CoScho', 'Sports', 'Affirmative', 'Staff', 'GSIS']
    db_types = list(Scholarship.objects.values_list('type', flat=True).distinct())
    archive_types = base_types + [t for t in db_types if t not in base_types] + [UNAWARDED_TAB]
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active_label = settings_obj.academic_year  # e.g. '26-1'
    parsed = SystemSettings.parse_label(active_label)
    active_sy = parsed['sy']           # '2025-2026'
    active_semester = parsed['semester']  # '1st Semester'

    history = ScholarListImport.objects.filter(scholarship_type=stype).order_by('-created_at')
    all_labels = list(
    ScholarListImport.objects.filter(scholarship_type=stype)
    .values_list('term_label', flat=True).distinct().order_by('-term_label')
)
    if active_label not in all_labels:
        all_labels.insert(0, active_label)
    selected_label = request.GET.get('sy', active_label)
    if selected_label not in all_labels:
        selected_label = active_label

    selected_parsed = SystemSettings.parse_label(selected_label)
    selected_sy = selected_parsed['sy']   # e.g. '2025-2026'
    sy_start = selected_parsed['sy_start']
    sy_end = selected_parsed['sy_end']

    next_label = settings_obj.next_label()
    # Human-readable label for display: '2025-2026 — 1st Semester'
    active_display = f"{active_sy} — {active_semester}"
    # Build (label, display) pairs for the SY dropdown
    all_sy_display = []
    for lbl in all_labels:
        p = SystemSettings.parse_label(lbl)
        all_sy_display.append((lbl, f"{p['sy']} — {p['semester']}"))

    # Previous label = one step back from active
    yy, s = active_label.split('-')
    if s == '2':
        prev_label = f'{yy}-1'
    else:
        prev_label = f'{int(yy)-1}-2'
    prev_parsed = SystemSettings.parse_label(prev_label)
    prev_display = f"{prev_parsed['sy']} — {prev_parsed['semester']}"

    import json as _json
    base_ctx = {
        'archive_types': archive_types,
        'bipsu_schools': BIPSU_SCHOOLS,
        'bipsu_courses_json': _json.dumps(BIPSU_COURSES),
        'active_type': stype,
        'history': history,
        'all_sy': all_labels,
        'all_sy_display': all_sy_display,
        'selected_sy': selected_label,
        'active_sy': active_label,
        'active_sy_display': active_display,
        'active_semester': active_semester,
        'next_sy': next_label,
        'prev_sy': prev_label,
        'prev_sy_display': prev_display,
        'add_docs': [
            (1, 'Certificate of Grades', 'Official COG from the Registrar for the previous semester.', 'doc_certificate_of_grades'),
            (2, 'Certificate of Enrollment', 'Official COE from the Registrar for the current semester.', 'doc_certificate_of_enrollment'),
            (3, 'Prospectus / Subject Checklist', 'Program prospectus or subject checklist showing enrolled subjects.', 'doc_prospectus'),
            (4, '2Ã—2 ID Photo', 'Recent 2Ã—2 ID photo with white background.', 'doc_id_photo'),
            (5, 'Application Form', 'Signed and accomplished scholarship application form.', 'doc_application_form'),
        ],
        'col_hint': COLUMN_HINTS.get(stype, ''),
        'recent_imports': ActivityLog.objects.filter(action__icontains='Imported').order_by('-created_at')[:5],
        'import_message': f"Successfully imported {request.GET.get('import_ok')} records." if request.GET.get('import_ok') else None,
        'import_error': request.GET.get('import_error'),
    }

    if stype == UNAWARDED_TAB:
        rows = _unawarded_rows(selected_label)
        return render(request, 'vpsea/archives_unawarded.html', {
            **base_ctx,
            'rows': rows,
            'total': len(rows),
        })

    # Renewal-gated: only scholars with an Approved AcademicRenewal are shown
    # For Affirmative/Staff there is no AcademicRenewal flow — show all approved
    from .models import ImportedScholar
    # Rows claimed through an approved link request are excluded — that scholar
    # now has a live Application row below, so listing both double counts them.
    imported_rows = ImportedScholar.objects.filter(
        scholarship_type=stype, term_label=selected_label, claimed_by__isnull=True,
    ).order_by('last_name', 'first_name')

    if stype in ('Affirmative', 'Staff'):
        aff_scholars = AffirmativeStaffApplication.objects.filter(
            status='Approved', qualified_for=stype
        ).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name')
        return render(request, 'vpsea/archives.html', {
            **base_ctx,
            **_scholar_groups(stype, [
                (None, list(aff_scholars) + list(imported_rows),
                 f'No approved {stype} scholars yet.'),
            ]),
            'total': aff_scholars.count() + imported_rows.count(),
        })

    def sy_filter(qs):
        # The term is an indexed column now, so the selected one is simply
        # matched. The old version tried four different form_data keys — the
        # office wrote 'academic_year', the student form wrote 'school_year' —
        # and then gave up for the active term and returned every approved
        # application ever, whatever semester it belonged to.
        return qs.filter(term_label=selected_label)

    if stype == 'CHED':
        all_scholars = sy_filter(Application.objects.filter(
            status='Approved', scholarship__type='CHED',
        )).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')
        full_scholars, half_scholars = split_ched(all_scholars)
        return render(request, 'vpsea/archives.html', {
            **base_ctx,
            # Imported CHED rows carry no tier, and the report has always
            # printed an unclassified scholar under Full rather than dropping
            # them — see split_ched.
            **_scholar_groups(stype, [
                ('Full Merit / Full Scholar',
                 list(full_scholars) + list(imported_rows),
                 'No approved CHED full scholars yet.'),
                ('Half Merit / Partial Scholar', half_scholars,
                 'No approved CHED half scholars yet.'),
            ]),
            'total': all_scholars.count() + imported_rows.count(),
        })

    scholars = sy_filter(Application.objects.filter(
        status='Approved', scholarship__type=stype,
    )).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name').distinct()
    return render(request, 'vpsea/archives.html', {
        **base_ctx,
        **_scholar_groups(stype, [
            (None, list(scholars) + list(imported_rows),
             f'No approved {stype} scholars yet.'),
        ]),
        'total': scholars.count() + imported_rows.count(),
    })


@_vpsea_required
def vpsea_archive_add(request):
    from .models import Application, Scholarship, StudentProfile, User, AffirmativeStaffApplication, SystemSettings, ApplicationDocument
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    p = request.POST
    f = request.FILES
    stype = p.get('scholarship_type', 'Academic')
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    active_sy = parsed['sy']
    active_semester = parsed['semester']

    # Whether this scholar gets a login. The office adds plenty of records for
    # people who will never open the portal — a graduating batch typed up from
    # the agency's own list, say — and those used to get an account anyway, with
    # a password guessable from the student number. Answering "no" keeps the
    # profile every report reads and leaves no usable password behind.
    wants_account = p.get('create_account') == 'yes'
    supplied_email = p.get('email', '').strip()

    if wants_account:
        # An account needs both, and the form asks for neither. Without an
        # address the account is created against a fabricated one the scholar
        # cannot receive mail at; without a student number the password falls
        # back to the literal 'bipsu1234', shared by every account made that
        # way. Refused rather than half-made — an import is the option for a
        # scholar whose details the office does not have.
        missing = []
        if not supplied_email:
            missing.append('an email address')
        if not p.get('student_id', '').strip():
            missing.append('a student number')
        if missing:
            from urllib.parse import quote
            return redirect(f'/vpsea/archives/?type={stype}&error=' + quote(
                'Creating an account needs ' + ' and '.join(missing) + '. The '
                'address is where the scholar is emailed, and the student number '
                'is their first password. Choose "Just an import" to record this '
                'scholar without an account.'))

    if not wants_account:
        # "Just an import": exactly the row an uploaded spreadsheet produces, for
        # a scholar the office is recording rather than enrolling. No account, no
        # password, no profile — ImportedScholar is where every other import
        # lands, and the archive tables already read it for every programme.
        from .models import ImportedScholar
        try:
            year_level = int(p.get('year_level', 0) or 0)
        except (TypeError, ValueError):
            year_level = 0
        try:
            gwa = float(p.get('gwa', 0) or 0)
        except (TypeError, ValueError):
            gwa = 0.0
        ImportedScholar.objects.create(
            scholarship_type=stype,
            term_label=settings_obj.academic_year,
            last_name=p.get('last_name', '').strip(),
            first_name=p.get('first_name', '').strip(),
            middle_name=p.get('middle_name', '').strip(),
            gender=p.get('gender', ''),
            course=p.get('course', ''),
            year_level=year_level,
            gwa=gwa,
            student_id=p.get('student_id', '').strip(),
            award_number=p.get('award_number', ''),
            congress_district=p.get('congress_district', ''),
            barangay=p.get('barangay', ''),
            municipality=p.get('municipality', ''),
            province=p.get('province', ''),
            imported_from='Added by SDSO',
        )
        return redirect(f'/vpsea/archives/?type={stype}&added=1')

    if stype in ('Affirmative', 'Staff'):
        full_name = f"{p.get('first_name','').strip()} {p.get('last_name','').strip()}".strip()
        # A real address when the office has one; otherwise the placeholder this
        # has always fabricated, still deduped because several records can share
        # a blank student number.
        email = supplied_email
        if not email:
            email = f"{p.get('student_id','').strip() or full_name.replace(' ','_').lower()}_{stype.lower()}@bipsu.edu.ph"
            base = email
            counter = 1
            while AffirmativeStaffApplication.objects.filter(email=email).exists():
                email = f"{base.split('@')[0]}_{counter}@bipsu.edu.ph"
                counter += 1
        AffirmativeStaffApplication.objects.create(
            full_name=full_name,
            email=email,
            contact_number=p.get('contact_number', ''),
            barangay=p.get('barangay', ''),
            municipality=p.get('municipality', ''),
            province=p.get('province', ''),
            date_of_birth=p.get('date_of_birth') or '2000-01-01',
            gender=p.get('gender', ''),
            course=p.get('course', ''),
            year_level=int(p.get('year_level', 1) or 1),
            student_id=p.get('student_id', ''),
            qualified_for=stype,
            status='Approved',
            is_nsu_staff=(stype == 'Staff'),
        )
    else:
        email = supplied_email
        student_id = p.get('student_id', '').strip()
        if not email:
            email = f"{student_id}@bipsu.edu.ph"
        user = User.objects.filter(email=email).first()
        account_created = False
        if not user:
            user = User.objects.create_user(
                username=email, email=email,
                password=student_id or 'bipsu1234',
                first_name=p.get('first_name', ''),
                last_name=p.get('last_name', ''),
                role='student',
            )
            account_created = True
        profile = StudentProfile.objects.filter(user=user).first()
        if not profile:
            profile = StudentProfile.objects.create(
                user=user,
                student_id=student_id,
                course=p.get('course', ''),
                year_level=int(p.get('year_level', 1) or 1),
                gwa=float(p.get('gwa', 0) or 0),
                gender=p.get('gender', ''),
                barangay=p.get('barangay', ''),
                municipality=p.get('municipality', ''),
                province=p.get('province', ''),
                contact_number=p.get('contact_number', ''),
                date_of_birth=p.get('date_of_birth') or None,
            )
        else:
            profile.course = p.get('course', profile.course)
            profile.year_level = int(p.get('year_level', profile.year_level) or profile.year_level)
            profile.gwa = float(p.get('gwa', profile.gwa) or profile.gwa)
            profile.gender = p.get('gender', profile.gender)
            profile.barangay = p.get('barangay', profile.barangay)
            profile.municipality = p.get('municipality', profile.municipality)
            profile.province = p.get('province', profile.province)
            profile.save()
        scholarship = Scholarship.objects.filter(type=stype).first()
        if scholarship:
            award_fields = {
                'source': 'import',
                'school_year': active_sy,
                'semester': active_semester,
                'award_number': p.get('award_number', ''),
                'congress_district': p.get('congress_district', ''),
            }
            form_data = {}
            if stype == 'Academic':
                form_data.update({
                    'elementary': p.get('elementary', ''),
                    'highschool': p.get('highschool', ''),
                    'last_school': p.get('last_school', ''),
                })
                # Parents' names belong on the profile, in parts. They used to be
                # written into form_data as one string per parent, so the columns
                # the masterlists and the TES form read stayed empty.
                parent_fields = [
                    'father_last_name', 'father_first_name', 'father_middle_name',
                    'father_occupation', 'mother_last_name', 'mother_first_name',
                    'mother_middle_name', 'mother_occupation',
                ]
                changed = []
                for field in parent_fields:
                    value = p.get(field, '').strip()
                    if value:
                        setattr(profile, field, value)
                        changed.append(field)
                for field in ('elementary', 'highschool', 'last_school'):
                    value = p.get(field, '').strip()
                    if value:
                        setattr(profile, field, value)
                        changed.append(field)
                if changed:
                    profile.save(update_fields=changed)
            already = Application.objects.filter(
                student=profile, scholarship=scholarship,
                school_year=active_sy, semester=active_semester,
            ).exists()
            if not already:
                app = Application.objects.create(
                    student=profile,
                    scholarship=scholarship,
                    status='Approved',
                    form_data=form_data,
                    **award_fields,
                )
            else:
                app = Application.objects.filter(
                    student=profile, scholarship=scholarship,
                    school_year=active_sy, semester=active_semester,
                ).first()
            # Save uploaded documents
            doc_fields = [
                ('doc_certificate_of_grades', 'Certificate Of Grades'),
                ('doc_certificate_of_enrollment', 'Certificate Of Enrollment'),
                ('doc_prospectus', 'Prospectus'),
                ('doc_id_photo', 'Id Photo'),
                ('doc_application_form', 'Application Form'),
                ('proof_document', 'Proof Document'),
            ]
            for field, label in doc_fields:
                uploaded = f.get(field)
                if uploaded:
                    ApplicationDocument.objects.create(application=app, name=label, file=uploaded)
        if account_created and supplied_email:
            notify.notify(
                user, 'Your SRMS account is ready',
                'The VPSEA office has added you to the Scholarship Records '
                'Management System.\n\n'
                f'Sign in with this email address. Your initial password is your '
                f'student number ({student_id}). Contact the VPSEA office to have '
                'it changed.',
                tone='success',
            )
    return redirect(f'/vpsea/archives/?type={stype}&added=1')


def _apply_student_record_edits(profile, p):
    """Correct a student's own details — the typo fixing the archives screens do.

    Shared by the scholarship tabs, which reach a student through their
    Application, and the No Scholarship tab, which has no application to reach
    through. One implementation, so the same Edit button means the same thing
    wherever it is opened.

    Returns an error string, or '' when the edits were saved.
    """
    from .models import StudentProfile

    user = profile.user
    new_sid = (p.get('student_id') or '').strip()
    if (new_sid and new_sid != profile.student_id
            and StudentProfile.objects.filter(student_id=new_sid).exclude(pk=profile.pk).exists()):
        # student_id is unique at the database level; without this the office
        # would get a 500 instead of being told what went wrong.
        return f'Student number {new_sid} already belongs to another student.'

    user.first_name = p.get('first_name', user.first_name)
    user.last_name = p.get('last_name', user.last_name)
    new_pw = (p.get('new_password') or '').strip()
    if new_pw:
        user.set_password(new_pw)
    user.save()

    profile.course = p.get('course', profile.course)
    if p.get('school'):
        profile.school = p.get('school')
    profile.year_level = int(p.get('year_level', profile.year_level) or profile.year_level)
    profile.gender = p.get('gender', profile.gender)
    profile.barangay = p.get('barangay', profile.barangay)
    profile.municipality = p.get('municipality', profile.municipality)
    profile.province = p.get('province', profile.province)
    if new_sid:
        profile.student_id = new_sid
    if p.get('contact_number'):
        profile.contact_number = p.get('contact_number')
    if p.get('gwa'):
        profile.gwa = float(p.get('gwa'))

    # Civil status, educational background and family background lock for the
    # student after their first save. This screen is the office's way to fix one
    # that was entered wrong, so here they stay writable — but only fields the
    # form actually submitted, so a modal that omits them changes nothing.
    for field in ('civil_status', 'elementary', 'highschool', 'last_school',
                  'father_last_name', 'father_first_name', 'father_middle_name',
                  'father_occupation', 'mother_last_name', 'mother_first_name',
                  'mother_middle_name', 'mother_occupation'):
        value = (p.get(field) or '').strip()
        if value:
            setattr(profile, field, value)

    profile.save()
    return ''


@_vpsea_required
def vpsea_student_record_edit(request, pk):
    """Edit a student who holds no award, from the No Scholarship tab.

    Same edits as the scholarship tabs, reached by StudentProfile instead of
    Application — these students have no application to key off.
    """
    from .models import StudentProfile
    from urllib.parse import quote
    back = f'/vpsea/archives/?type={quote(UNAWARDED_TAB)}'
    if request.method != 'POST':
        return redirect(back)
    profile = StudentProfile.objects.select_related('user').filter(pk=pk).first()
    if not profile:
        return redirect(back)
    error = _apply_student_record_edits(profile, request.POST)
    if error:
        return redirect(f'{back}&error={quote(error)}')
    return redirect(f'{back}&edited=1')


@_vpsea_required
def vpsea_archive_edit(request, pk):
    from .models import Application, AffirmativeStaffApplication, StudentProfile, SystemSettings
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    p = request.POST
    stype = p.get('scholarship_type', 'Academic')
    is_aff = stype in ('Affirmative', 'Staff')

    if is_aff:
        try:
            obj = AffirmativeStaffApplication.objects.get(pk=pk)
        except AffirmativeStaffApplication.DoesNotExist:
            return redirect(f'/vpsea/archives/?type={stype}')
        obj.full_name = f"{p.get('first_name','').strip()} {p.get('last_name','').strip()}".strip()
        obj.gender = p.get('gender', obj.gender)
        obj.course = p.get('course', obj.course)
        if p.get('school'):
            obj.school = p.get('school')
        obj.year_level = int(p.get('year_level', obj.year_level) or obj.year_level)
        obj.barangay = p.get('barangay', obj.barangay)
        obj.municipality = p.get('municipality', obj.municipality)
        obj.province = p.get('province', obj.province)
        obj.student_id = p.get('student_id', obj.student_id)
        if p.get('contact_number'):
            obj.contact_number = p.get('contact_number')
        if p.get('date_of_birth'):
            obj.date_of_birth = p.get('date_of_birth')
        # Password reset. The login lives on the User account, not on this row:
        # migration 0045 dropped the password field, so the obj.set_password()
        # this used to call raised AttributeError the moment the field was
        # filled in. Office-added archive rows carry a fabricated email and have
        # no account behind them at all, which is what the miss check is for.
        new_pw = (p.get('new_password') or '').strip()
        if new_pw:
            account = User.objects.filter(email=obj.email).first()
            if not account:
                from urllib.parse import quote
                return redirect(
                    f'/vpsea/archives/?type={stype}&error=' + quote(
                        'This scholar has no login account, so the password '
                        'cannot be reset.')
                )
            account.set_password(new_pw)
            account.save()
        obj.save()
    else:
        try:
            app = Application.objects.select_related('student__user', *STUDENT_DETAILS).get(pk=pk)
        except Application.DoesNotExist:
            return redirect(f'/vpsea/archives/?type={stype}')
        error = _apply_student_record_edits(app.student, p)
        if error:
            from urllib.parse import quote
            return redirect(f'/vpsea/archives/?type={stype}&error={quote(error)}')
        if p.get('award_number') is not None:
            app.award_number = p.get('award_number')
        if p.get('congress_district') is not None:
            app.congress_district = p.get('congress_district')
        app.save()
    return redirect(f'/vpsea/archives/?type={stype}&edited=1')


@_vpsea_required
def vpsea_archive_delete(request, pk):
    from .models import Application, AffirmativeStaffApplication
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    stype = request.POST.get('scholarship_type', 'Academic')
    if stype in ('Affirmative', 'Staff'):
        AffirmativeStaffApplication.objects.filter(pk=pk).delete()
    else:
        Application.objects.filter(pk=pk).delete()
    return redirect(f'/vpsea/archives/?type={stype}&deleted=1')


@_vpsea_required
def vpsea_new_semester(request):
    from .models import SystemSettings, ScholarListImport, ActivityLog, Application, AffirmativeStaffApplication
    from django.core.files.base import ContentFile
    import openpyxl
    from io import BytesIO
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    label = request.POST.get('school_year', '').strip()
    stype = request.POST.get('type', 'Academic')
    if not label or '-' not in label:
        return redirect(f'/vpsea/archives/?type={stype}')

    parsed = SystemSettings.parse_label(label)
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    settings_obj.academic_year = label
    settings_obj.active_semester = parsed['semester']
    settings_obj.save()

    _base = ['Academic', 'TDP', 'DOST', 'CHED', 'CoScho', 'Sports', 'Affirmative', 'Staff', 'GSIS']
    ALL_TYPES = _base + [t for t in Scholarship.objects.values_list('type', flat=True).distinct() if t not in _base]

    def _build_excel(scholarship_type):
        # Header labels and cell positions both come from the import contract,
        # so a rollover file can be uploaded straight back through
        # vpsea_archive_import. They were written out by hand here before and
        # had drifted from COLUMN_MAPS for every scholarship type — Affirmative
        # and Staff in all of their columns, the rest from the middle name on,
        # which re-imported a gender as somebody's first name. An unmapped type
        # falls back to the layout the importer falls back to.
        col_map = COLUMN_MAPS.get(scholarship_type, COLUMN_MAPS['CoScho'])
        hint = COLUMN_HINTS.get(scholarship_type, COLUMN_HINTS['CoScho'])
        header = [h.strip() for h in hint.split('|')]

        if scholarship_type in ('Affirmative', 'Staff'):
            qs = AffirmativeStaffApplication.objects.filter(
                status='Approved', qualified_for=scholarship_type
            ).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name')
        else:
            qs = Application.objects.filter(
                status='Approved', scholarship__type=scholarship_type
            ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Scholars'
        ws.append(header)
        for i, record in enumerate(qs, 1):
            cells = _rollover_fields(record)
            row = [''] * len(header)
            row[0] = i
            for idx, field in col_map:
                row[idx] = cells.get(field, '')
            ws.append(row)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, qs.count()

    created = 0
    for t in ALL_TYPES:
        if ScholarListImport.objects.filter(term_label=label, scholarship_type=t).exists():
            continue
        buf, count = _build_excel(t)
        rollover = ScholarListImport(
            scholarship_type=t,
            school_year=parsed['sy'],
            semester=parsed['semester'],
            term_label=label,
            scholar_count=count,
            imported_by=request.user,
        )
        rollover.excel_file.save(f'{t}_{label}.xlsx', ContentFile(buf.read()), save=True)
        created += 1

    ActivityLog.objects.create(
        user=request.user,
        action=f'New semester started: {label} ({parsed["sy"]} {parsed["semester"]}). Saved lists for {created} scholarship types.'
    )
    return redirect(f'/vpsea/archives/?type={stype}')


@_vpsea_required
def vpsea_undo_semester(request):
    from .models import SystemSettings, ActivityLog
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    prev_label = request.POST.get('prev_label', '').strip()
    stype = request.POST.get('type', 'Academic')
    if not prev_label or '-' not in prev_label:
        return redirect(f'/vpsea/archives/?type={stype}')
    parsed = SystemSettings.parse_label(prev_label)
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    old_label = settings_obj.academic_year
    settings_obj.academic_year = prev_label
    settings_obj.active_semester = parsed['semester']
    settings_obj.save()
    ActivityLog.objects.create(
        user=request.user,
        action=f'Semester undone from {old_label} back to {prev_label} ({parsed["sy"]} {parsed["semester"]}).'
    )
    return redirect(f'/vpsea/archives/?type={stype}')


COLUMN_HINTS = {
    'Academic':    'No. | Last Name | First Name | Middle Name | Sex | Brgy./St. | Municipality | Province | Course | Year | GWA | % / Type | Scholarship',
    'Staff':       'No. | Last Name | First Name | Middle Initial | Sex | Course | Year Level | Student Number | % | Scholarship Program',
    'CHED':        'No. | Award Number | Last Name | First Name | Middle Name | Sex | Brgy./St. | Municipality | Province | Congress District | Course | Yr. | Scholarship Program',
    'TDP':         'No. | Award Number | Last Name | First Name | Middle Name | Sex | Brgy./St. | Municipality | Province | Congress District | Course | Yr. | Scholarship Program',
    'DOST':        'No. | Award Number | Last Name | First Name | Middle Name | Sex | Brgy./St. | Municipality | Province | Congress District | Course | Yr. | Scholarship Program',
    'GSIS':        'No. | Last Name | First Name | Middle Initial | Brgy./St. | Municipality | Province | Sex | Course | Year Level | Student Number | Scholarship Program',
    'Affirmative': 'No. | Award Number | Last Name | First Name | Middle Name | Sex | Brgy./St. | Municipality | Province | Congress District | Course | Yr. | Scholarship Program',
    'CoScho':      'No. | Last Name | First Name | Middle Initial | Sex | Brgy./St. | Municipality | Province | Course | Year Level | Student Number | Scholarship Program',
    'Sports':      'No. | Last Name | First Name | Middle Initial | Sex | Brgy./St. | Municipality | Province | Course | Year Level | Student Number | Scholarship Program',
}

COLUMN_MAPS = {
    'Academic':    [(1,'last_name'),(2,'first_name'),(3,'middle_name'),(4,'sex'),(5,'barangay'),(6,'municipality'),(7,'province'),(8,'course'),(9,'year'),(10,'gwa'),(11,'pct_type'),(12,'scholarship')],
    'Staff':       [(1,'last_name'),(2,'first_name'),(3,'middle_initial'),(4,'sex'),(5,'course'),(6,'year'),(7,'student_number'),(8,'pct'),(9,'scholarship_program')],
    'CHED':        [(1,'award_number'),(2,'last_name'),(3,'first_name'),(4,'middle_name'),(5,'sex'),(6,'barangay'),(7,'municipality'),(8,'province'),(9,'congress_district'),(10,'course'),(11,'year'),(12,'scholarship_program')],
    'TDP':         [(1,'award_number'),(2,'last_name'),(3,'first_name'),(4,'middle_name'),(5,'sex'),(6,'barangay'),(7,'municipality'),(8,'province'),(9,'congress_district'),(10,'course'),(11,'year'),(12,'scholarship_program')],
    'DOST':        [(1,'award_number'),(2,'last_name'),(3,'first_name'),(4,'middle_name'),(5,'sex'),(6,'barangay'),(7,'municipality'),(8,'province'),(9,'congress_district'),(10,'course'),(11,'year'),(12,'scholarship_program')],
    'GSIS':        [(1,'last_name'),(2,'first_name'),(3,'middle_initial'),(4,'barangay'),(5,'municipality'),(6,'province'),(7,'sex'),(8,'course'),(9,'year'),(10,'student_number'),(11,'scholarship_program')],
    'Affirmative': [(1,'award_number'),(2,'last_name'),(3,'first_name'),(4,'middle_name'),(5,'sex'),(6,'barangay'),(7,'municipality'),(8,'province'),(9,'congress_district'),(10,'course'),(11,'year'),(12,'scholarship_program')],
    'CoScho':      [(1,'last_name'),(2,'first_name'),(3,'middle_initial'),(4,'sex'),(5,'barangay'),(6,'municipality'),(7,'province'),(8,'course'),(9,'year'),(10,'student_number'),(11,'scholarship_program')],
    'Sports':      [(1,'last_name'),(2,'first_name'),(3,'middle_initial'),(4,'sex'),(5,'barangay'),(6,'municipality'),(7,'province'),(8,'course'),(9,'year'),(10,'student_number'),(11,'scholarship_program')],
}


def _rollover_fields(record):
    """One approved scholar's cells, keyed by the field names COLUMN_MAPS uses.

    Both shapes an approved scholar can arrive in — an Application with a
    StudentProfile behind it, or an AffirmativeStaffApplication carrying its own
    copy of the details — reduce to the same dict here. That is what lets
    _build_excel lay a rollover out by the import contract without caring which
    table the row came from.
    """
    if isinstance(record, AffirmativeStaffApplication):
        pct = '100' if record.is_nsu_staff else '75'
        name = ('BiPSU Staff Scholarship' if record.is_nsu_staff
                else 'Affirmative Action Scholarship')
        return {
            'last_name': record.last_name,
            'first_name': record.first_name,
            'middle_name': record.middle_name,
            'middle_initial': record.middle_initial,
            'sex': record.gender or '',
            'barangay': record.barangay or '',
            'municipality': record.municipality or '',
            'province': record.province or '',
            'course': record.course or '',
            'year': record.year_level or '',
            'gwa': '',
            'student_number': record.student_id or '',
            'award_number': '',
            'congress_district': '',
            'pct': pct, 'pct_type': pct,
            'scholarship': name, 'scholarship_program': name,
        }

    profile = record.student
    gwa = profile.gwa or 0
    # Academic reports a rank where the other programmes report a percentage —
    # the same rule masterlist_report._application_row applies.
    if record.scholarship.type == 'Academic':
        pct = 'Univ. Scholar' if gwa <= 1.29 else ('College Scholar' if gwa <= 1.50 else '')
    else:
        pct = ''
    return {
        'last_name': profile.user.last_name or '',
        'first_name': profile.user.first_name or '',
        'middle_name': profile.middle_name or '',
        'middle_initial': profile.middle_initial,
        'sex': profile.gender or '',
        'barangay': profile.barangay or '',
        'municipality': profile.municipality or '',
        'province': profile.province or '',
        'course': profile.course or '',
        'year': profile.year_level or '',
        'gwa': f'{gwa:.2f}' if gwa else '',
        'student_number': profile.student_id or '',
        'award_number': record.award_number or '',
        'congress_district': record.congress_district or '',
        'pct': pct, 'pct_type': pct,
        'scholarship': record.scholarship.name,
        'scholarship_program': record.scholarship.name,
    }


def _delete_import_with_scholars(record):
    """Delete a ScholarListImport together with the scholars it brought in.

    The rows an import creates are keyed by scholarship type and term rather
    than by a foreign key back to the import, so nothing cascaded and deleting
    the import used to leave every one of them behind — still counted on the
    analytics screen, still listed in the archive tables, and still holding the
    term open in the semester dropdown, which reads its labels off these rows.

    A row a student has since claimed is deleted too. Both foreign keys that
    point at one — Application.claimed_archive and
    ScholarshipLinkRequest.matched_archive — are SET_NULL, so an approved award
    survives its provenance being removed.

    Returns ``(rows_removed, scholarship_type, term_label)``.
    """
    from .models import ImportedScholar

    stype, label = record.scholarship_type, record.term_label
    with transaction.atomic():
        removed, _ = ImportedScholar.objects.filter(
            scholarship_type=stype, term_label=label,
        ).delete()
        record.excel_file.delete(save=False)
        record.delete()
    return removed, stype, label


@_vpsea_required
def vpsea_imported_delete(request, pk):
    """Remove one imported scholar row.

    The archive tables list imported rows beside portal awards but offered no
    way to remove one — the only delete was the whole import, which takes every
    row with it. A single bad line from a spreadsheet had to be fixed by
    deleting and re-uploading the lot.
    """
    from .models import ImportedScholar, ActivityLog
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    stype = request.POST.get('scholarship_type', 'Academic')
    row = ImportedScholar.objects.filter(pk=pk).first()
    if not row:
        return redirect(f'/vpsea/archives/?type={stype}')
    label = f'{row.full_name} ({row.scholarship_type} {row.term_label})'
    row.delete()
    ActivityLog.objects.create(
        user=request.user,
        action=f'Deleted imported scholar {label}.',
    )
    return redirect(f'/vpsea/archives/?type={stype}&deleted=1')


@_vpsea_required
def vpsea_rollover_delete(request, pk):
    from .models import ScholarListImport, ActivityLog
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    stype = request.POST.get('type', 'Academic')
    try:
        r = ScholarListImport.objects.get(pk=pk)
    except ScholarListImport.DoesNotExist:
        return redirect(f'/vpsea/archives/?type={stype}')
    removed, imported_type, label = _delete_import_with_scholars(r)
    ActivityLog.objects.create(
        user=request.user,
        action=f'Deleted the {imported_type} import for "{label}" and the '
               f'{removed} scholar row(s) it had created.'
    )
    return redirect(f'/vpsea/archives/?type={stype}')


@_vpsea_required
def vpsea_archive_import(request):
    import openpyxl
    from django.core.files.base import ContentFile
    from .models import ScholarListImport, ActivityLog, SystemSettings, ImportedScholar
    if request.method != 'POST':
        return redirect('/vpsea/archives/')
    stype = request.POST.get('type', 'Academic')
    rollover_label = request.POST.get('rollover_label', '').strip()
    file = request.FILES.get('file')
    if not file:
        return redirect(f'/vpsea/archives/?type={stype}&import_error=No+file+provided')
    if not rollover_label:
        return redirect(f'/vpsea/archives/?type={stype}&import_error=Rollover+name+is+required')
    try:
        settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
        parsed = SystemSettings.parse_label(settings_obj.academic_year)
        active_semester = parsed['semester']
        rollover_parsed = SystemSettings.parse_label(rollover_label) if '-' in rollover_label else {'sy': rollover_label, 'semester': active_semester}

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        col_map = COLUMN_MAPS.get(stype, COLUMN_MAPS['CoScho'])

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                int(row[0])
            except (ValueError, TypeError):
                continue

            extra = {}
            for idx, field in col_map:
                val = row[idx] if idx < len(row) else None
                extra[field] = str(val).strip() if val is not None else ''

            last = extra.get('last_name', '').strip()
            first = extra.get('first_name', '').strip()
            if not last and not first:
                continue

            try:
                year_level = int(extra.get('year', 0) or 0)
            except (ValueError, TypeError):
                year_level = 0
            try:
                gwa = float(extra.get('gwa', 0) or 0)
            except (ValueError, TypeError):
                gwa = 0.0

            address = extra.get('address', '')
            addr_parts = [x.strip() for x in address.split(',')] if address else []

            records.append(ImportedScholar(
                scholarship_type=stype,
                term_label=rollover_label,
                last_name=last,
                first_name=first,
                middle_name=extra.get('middle_name', extra.get('middle_initial', '')),
                gender=extra.get('sex', ''),
                course=extra.get('course', ''),
                year_level=year_level,
                gwa=gwa,
                barangay=extra.get('barangay', addr_parts[0] if len(addr_parts) > 0 else ''),
                municipality=extra.get('municipality', addr_parts[1] if len(addr_parts) > 1 else ''),
                province=extra.get('province', addr_parts[2] if len(addr_parts) > 2 else ''),
                student_id=extra.get('student_number', extra.get('student_id', '')),
                award_number=extra.get('award_number', ''),
                congress_district=extra.get('congress_district', ''),
                imported_from=file.name,
            ))

        # Replace the term's rows only once the new ones are in hand, and in one
        # transaction. The delete used to run before the sheet was parsed, so a
        # file the parser choked on destroyed the term and imported nothing.
        with transaction.atomic():
            ImportedScholar.objects.filter(
                scholarship_type=stype, term_label=rollover_label).delete()
            ImportedScholar.objects.bulk_create(records)
        created = len(records)

        # Save the uploaded file as a rollover record for download history
        file.seek(0)
        if not ScholarListImport.objects.filter(term_label=rollover_label, scholarship_type=stype).exists():
            rollover = ScholarListImport(
                scholarship_type=stype,
                school_year=rollover_parsed['sy'],
                semester=rollover_parsed.get('semester', active_semester),
                term_label=rollover_label,
                scholar_count=created,
                imported_by=request.user,
            )
            rollover.excel_file.save(f'{stype}_{rollover_label}.xlsx', ContentFile(file.read()), save=True)
        else:
            ScholarListImport.objects.filter(term_label=rollover_label, scholarship_type=stype).update(scholar_count=created)

        ActivityLog.objects.create(
            user=request.user,
            action=f'Imported {file.name} ({created} rows) for {stype} as "{rollover_label}"'
        )
    except Exception as e:
        return redirect(f'/vpsea/archives/?type={stype}&import_error={e}')
    return redirect(f'/vpsea/archives/?type={stype}&import_ok={created}')



@_vpsea_required
def vpsea_archive_download(request):
    from .models import Application, AffirmativeStaffApplication, SystemSettings, AcademicRenewal
    stype = request.GET.get('type', 'Academic')
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    school_year = settings_obj.academic_year
    semester = settings_obj.active_semester

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{stype} Scholars'
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hfont = Font(bold=True)
    hfill = PatternFill('solid', fgColor='D9E1F2')

    def hrow(ws, r, n):
        for c in range(1, n+1):
            cell = ws.cell(row=r, column=c)
            cell.font = hfont; cell.fill = hfill; cell.border = border; cell.alignment = center

    def drow(ws, r, vals):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    def name_parts(user):
        return user.last_name or '', user.first_name or '', ''

    def split_name(full):
        p = full.strip().split()
        if len(p) == 0: return '', '', ''
        if len(p) == 1: return p[0], '', ''
        if len(p) == 2: return p[-1], p[0], ''
        return p[-1], p[0], ' '.join(p[1:-1])[0] + '.'

    renewed_ids = AcademicRenewal.objects.filter(status='Approved').values_list('student_id', flat=True)

    row = 1
    if stype in ('Affirmative', 'Staff'):
        scholars = AffirmativeStaffApplication.objects.filter(
            status='Approved', qualified_for=stype).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name')
        if stype == 'Staff':
            headers = ['No.', 'Last', 'First', 'M.I.', 'Sex', 'Course', 'Year Level', 'Student No.', '%', 'Scholarship Program']
        else:
            headers = ['No.', 'Last Name', 'First Name', 'Middle Name', 'Sex', 'Address', 'Course', 'Yr.', 'Scholarship Program']
        ws.append(headers); hrow(ws, row, len(headers)); row += 1
        for i, app in enumerate(scholars, 1):
            last, first, mi = split_name(app.full_name)
            if stype == 'Staff':
                drow(ws, row, [i, last, first, mi, app.gender, app.course, app.year_level, app.student_id, '100%' if app.is_nsu_staff else '75%', 'BiPSU Staff Scholarship'])
            else:
                drow(ws, row, [i, last, first, mi, app.gender, app.address, app.course, app.year_level, 'Affirmative Action Scholarship'])
            row += 1
    elif stype == 'Academic':
        scholars = Application.objects.filter(status='Approved', scholarship__type='Academic', student_id__in=renewed_ids).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')
        headers = ['No.', 'Last Name', 'First', 'Middle Name', 'Sex', 'Brgy./St.', 'Municipality', 'Province', 'Course', 'Yr', 'GWA', '% / Type', 'Scholarship']
        ws.append(headers); hrow(ws, row, len(headers)); row += 1
        for i, app in enumerate(scholars, 1):
            p = app.student; last, first, mi = name_parts(p.user)
            pct = 'University Scholar' if p.gwa <= 1.29 else ('College Scholar' if p.gwa <= 1.50 else '')
            drow(ws, row, [i, last, first, mi, p.gender, p.barangay, p.municipality, p.province, p.course, p.year_level, p.gwa, pct, app.scholarship.name]); row += 1
    else:
        scholars = Application.objects.filter(status='Approved', scholarship__type=stype, student_id__in=renewed_ids).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')
        headers = ['No.', 'Last Name', 'First Name', 'Middle Name', 'Sex', 'Brgy./St.', 'Municipality', 'Province', 'Course', 'Year', 'Student No.', 'Scholarship Program']
        ws.append(headers); hrow(ws, row, len(headers)); row += 1
        for i, app in enumerate(scholars, 1):
            p = app.student; last, first, mi = name_parts(p.user)
            drow(ws, row, [i, last, first, mi, p.gender, p.barangay, p.municipality, p.province, p.course, p.year_level, p.student_id, app.scholarship.name]); row += 1

    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f'{stype}_scholars_{school_year}_{semester.replace(" ", "_")}.xlsx'
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



# What a scholar is counted under when no school is recorded and their course
# matches none of BiPSU's. Named rather than left blank so it reads as a gap in
# the record on the chart, which is what it is.
UNRECORDED_SCHOOL = 'Not recorded'


def _analytics_context(request, all_types, include_gwa=True):
    """Build the analytics context for a given set of scholarship types.

    Shared by the VPSEA portal (every type) and the UniFAST portal (TES and TDP
    only), so both read the same numbers from the same code path.
    """
    from .models import ScholarListImport, SystemSettings
    import openpyxl
    from collections import defaultdict

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active_label = settings_obj.academic_year
    ALL_TYPES = list(all_types)

    from .models import ImportedScholar

    def _imported_current(stype):
        """Imported scholars for the active term, the ones a live count misses.

        Claimed rows are excluded: that scholar has an Application of their own,
        and counting both would show them twice. Same rule the archive tables
        use.
        """
        return ImportedScholar.objects.filter(
            scholarship_type=stype, term_label=active_label, claimed_by__isnull=True,
        )

    all_labels = list(
        ScholarListImport.objects.values_list('term_label', flat=True)
        .distinct().order_by('-term_label')
    )
    # Also include labels that only exist in ImportedScholar (import-only semesters)
    ar_labels = list(
        ImportedScholar.objects.exclude(term_label='')
        .values_list('term_label', flat=True).distinct()
    )
    for lbl in ar_labels:
        if lbl not in all_labels:
            all_labels.append(lbl)
    all_labels = sorted(set(all_labels), reverse=True)
    if active_label not in all_labels:
        all_labels.insert(0, active_label)
    all_sy_display = [(lbl, f"{SystemSettings.parse_label(lbl)['sy']} — {SystemSettings.parse_label(lbl)['semester']}") for lbl in all_labels]

    selected_label = request.GET.get('sy', active_label)
    if selected_label not in all_labels:
        selected_label = active_label
    selected_parsed = SystemSettings.parse_label(selected_label)
    selected_sy = selected_parsed['sy']
    selected_semester = selected_parsed['semester']
    selected_type = request.GET.get('stype', '')

    # Scholar count per type: rollover records + ImportedScholar imports
    from .models import ImportedScholar
    # Scholar count per type: if selected = active semester, use live DB counts
    # otherwise use ImportedScholar imports or ScholarListImport snapshot
    rollover_counts = {}
    for t in ALL_TYPES:
        if selected_label == active_label:
            # Current semester — no rollover yet, count from live approved records
            if t in ('Affirmative', 'Staff'):
                from .models import AffirmativeStaffApplication
                rollover_counts[t] = (
                    AffirmativeStaffApplication.objects.filter(
                        status='Approved', qualified_for=t
                    ).count()
                    + _imported_current(t).count()
                )
            else:
                # Imported scholars count too. Only the past-semester branch
                # below ever looked at them, so a list uploaded into the active
                # term showed nothing here until the term rolled over.
                rollover_counts[t] = (
                    Application.objects.filter(
                        status='Approved', scholarship__type=t
                    ).count()
                    + _imported_current(t).count()
                )
        else:
            # Past semester — prefer ImportedScholar rows, fall back to rollover snapshot
            import_count = ImportedScholar.objects.filter(scholarship_type=t, term_label=selected_label).count()
            if import_count:
                rollover_counts[t] = import_count
            else:
                r = ScholarListImport.objects.filter(scholarship_type=t, term_label=selected_label).first()
                rollover_counts[t] = r.scholar_count if r else 0

    def _course_counts_from_rollover(stype):
        # Current semester — pull from live approved records
        if selected_label == active_label:
            from django.db.models import Count as DCount
            counts = {}
            if stype in ('Affirmative', 'Staff'):
                from .models import AffirmativeStaffApplication
                qs = AffirmativeStaffApplication.objects.filter(
                    status='Approved', qualified_for=stype
                ).values('enrollment__course').annotate(n=DCount('id'))
                for r in qs:
                    key = r['enrollment__course'] or 'Unknown'
                    counts[key] = counts.get(key, 0) + r['n']
            else:
                qs = Application.objects.filter(
                    status='Approved', scholarship__type=stype
                ).values('student__enrollment__course').annotate(n=DCount('id'))
                for r in qs:
                    key = r['student__enrollment__course'] or 'Unknown'
                    counts[key] = counts.get(key, 0) + r['n']
            # …and the imported rows for the same term, which this branch used
            # to leave out entirely.
            for r in _imported_current(stype).values('course').annotate(n=DCount('id')):
                key = r['course'] or 'Unknown'
                counts[key] = counts.get(key, 0) + r['n']
            return counts
        # Past semester — first try ImportedScholar (imported rows)
        ar_counts = {}
        for rec in ImportedScholar.objects.filter(scholarship_type=stype, term_label=selected_label).values('course'):
            c = rec['course'] or 'Unknown'
            ar_counts[c] = ar_counts.get(c, 0) + 1
        if ar_counts:
            return ar_counts
        # Fall back to rollover excel file
        r = ScholarListImport.objects.filter(scholarship_type=stype, term_label=selected_label).first()
        if not r or not r.excel_file:
            return {}
        try:
            wb = openpyxl.load_workbook(r.excel_file.path)
            ws = wb.active
            course_col = next((cell.column - 1 for cell in ws[1] if cell.value and 'course' in str(cell.value).lower()), None)
            if course_col is None:
                return {}
            counts = defaultdict(int)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] is not None and course_col < len(row) and row[course_col]:
                    counts[str(row[course_col]).strip()] += 1
            return dict(counts)
        except Exception:
            return {}

    def _school_counts(stype):
        """Scholars per BiPSU school for one programme.

        Only an award and a staff application record a school of their own. An
        imported row and a rollover sheet carry a course and nothing else, so
        theirs is worked out from it — and `school_for_course` matches a course
        exactly or not at all, so a scholar whose course was typed free-hand is
        reported as unrecorded rather than filed under a school someone guessed.
        """
        from .constants import school_for_course

        counts = defaultdict(int)

        def add(school, course, n=1):
            named = (school or '').strip()
            counts[named or school_for_course(course) or UNRECORDED_SCHOOL] += n

        if selected_label == active_label:
            if stype in ('Affirmative', 'Staff'):
                from .models import AffirmativeStaffApplication
                rows = AffirmativeStaffApplication.objects.filter(
                    status='Approved', qualified_for=stype
                ).values('enrollment__school', 'enrollment__course')
                for r in rows:
                    add(r['enrollment__school'], r['enrollment__course'])
            else:
                rows = Application.objects.filter(
                    status='Approved', scholarship__type=stype
                ).values('student__enrollment__school', 'student__enrollment__course')
                for r in rows:
                    add(r['student__enrollment__school'],
                        r['student__enrollment__course'])
            for r in _imported_current(stype).values('course'):
                add('', r['course'])
            return dict(counts)

        # A past semester is read from imported rows or the uploaded sheet, and
        # neither carries a school — so the course counts already built for that
        # term are rolled up rather than queried a second time.
        for course, n in _course_counts_from_rollover(stype).items():
            add('', '' if course == 'Unknown' else course, n)
        return dict(counts)

    # School distribution
    if selected_type and selected_type in ALL_TYPES:
        raw_schools = _school_counts(selected_type)
    else:
        raw_schools = defaultdict(int)
        for t in ALL_TYPES:
            for k, v in _school_counts(t).items():
                raw_schools[k] += v
    school_dist = [{'school': k, 'scholars': v}
                   for k, v in sorted(raw_schools.items(), key=lambda x: -x[1])]

    # Course distribution
    if selected_type and selected_type in ALL_TYPES:
        raw = _course_counts_from_rollover(selected_type)
    else:
        raw = defaultdict(int)
        for t in ALL_TYPES:
            for k, v in _course_counts_from_rollover(t).items():
                raw[k] += v
    course_dist = [{'course': k, 'scholars': v} for k, v in sorted(raw.items(), key=lambda x: -x[1])]

    # GWA distribution — current sem: live DB; past sem: ImportedScholar or rollover excel
    gpa_ranges = [{'range': r, 'count': 0} for r in ['1.00-1.25', '1.26-1.50', '1.51-1.75', '1.76-2.00', '2.01-2.50']]
    if not include_gwa:
        # TES and TDP are needs-based — they are not banded by GWA.
        gpa_ranges = []
    elif selected_label == active_label:
        buckets = {'1.00-1.25': 0, '1.26-1.50': 0, '1.51-1.75': 0, '1.76-2.00': 0, '2.01-2.50': 0}
        for p in Application.objects.filter(
            status='Approved', scholarship__type='Academic'
        ).values('student__enrollment__gwa'):
            g = p['student__enrollment__gwa'] or 0
            if 1.0 <= g <= 1.25: buckets['1.00-1.25'] += 1
            elif g <= 1.50: buckets['1.26-1.50'] += 1
            elif g <= 1.75: buckets['1.51-1.75'] += 1
            elif g <= 2.00: buckets['1.76-2.00'] += 1
            elif g <= 2.50: buckets['2.01-2.50'] += 1
        gpa_ranges = [{'range': k, 'count': v} for k, v in buckets.items()]
    else:
        ar_academic = ImportedScholar.objects.filter(scholarship_type='Academic', term_label=selected_label)
        if ar_academic.exists():
            buckets = {'1.00-1.25': 0, '1.26-1.50': 0, '1.51-1.75': 0, '1.76-2.00': 0, '2.01-2.50': 0}
            for rec in ar_academic.values('gwa'):
                g = rec['gwa'] or 0
                if 1.0 <= g <= 1.25: buckets['1.00-1.25'] += 1
                elif g <= 1.50: buckets['1.26-1.50'] += 1
                elif g <= 1.75: buckets['1.51-1.75'] += 1
                elif g <= 2.00: buckets['1.76-2.00'] += 1
                elif g <= 2.50: buckets['2.01-2.50'] += 1
            gpa_ranges = [{'range': k, 'count': v} for k, v in buckets.items()]
        else:
            acad_r = ScholarListImport.objects.filter(scholarship_type='Academic', term_label=selected_label).first()
            if acad_r and acad_r.excel_file:
                try:
                    wb = openpyxl.load_workbook(acad_r.excel_file.path)
                    ws = wb.active
                    gwa_col = next((cell.column - 1 for cell in ws[1] if cell.value and 'gwa' in str(cell.value).lower()), None)
                    if gwa_col is not None:
                        buckets = {'1.00-1.25': 0, '1.26-1.50': 0, '1.51-1.75': 0, '1.76-2.00': 0, '2.01-2.50': 0}
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if row and row[0] is not None:
                                try:
                                    g = float(row[gwa_col]) if gwa_col < len(row) and row[gwa_col] else 0
                                    if 1.0 <= g <= 1.25: buckets['1.00-1.25'] += 1
                                    elif g <= 1.50: buckets['1.26-1.50'] += 1
                                    elif g <= 1.75: buckets['1.51-1.75'] += 1
                                    elif g <= 2.00: buckets['1.76-2.00'] += 1
                                    elif g <= 2.50: buckets['2.01-2.50'] += 1
                                except (ValueError, TypeError):
                                    pass
                        gpa_ranges = [{'range': k, 'count': v} for k, v in buckets.items()]
                except Exception:
                    pass

    # ── Scholars-over-time trend ──────────────────────────────────────────────
    # Build a chronological list of (label, display, total_scholars) covering
    # every known semester so the line chart spans the full history.
    # Labels use the compact format "YY-S" (e.g. "25-1") — sort numerically.
    def _label_sort_key(lbl):
        try:
            yy, s = lbl.split('-')
            return int(yy) * 10 + int(s)
        except Exception:
            return 0

    trend_labels_sorted = sorted(set(all_labels), key=_label_sort_key)

    trend_data = []
    for lbl in trend_labels_sorted:
        parsed = SystemSettings.parse_label(lbl)
        display = f"{parsed['sy']}\n{parsed['semester']}"   # two-line label for x-axis
        short   = f"{parsed['sy']} S{parsed['sy_start'] and lbl.split('-')[1] or '?'}"

        # One pass per semester. The total and the per-programme breakdown are
        # the same counts, and this used to run every one of these queries twice
        # to produce both.
        counts = {}
        for t in ALL_TYPES:
            if lbl == active_label:
                if t in ('Affirmative', 'Staff'):
                    from .models import AffirmativeStaffApplication
                    c = AffirmativeStaffApplication.objects.filter(
                        status='Approved', qualified_for=t
                    ).count()
                else:
                    c = Application.objects.filter(
                        status='Approved', scholarship__type=t
                    ).count()
            else:
                # Prefer ImportedScholar row count, fall back to rollover snapshot
                c = ImportedScholar.objects.filter(
                    scholarship_type=t, term_label=lbl
                ).count()
                if not c:
                    r = ScholarListImport.objects.filter(
                        scholarship_type=t, term_label=lbl
                    ).first()
                    c = r.scholar_count if r else 0
            counts[t] = c

        parsed_display = f"{parsed['sy']} — {parsed['semester']}"
        trend_data.append({
            'label': lbl,
            'display': parsed_display,
            'total': sum(counts.values()),
            'counts': counts,
            'per_type': {t: c for t, c in counts.items() if c},
        })

    # One line per programme rather than a single blended total: the total is
    # what the summary tiles already say, while the comparison between
    # programmes is what this chart is for. A programme with no scholars in a
    # semester contributes a zero rather than a gap, so the lines stay aligned;
    # a programme with no scholars in any semester is left out entirely.
    # Filtering the page to one programme draws that one alone.
    if selected_type and selected_type in ALL_TYPES:
        series_types = [selected_type]
    else:
        series_types = [t for t in ALL_TYPES
                        if any(d['counts'].get(t) for d in trend_data)]
    trend_series = [
        {'type': t, 'counts': [d['counts'].get(t, 0) for d in trend_data]}
        for t in series_types
    ]

    return {
        'rollover_counts': rollover_counts,
        'all_types': ALL_TYPES,
        'course_dist': course_dist,
        'school_dist': school_dist,
        'gpa_ranges': gpa_ranges,
        'all_sy_display': all_sy_display,
        'selected_sy': selected_label,
        'selected_type': selected_type,
        'selected_sy_display': f"{selected_sy} — {selected_semester}",
        'active_sy': active_label,
        'trend_data': trend_data,
        'trend_series': trend_series,
    }


@_vpsea_required
def vpsea_analytics(request):
    """Every scholarship type, GWA distribution included."""
    _base = ['Academic', 'TDP', 'DOST', 'CHED', 'CoScho', 'Sports', 'Affirmative', 'Staff', 'GSIS']
    all_types = _base + [
        t for t in Scholarship.objects.values_list('type', flat=True).distinct()
        if t not in _base
    ]
    return render(request, 'vpsea/analytics.html', _analytics_context(request, all_types))


@_unifast_required
def unifast_analytics(request):
    """Scoped to the programmes UniFAST administers: TES and TDP."""
    ctx = _analytics_context(request, UNIFAST_TYPES, include_gwa=False)
    return render(request, 'unifast/analytics.html', ctx)


@_unifast_required
def unifast_announcements(request):
    from .models import Announcement
    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        body = (request.POST.get('body') or '').strip()
        if title and body:
            Announcement.objects.create(title=title, body=body, published_by=request.user)
            reached = notify.broadcast(title, body)
            return redirect(f'/unifast/announcements/?posted={reached}')
        return render(request, 'unifast/announcements.html', {
            'announcements': Announcement.objects.select_related('published_by').order_by('-created_at'),
            'errors': ['Both a title and a body are required.'],
            'post': request.POST,
        })
    return render(request, 'unifast/announcements.html', {
        'announcements': Announcement.objects.select_related('published_by').order_by('-created_at'),
        'posted': request.GET.get('posted'),
    })


UNIFAST_PROGRAMMES = (
    ('TDP', 'TDP — TULONG DUNONG PROGRAM'),
    ('TES', 'TES — TERTIARY EDUCATION SUBSIDY'),
)


def _unifast_report_sections(school_year='', types=None):
    """Approved TES and TDP scholars, split by gender, in masterlist order.

    Mirrors the VPSEA report layout but covers only the two programmes UniFAST
    administers. Both are sourced from approved Applications — TES rows are
    created when a TES application is approved in this portal.

    ``types`` narrows it to one programme, which is how the TDP report is built
    from the same rows the combined masterlist uses; ``school_year`` scopes it
    to one term, blank meaning every one of them.
    """
    def _row(app):
        p = app.student
        u = p.user
        return {
            'last': u.last_name, 'first': u.first_name, 'mi': '',
            'sex': p.gender or '', 'brgy': p.barangay, 'mun': p.municipality,
            'prov': p.province, 'course': p.course, 'yr': p.year_level,
            'student_no': p.student_id,
            'scholarship': app.scholarship.name,
            'award': app.award_number,
            'cong': app.congress_district,
        }

    def _split_gender(apps):
        female = [_row(a) for a in apps if (a.student.gender or '').upper() in ('F', 'FEMALE')]
        male = [_row(a) for a in apps if (a.student.gender or '').upper() not in ('F', 'FEMALE')]
        return female, male

    headers = ['NO.', 'AWARD NO.', 'LAST NAME', 'FIRST NAME', 'M.I.', 'SEX',
               'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'STUDENT NO.',
               'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']

    sections = []
    for stype, title in UNIFAST_PROGRAMMES:
        if types and stype not in types:
            continue
        qs = (
            Application.objects.filter(status='Approved', scholarship__type=stype)
            .select_related('student__user', 'scholarship', *STUDENT_DETAILS)
            .order_by('student__user__last_name')
        )
        if school_year:
            qs = qs.filter(school_year=school_year)
        apps = list(qs)
        female, male = _split_gender(apps)
        sections.append({
            'title': title, 'key': stype.lower(), 'headers': headers,
            'groups': [('FEMALE', female), ('MALE', male)],
            'female_rows': female, 'male_rows': male,
            'total': len(apps),
        })
    return sections


@_unifast_required
def unifast_reports(request):
    """On-screen replica of the CHED 'Official List' validation sheet."""
    import os
    from .models import SystemSettings
    from . import annex1_report, doc_convert, tes_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    # No batch is asked for here any more. It is a CHED bookkeeping label, not
    # something the office filters its own reports by, and the workbook's own
    # header already reads 'On-going' when none is given. The download endpoint
    # still honours ?batch= for anyone who needs to stamp one.
    #
    # Which school year the office is generating for. Blank means every year,
    # which is what this page reported before the year could be picked — so an
    # existing deployment sees the same list until someone chooses one.
    school_year = request.GET.get('sy', '').strip()
    rows = tes_report.grantee_rows(school_year=school_year)
    sections = _unifast_report_sections(school_year=school_year)
    tdp = next((s for s in sections if s['key'] == 'tdp'), None)

    return render(request, 'unifast/reports.html', {
        'rows': rows,
        'headers': tes_report.OFFICIAL_LIST_HEADERS,
        'semester': parsed['semester'],
        # TDP gets the same treatment as TES: its own count, its own download
        # and its own preview frame, off the same school year.
        'tdp_total': tdp['total'] if tdp else 0,
        'tdp_female': len(tdp['female_rows']) if tdp else 0,
        'tdp_male': len(tdp['male_rows']) if tdp else 0,
        # What the report prints over the list: the year that was picked, or
        # the active term when the list spans all of them.
        'ay': school_year or parsed['sy'],
        'school_year': school_year,
        'school_years': tes_report.school_year_options(),
        'tes_total': len(rows),
        'pwd_count': sum(1 for r in rows if r['is_pwd']),
        'template_available': os.path.exists(tes_report.TEMPLATE_PATH),
        # Whether the frame is showing the workbook itself or the fallback
        # layout, so the page can say which one an officer is reading.
        'exact_preview': doc_convert.available(),
        # The plain TDP/TES masterlist stays available underneath.
        'sections': sections,
        'grand_total': sum(s['total'] for s in sections),
        'error': request.GET.get('error'),
    })


@_unifast_required
def unifast_billing(request):
    """What CHED pays each TES grantee, and the statement that follows from it.

    The Annex 2 workbook bills on two per-grantee figures and computes the rest
    — the column totals, the 1% management fee, the Form 1 statement. Those
    columns exported blank until this page existed, because a rate the system
    chose for itself would have been a guess an officer then signed. The rate is
    still not the system's to choose; this is where an officer records the one
    CHED gave them.
    """
    from .models import SystemSettings, TESBilling
    from . import tes_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active = SystemSettings.parse_label(settings_obj.academic_year)['sy']
    # A rate belongs to a term — CHED revises it, and last year's figure must
    # not quietly bill this year's grantees — so unlike the report pages there
    # is no 'all school years' here.
    school_year = request.GET.get('sy', '').strip() or active

    errors = []
    if request.method == 'POST':
        school_year = request.POST.get('sy', '').strip() or active
        amounts = {}
        for field, label in (('tes_amount', 'TES amount'),
                             ('tes_3a_amount', 'TES-3A amount')):
            raw = (request.POST.get(field) or '').strip().replace(',', '')
            if not raw:
                amounts[field] = None
                continue
            try:
                value = Decimal(raw)
            except InvalidOperation:
                errors.append(f'{label} must be a number.')
                continue
            if value < 0:
                errors.append(f'{label} cannot be negative.')
                continue
            amounts[field] = value

        if not errors:
            TESBilling.objects.update_or_create(
                school_year=school_year,
                defaults=dict(
                    reference_no=(request.POST.get('reference_no') or '').strip(),
                    statement_date=request.POST.get('statement_date') or None,
                    updated_by=request.user,
                    **amounts,
                ),
            )
            return redirect(f'/unifast/billing/?sy={school_year}&saved=1')

    summary = tes_report.billing_summary(school_year)
    saved_row = summary['billing']

    # What the boxes show. Built here rather than juggling template filters: on
    # a GET `request.POST` is empty, and an empty string is not None, so a
    # `default_if_none` fallback never fires and the saved figures render blank.
    def shown(field, stored):
        if request.method == 'POST':
            return request.POST.get(field, '')
        return '' if stored is None else stored

    form = {
        'tes_amount': shown('tes_amount', saved_row.tes_amount if saved_row else None),
        'tes_3a_amount': shown('tes_3a_amount',
                               saved_row.tes_3a_amount if saved_row else None),
        'reference_no': shown('reference_no',
                              saved_row.reference_no if saved_row else ''),
        'statement_date': shown(
            'statement_date',
            saved_row.statement_date.strftime('%Y-%m-%d')
            if saved_row and saved_row.statement_date else ''),
    }

    return render(request, 'unifast/billing.html', dict(summary, **{
        'school_year': school_year,
        'school_years': tes_report.school_year_options(),
        'semester': SystemSettings.parse_label(settings_obj.academic_year)['semester'],
        'errors': errors,
        'saved': request.GET.get('saved'),
        'form': form,
    }))


def _decimal_or_error(raw, label, errors):
    """A money field as typed by a person, or None with the reason recorded.

    Blank is None rather than zero, and stays that way all the way to the
    column — the distinction TESBilling and TESLiquidation are both built on.
    Thousands separators are stripped because a cashier reading off a voucher
    types them.
    """
    raw = (raw or '').strip().replace(',', '')
    if not raw:
        return None
    try:
        value = Decimal(raw)
    except InvalidOperation:
        errors.append(f'{label} must be a number.')
        return None
    if value < 0:
        errors.append(f'{label} cannot be negative.')
        return None
    return value


@_unifast_required
def unifast_liquidation(request):
    """What CHED's money actually did, once it arrived.

    The Billing tab states what the office asked CHEDRO for. This is the other
    half: what was remitted, what the cashier paid out, and what is still in
    hand. Those are separate facts and they routinely differ — a grantee who
    never collects is the ordinary case — so they are separate records rather
    than one number doing both jobs.

    The grantee list is not maintained here. It is exactly the list the billing
    and the CHED workbook use, so the report can only ever account for the
    people the office actually billed for; the page adds what became of each
    one's share.

    Nothing is defaulted from the billed rate. The billed rate is what was asked
    for, and a liquidation that assumes everyone was paid it cannot detect the
    thing it exists to detect. The 'Fill in the billed amount' button on the
    page fills the boxes in the browser and saves nothing — an officer still
    reads the rows and submits them.
    """
    from django.utils import timezone

    from .constants import TES_DISBURSEMENT_STATUSES, TES_RELEASED
    from .models import SystemSettings, TESDisbursement, TESLiquidation
    from . import tes_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active = SystemSettings.parse_label(settings_obj.academic_year)['sy']
    # Scoped to one term for the same reason billing is: money is released
    # against a school year, and a balance that quietly spans two of them
    # reconciles to a number nobody can explain to an auditor.
    school_year = request.GET.get('sy', '').strip() or active

    errors = []
    if request.method == 'POST':
        school_year = request.POST.get('sy', '').strip() or active
        action = request.POST.get('action') or 'remittance'

        if action == 'remittance':
            received = _decimal_or_error(
                request.POST.get('funds_received'), 'Funds received', errors)
            if not errors:
                TESLiquidation.objects.update_or_create(
                    school_year=school_year,
                    defaults={
                        'funds_received': received,
                        'received_date': request.POST.get('received_date') or None,
                        'credit_advice_no': (request.POST.get('credit_advice_no') or '').strip(),
                        'report_no': (request.POST.get('report_no') or '').strip(),
                        'report_date': request.POST.get('report_date') or None,
                        'updated_by': request.user,
                    },
                )
                return redirect(f'/unifast/liquidation/?sy={school_year}&saved=1')

        elif action == 'disbursements':
            # The grantee list is read from the database, not from the form. A
            # row posted for somebody who is not an approved grantee of this
            # term records nothing, however the request was assembled.
            rows = tes_report.liquidation_rows(school_year)
            allowed = [row['app_id'] for row in rows]
            valid = {value for value, _ in TES_DISBURSEMENT_STATUSES}

            entries = []
            bad_status = False
            for app_id in allowed:
                status = (request.POST.get(f'status-{app_id}') or '').strip()
                if not status:
                    continue
                if status not in valid:
                    bad_status = True
                    continue
                amount = _decimal_or_error(
                    request.POST.get(f'amount-{app_id}'), 'Amount released', errors)
                # A release with no figure on it is not a release, it is a
                # half-finished entry — and it would report as one grantee paid
                # nothing, which reads as a payment of zero rather than as the
                # omission it is. Every other status carries no amount by
                # design, so only this one has to insist.
                if status == TES_RELEASED and amount is None:
                    errors.append('A released row needs the amount that was paid.')
                entries.append((app_id, status, amount,
                                request.POST.get(f'date-{app_id}') or None,
                                (request.POST.get(f'receipt-{app_id}') or '').strip()))

            if bad_status:
                errors.append('One of the rows carried a status this office does not use.')

            # These are per-row faults on a form of up to a thousand rows, so
            # the same sentence can be reached many times over. Said once each,
            # in the order they were first hit: ten identical lines tell the
            # officer nothing the first one did not.
            errors[:] = list(dict.fromkeys(errors))

            if not errors:
                # A liquidation row has to exist for the disbursements to hang
                # off. Created empty here rather than refusing the save: the
                # office often pays grantees before the credit advice is filed,
                # and losing that work to a validation error would be its own
                # small disaster.
                liquidation, _ = TESLiquidation.objects.get_or_create(
                    school_year=school_year,
                    defaults={'updated_by': request.user},
                )
                with transaction.atomic():
                    for app_id, status, amount, date, receipt in entries:
                        TESDisbursement.objects.update_or_create(
                            liquidation=liquidation, tes_application_id=app_id,
                            defaults={
                                'status': status,
                                # Only a release moves money. A row switched
                                # back to Unclaimed keeps no amount, or the
                                # totals would go on counting it.
                                'amount_released': amount if status == TES_RELEASED else None,
                                'date_released': date if status == TES_RELEASED else None,
                                'receipt_no': receipt,
                                'updated_by': request.user,
                            },
                        )
                return redirect(f'/unifast/liquidation/?sy={school_year}&saved=rows')

    summary = tes_report.liquidation_summary(school_year)
    saved_row = summary['liquidation']

    # Same reason the Billing tab builds this here: on a GET request.POST is
    # empty, and '' is not None, so a template-level default never fires and the
    # saved figures would render blank.
    def shown(field, stored):
        if request.method == 'POST':
            return request.POST.get(field, '')
        return '' if stored is None else stored

    def shown_date(field, stored):
        return shown(field, stored.strftime('%Y-%m-%d') if stored else None)

    form = {
        'funds_received': shown('funds_received',
                                saved_row.funds_received if saved_row else None),
        'credit_advice_no': shown('credit_advice_no',
                                  saved_row.credit_advice_no if saved_row else ''),
        'report_no': shown('report_no', saved_row.report_no if saved_row else ''),
        'received_date': shown_date('received_date',
                                    saved_row.received_date if saved_row else None),
        'report_date': shown_date('report_date',
                                  saved_row.report_date if saved_row else None),
    }

    return render(request, 'unifast/liquidation.html', dict(summary, **{
        'school_year': school_year,
        'school_years': tes_report.school_year_options(),
        'semester': SystemSettings.parse_label(settings_obj.academic_year)['semester'],
        'statuses': TES_DISBURSEMENT_STATUSES,
        # What the 'billed amount' button puts in an empty date box. A prefill,
        # not a record: it reaches the database only if the officer saves.
        'today': timezone.localdate().strftime('%Y-%m-%d'),
        'errors': errors,
        'saved': request.GET.get('saved'),
        'form': form,
    }))


@_unifast_required
@xframe_options_exempt
def unifast_report_preview_pdf(request):
    """The CHED workbook as a PDF page, for the preview frame on the Reports tab.

    The Annex 2 workbook the office submits is filled and converted, so all four
    sheets preview exactly as they print. Without a converter installed the
    Official List is laid out here instead, and the tab says so.
    """
    from .models import SystemSettings
    from . import doc_convert, report_pdf, tes_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    batch = request.GET.get('batch', '').strip()
    school_year = request.GET.get('sy', '').strip()
    ay = school_year or parsed['sy']
    label = (school_year or settings_obj.academic_year).replace('-', '_')

    pdf = None
    if doc_convert.available():
        try:
            buf, _written, _overflow = tes_report.build_workbook(
                ay, parsed['semester'], batch=batch, school_year=school_year)
            pdf = doc_convert.to_pdf(buf.getvalue(), '.xlsx')
        except (FileNotFoundError, doc_convert.ConversionUnavailable,
                doc_convert.ConversionFailed):
            pdf = None
    if pdf is None:
        buf, _rows = report_pdf.tes_official_list_pdf(
            ay, parsed['semester'], batch=batch, school_year=school_year)
        pdf = buf.read()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="TES_Validation_List_{label}.pdf"')
    return response


@_unifast_required
def unifast_report_download_tes(request):
    """The filled CHED TES validation & billing workbook, all four sheets."""
    from .models import SystemSettings
    from . import tes_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    batch = request.GET.get('batch', '').strip()
    school_year = request.GET.get('sy', '').strip()
    ay = school_year or parsed['sy']
    try:
        buf, written, overflow = tes_report.build_workbook(
            ay, parsed['semester'], batch=batch, school_year=school_year)
    except FileNotFoundError as exc:
        from urllib.parse import quote
        return redirect(f'/unifast/reports/?error={quote(str(exc))}')

    label = (school_year or settings_obj.academic_year).replace('-', '_')
    filename = f'TES_Validation_Billing_{label}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    if overflow:
        response['X-TES-Overflow'] = str(overflow)
    return response


def _unifast_scholars_workbook(sections, ay, semester, sheet_title, banner_text):
    """The UniFAST scholars masterlist as a styled workbook.

    Shared by the combined TDP/TES download and the TDP-only one, so a change to
    the layout reaches both rather than one of them drifting.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    row_at = [1]

    def banner(text, ncols, fill, font, height=None):
        r = row_at[0]
        ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1)
        cell.font, cell.fill, cell.alignment, cell.border = font, fill, center, border
        if height:
            ws.row_dimensions[r].height = height
        row_at[0] += 1

    def label(text, ncols):
        r = row_at[0]
        ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        ws.cell(row=r, column=1).font = Font(bold=True, size=9)
        row_at[0] += 1

    def table(headers, rows):
        r = row_at[0]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
            cell.alignment, cell.border = center, border
        row_at[0] += 1
        for i, row in enumerate(rows, 1):
            r = row_at[0]
            values = [i, row['award'], row['last'], row['first'], row['mi'], row['sex'],
                      row['brgy'], row['mun'], row['prov'], row['cong'],
                      row['student_no'], row['course'], row['yr'], row['scholarship']]
            for ci, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = Font(size=9)
                cell.border = border
            row_at[0] += 1
        if not rows:
            r = row_at[0]
            ws.cell(row=r, column=1, value='No approved scholars for this programme.')
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
            ws.cell(row=r, column=1).font = Font(size=9, italic=True)
            row_at[0] += 1
        row_at[0] += 1

    ncols = len(sections[0]['headers'])
    banner(banner_text, ncols, PatternFill('solid', fgColor='1F4E79'),
           Font(bold=True, size=11, color='FFFFFF'), height=18)
    row_at[0] += 1

    for section in sections:
        banner(f"{section['title']}  ({section['total']} scholars)", ncols,
               PatternFill('solid', fgColor='BDD7EE'), Font(bold=True, size=10))
        label('FEMALE', ncols)
        table(section['headers'], section['female_rows'])
        label('MALE', ncols)
        table(section['headers'], section['male_rows'])

    widths = [5, 14, 18, 18, 6, 6, 16, 14, 14, 12, 14, 22, 5, 26]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_response(buf, filename):
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@_unifast_required
def unifast_report_download_excel(request):
    """The combined TDP/TES scholars masterlist."""
    from .models import SystemSettings

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    school_year = request.GET.get('sy', '').strip()
    ay = school_year or parsed['sy']
    semester = parsed['semester']

    buf = _unifast_scholars_workbook(
        _unifast_report_sections(school_year=school_year), ay, semester,
        'UniFAST Scholars',
        f'BILIRAN PROVINCE STATE UNIVERSITY — UniFAST SCHOLARS  |  {ay} {semester}',
    )
    return _xlsx_response(
        buf,
        f'UniFAST_Scholars_{ay.replace("-", "_")}_{semester.replace(" ", "_")}.xlsx',
    )


@_unifast_required
def unifast_report_download_tdp(request):
    """The TDP scholars masterlist on its own, for the school year chosen."""
    from .models import SystemSettings

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    school_year = request.GET.get('sy', '').strip()
    ay = school_year or parsed['sy']
    semester = parsed['semester']

    buf = _unifast_scholars_workbook(
        _unifast_report_sections(school_year=school_year, types=('TDP',)),
        ay, semester, 'TDP Scholars',
        f'BILIRAN PROVINCE STATE UNIVERSITY — TULONG DUNONG PROGRAM  |  {ay} {semester}',
    )
    return _xlsx_response(buf, f'TDP_Scholars_{ay.replace("-", "_")}.xlsx')


@_unifast_required
@xframe_options_exempt
def unifast_report_preview_tdp(request):
    """The TDP masterlist as a PDF page, for the preview frame on Reports.

    The same two-step the TES frame uses: convert the workbook that actually
    downloads when a converter is installed, so the preview is the file; draw
    the list here when one is not.
    """
    from .models import SystemSettings
    from . import doc_convert, report_pdf

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    school_year = request.GET.get('sy', '').strip()
    ay = school_year or parsed['sy']
    semester = parsed['semester']
    sections = _unifast_report_sections(school_year=school_year, types=('TDP',))

    pdf = None
    if doc_convert.available():
        try:
            buf = _unifast_scholars_workbook(
                sections, ay, semester, 'TDP Scholars',
                f'BILIRAN PROVINCE STATE UNIVERSITY — TULONG DUNONG PROGRAM  |  {ay} {semester}',
            )
            pdf = doc_convert.to_pdf(buf.getvalue(), '.xlsx')
        except (doc_convert.ConversionUnavailable, doc_convert.ConversionFailed):
            pdf = None
    if pdf is None:
        pdf = report_pdf.programme_masterlist_pdf(
            'TULONG DUNONG PROGRAM (TDP)', sections, ay, semester).read()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="TDP_Scholars_{ay.replace("-", "_")}.pdf"')
    return response


@_vpsea_required
def vpsea_announcements(request):
    from .models import Announcement
    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        body = (request.POST.get('body') or '').strip()
        if not title or not body:
            return redirect('/vpsea/announcements/?error=1')
        Announcement.objects.create(title=title, body=body, published_by=request.user)
        # The row alone reached nobody: it showed only in the dashboard's top
        # three and raised no notification at all.
        reached = notify.broadcast(title, body)
        return redirect(f'/vpsea/announcements/?posted={reached}')
    announcements = Announcement.objects.all().order_by('-created_at')
    return render(request, 'vpsea/announcements.html', {'announcements': announcements})


@_vpsea_required
def vpsea_reports(request):
    """Preview of the BiPSU scholars masterlist.

    Built from the same context that renders the Word document, so what the page
    shows and what downloads can never drift apart.
    """
    import os
    from .models import SystemSettings
    from . import doc_convert, masterlist_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    context, summary = masterlist_report.build_context()

    sections = []
    for entry in summary:
        slot = context[entry['slot']]
        gendered = entry['layout'] == 'gendered'
        headers = entry['headers']

        def cells(rows):
            # Resolved against this table's own headings, so the preview shows
            # exactly the columns the document will — no more, no fewer.
            return [masterlist_report.cells_for(r, headers) for r in rows]

        sections.append((
            entry['heading'],
            headers,
            cells(slot['female'] if gendered else slot['students']),
            cells(slot['male']) if gendered else [],
            gendered,
            entry['key'].lower(),
        ))

    return render(request, 'vpsea/reports.html', {
        'sections': sections,
        'semester': parsed['semester'],
        'ay': parsed['sy'],
        'grand_total': sum(e['total'] for e in summary),
        'summary': summary,
        'template_available': os.path.exists(masterlist_report.TEMPLATE_PATH),
        # Whether the frame is showing the Word document itself or the fallback
        # layout, so the page can say which one an officer is reading.
        'exact_preview': doc_convert.available(),
        'error': request.GET.get('error'),
    })


@_vpsea_required
@xframe_options_exempt
def vpsea_report_preview_pdf(request):
    """The masterlist as a PDF page, for the preview frame on the Reports tab.

    The Word document the office files is generated and converted, so the page
    on screen is that document rather than a picture of it. Without a converter
    installed the same rows are laid out here instead, and the tab says so.
    """
    from .models import SystemSettings
    from . import doc_convert, masterlist_report, report_pdf

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    label = settings_obj.academic_year.replace('-', '_')

    pdf = None
    if doc_convert.available():
        try:
            buf, _summary = masterlist_report.build_document(
                parsed['sy'], parsed['semester'])
            pdf = doc_convert.to_pdf(buf.getvalue(), '.docx')
        except (FileNotFoundError, doc_convert.ConversionUnavailable,
                doc_convert.ConversionFailed):
            pdf = None
    if pdf is None:
        buf, _summary = report_pdf.masterlist_pdf(
            parsed['sy'], parsed['semester'])
        pdf = buf.read()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="BiPSU_List_of_Scholars_{label}.pdf"')
    return response


@_vpsea_required
def vpsea_report_download(request):
    """The BiPSU scholars masterlist as the office's own Word document."""
    from .models import SystemSettings
    from . import masterlist_report

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    try:
        buf, _summary = masterlist_report.build_document(parsed['sy'], parsed['semester'])
    except FileNotFoundError as exc:
        from urllib.parse import quote
        return redirect(f'/vpsea/reports/?error={quote(str(exc))}')

    label = settings_obj.academic_year.replace('-', '_')
    filename = f'BiPSU_List_of_Scholars_{label}.docx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

    # — Page layout: landscape, legal-ish wide ————————————————
    section = doc.sections[0]
    section.orientation = 1  # LANDSCAPE
    section.page_width = Inches(13.0)
    section.page_height = Inches(8.5)
    section.left_margin = Inches(0.5)
    section.right_margin = Inches(0.5)
    section.top_margin = Inches(0.9)
    section.bottom_margin = Inches(0.4)

    # — Helpers ————————————————————————————————
    def add_heading(text, bold=False, size=11, align=WD_ALIGN_PARAGRAPH.CENTER):
        p = doc.add_paragraph()
        p.alignment = align
        run = p.add_run(text)
        run.bold = bold
        run.font.size = Pt(size)
        p.paragraph_format.space_before = Pt(0)
        p.paragraph_format.space_after = Pt(0)
        return p

    def add_blank():
        p = doc.add_paragraph()
        p.paragraph_format.space_before = Pt(0)
        p.paragraph_format.space_after = Pt(0)

    def set_cell_border(cell):
        tc = cell._tc
        tcPr = tc.get_or_add_tcPr()
        for edge in ('top', 'left', 'bottom', 'right'):
            tag = OxmlElement(f'w:{edge}')
            tag.set(qn('w:val'), 'single')
            tag.set(qn('w:sz'), '4')
            tag.set(qn('w:space'), '0')
            tag.set(qn('w:color'), '000000')
            tcPr.append(tag)

    def style_header_row(row, bold=True, bg='D9E1F2'):
        for cell in row.cells:
            set_cell_border(cell)
            tc = cell._tc
            tcPr = tc.get_or_add_tcPr()
            shd = OxmlElement('w:shd')
            shd.set(qn('w:val'), 'clear')
            shd.set(qn('w:color'), 'auto')
            shd.set(qn('w:fill'), bg)
            tcPr.append(shd)
            for para in cell.paragraphs:
                para.alignment = WD_ALIGN_PARAGRAPH.CENTER
                for run in para.runs:
                    run.bold = bold
                    run.font.size = Pt(8)

    def style_data_row(row):
        for cell in row.cells:
            set_cell_border(cell)
            for para in cell.paragraphs:
                for run in para.runs:
                    run.font.size = Pt(8)

    def add_scholar_table(headers, sub_headers, rows_data):
        ncols = len(headers)
        table = doc.add_table(rows=0, cols=ncols)
        table.style = 'Table Grid'
        table.alignment = WD_TABLE_ALIGNMENT.CENTER
        # Header row 1
        hrow1 = table.add_row()
        for i, h in enumerate(headers):
            hrow1.cells[i].text = h
        style_header_row(hrow1)
        # Header row 2 (sub-headers)
        if sub_headers:
            hrow2 = table.add_row()
            for i, h in enumerate(sub_headers):
                hrow2.cells[i].text = h
            style_header_row(hrow2)
        # Data rows
        for row_vals in rows_data:
            drow = table.add_row()
            for i, val in enumerate(row_vals):
                drow.cells[i].text = str(val) if val is not None else ''
            style_data_row(drow)
        return table

    def _split_name(full_name):
        parts = full_name.strip().split()
        if len(parts) == 0: return ('', '', '')
        if len(parts) == 1: return (parts[0], '', '')
        if len(parts) == 2: return (parts[-1], parts[0], '')
        last = parts[-1]
        first = parts[0]
        middle = ' '.join(parts[1:-1])
        mi = middle[0] + '.' if middle else ''
        return (last, first, mi)

    def _name_parts(user):
        return (user.last_name or '', user.first_name or '', '')

    # — Document header ————————————————————————————
    add_heading('Republic of the Philippines', bold=False, size=11)
    add_heading('BILIRAN PROVINCE STATE UNIVERSITY', bold=True, size=11)
    add_heading('Naval, Biliran', bold=False, size=11)
    add_blank()
    add_heading(f'LIST OF SCHOLARS FOR {sem_label}', bold=True, size=16)
    add_blank()

    # — ACADEMIC ———————————————————————————————
    academic = Application.objects.filter(
        status='Approved', scholarship__type='Academic'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')

    females = [a for a in academic if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    males   = [a for a in academic if a not in females]

    add_heading('ACADEMIC (@)', bold=True, size=11)
    add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
    add_heading(f'{semester} SY: {ay}', bold=False, size=11)
    add_blank()

    headers_acad = ['NO.', 'NAME', 'NAME', 'NAME', 'SEX', 'ADDRESS', 'ADDRESS', 'ADDRESS', 'COURSE', 'YR.', 'GWA', '%', 'SCHOLARSHIP PROGRAM']
    sub_acad     = ['NO.', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'COURSE', 'YR.', 'GWA', '%', 'SCHOLARSHIP PROGRAM']

    def acad_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            p = app.student; u = p.user
            last, first, mi = _name_parts(u)
            addr = p.address or ''
            parts = [x.strip() for x in addr.split(',')]
            brgy = parts[0] if len(parts) > 0 else ''
            mun  = parts[1] if len(parts) > 1 else ''
            prov = parts[2] if len(parts) > 2 else ''
            pct = 'University Scholar' if p.gwa <= 1.29 else ('College Scholars' if p.gwa <= 1.50 else '')
            rows.append([i, last, first, mi, p.gender or '', brgy, mun, prov, p.course, p.year_level, p.gwa, pct, 'ACADEMIC'])
        return rows

    p_label = doc.add_paragraph('FEMALE')
    p_label.paragraph_format.space_before = Pt(0)
    p_label.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_acad, sub_acad, acad_rows(females))
    add_blank()
    p_label2 = doc.add_paragraph('MALE')
    p_label2.paragraph_format.space_before = Pt(0)
    p_label2.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_acad, sub_acad, acad_rows(males))
    add_blank()

    # — BiPSU STAFF ———————————————————————————————
    staff = AffirmativeStaffApplication.objects.filter(
        status='Approved', qualified_for='Staff'
    ).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name')

    add_heading('BiPSU STAFF (@)', bold=True, size=11)
    add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
    add_heading(f'{semester} SY: {ay}', bold=False, size=11)
    add_blank()

    headers_staff = ['NO. ', 'NAME ', 'NAME ', 'NAME ', 'SEX', 'COURSE ', 'YEAR LEVEL ', 'STUDENT ', '% ', 'SCHOLARSHIP ']
    sub_staff     = ['NO. ', 'LAST NAME ', 'FIRST NAME ', 'M.I. ', 'SEX', 'COURSE ', 'YEAR LEVEL ', 'NUMBER ', '% ', 'PROGRAM ']
    staff_rows = []
    for i, app in enumerate(staff, 1):
        last, first, mi = _split_name(app.full_name)
        pct = '100' if app.is_nsu_staff else '75'
        staff_rows.append([i, last, first, mi, app.gender or '', app.course, app.year_level, app.student_id or '', pct, 'BiPSU STAFF SCHOLARSHIP'])
    add_scholar_table(headers_staff, sub_staff, staff_rows)
    add_blank()

    # — AFFIRMATIVE ACTION ————————————————————————————
    affirmative = AffirmativeStaffApplication.objects.filter(
        status='Approved', qualified_for='Affirmative'
    ).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name')

    aff_females = [a for a in affirmative if a.gender and a.gender.upper() in ('F', 'FEMALE')]
    aff_males   = [a for a in affirmative if a not in aff_females]

    add_heading('AFFIRMATIVE ACTION (*)', bold=True, size=11)
    add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
    add_heading(f'{semester} SY: {ay}', bold=False, size=11)
    add_blank()

    headers_aff = ['NO.', 'AWARD NUMBER', 'NAME', 'NAME', 'NAME', 'SEX', 'ADDRESS', 'ADDRESS', 'ADDRESS', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']
    sub_aff     = ['NO.', 'AWARD NUMBER', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']

    def aff_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            last, first, mi = _split_name(app.full_name)
            addr = app.address or ''
            parts = [x.strip() for x in addr.split(',')]
            brgy = parts[0] if len(parts) > 0 else ''
            mun  = parts[1] if len(parts) > 1 else ''
            prov = parts[2] if len(parts) > 2 else ''
            rows.append([i, '', last, first, mi, app.gender or '', brgy, mun, prov, '', app.course, app.year_level, 'Affirmative Action Scholarship'])
        return rows

    p_f = doc.add_paragraph('FEMALE')
    p_f.paragraph_format.space_before = Pt(0)
    p_f.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_aff, sub_aff, aff_rows(aff_females))
    add_blank()
    p_m = doc.add_paragraph('MALE')
    p_m.paragraph_format.space_before = Pt(0)
    p_m.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_aff, sub_aff, aff_rows(aff_males))
    add_blank()

    # — CHED (FULL MERIT / HALF MERIT) ————————————————————
    ched_all = Application.objects.filter(
        status='Approved', scholarship__type='CHED'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')

    ched_full, ched_half = split_ched(ched_all)

    headers_ched = ['NO.', 'AWARD NUMBER', 'NAME', 'NAME', 'NAME', 'SEX', 'ADDRESS', 'ADDRESS', 'ADDRESS', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']
    sub_ched     = ['NO.', 'AWARD NUMBER', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']

    def ched_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            p = app.student; u = p.user
            last, first, mi = _name_parts(u)
            addr = p.address or ''
            parts = [x.strip() for x in addr.split(',')]
            brgy = parts[0] if len(parts) > 0 else ''
            mun  = parts[1] if len(parts) > 1 else ''
            prov = parts[2] if len(parts) > 2 else ''
            award = app.award_number
            cong  = app.congress_district
            rows.append([i, award, last, first, mi, p.gender or '', brgy, mun, prov, cong, p.course, p.year_level, app.scholarship.name])
        return rows

    for block_title, block_apps in [('FULL MERIT/ FULL SCHOLAR (*)', ched_full), ('HALF MERIT/ PARTIAL SCHOLAR (*)', ched_half)]:
        add_heading(block_title, bold=True, size=11)
        add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
        add_heading(f'{semester} SY: {ay}', bold=False, size=11)
        add_blank()
        ched_f = [a for a in block_apps if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
        ched_m = [a for a in block_apps if a not in ched_f]
        pf = doc.add_paragraph('FEMALE')
        pf.paragraph_format.space_before = Pt(0)
        pf.paragraph_format.space_after = Pt(0)
        add_scholar_table(headers_ched, sub_ched, ched_rows(ched_f))
        add_blank()
        pm = doc.add_paragraph('MALE')
        pm.paragraph_format.space_before = Pt(0)
        pm.paragraph_format.space_after = Pt(0)
        add_scholar_table(headers_ched, sub_ched, ched_rows(ched_m))
        add_blank()

    # — DOST —————————————————————————————————
    dost_all = Application.objects.filter(
        status='Approved', scholarship__type='DOST'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')

    add_heading('DOST (*)', bold=True, size=11)
    add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
    add_heading(f'{semester} SY: {ay}', bold=False, size=11)
    add_blank()

    dost_f = [a for a in dost_all if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    dost_m = [a for a in dost_all if a not in dost_f]
    pf = doc.add_paragraph('FEMALE')
    pf.paragraph_format.space_before = Pt(0)
    pf.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_ched, sub_ched, ched_rows(dost_f))
    add_blank()
    pm = doc.add_paragraph('MALE')
    pm.paragraph_format.space_before = Pt(0)
    pm.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_ched, sub_ched, ched_rows(dost_m))
    add_blank()

    # — GSIS —————————————————————————————————
    gsis_all = Application.objects.filter(
        status='Approved', scholarship__type='GSIS'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')

    add_heading('GSIS (*)', bold=True, size=11)
    add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
    add_heading(f'{semester} SY: {ay}', bold=False, size=11)
    add_blank()

    headers_gsis = ['NO.', 'NAME', 'NAME', 'NAME', 'SEX', 'ADDRESS', 'ADDRESS', 'ADDRESS', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']
    sub_gsis     = ['NO.', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']

    def gsis_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            p = app.student; u = p.user
            last, first, mi = _name_parts(u)
            addr = p.address or ''
            parts = [x.strip() for x in addr.split(',')]
            brgy = parts[0] if len(parts) > 0 else ''
            mun  = parts[1] if len(parts) > 1 else ''
            prov = parts[2] if len(parts) > 2 else ''
            cong = app.congress_district
            rows.append([i, last, first, mi, p.gender or '', brgy, mun, prov, cong, p.course, p.year_level, app.scholarship.name])
        return rows

    gsis_f = [a for a in gsis_all if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    gsis_m = [a for a in gsis_all if a not in gsis_f]
    pf = doc.add_paragraph('FEMALE')
    pf.paragraph_format.space_before = Pt(0)
    pf.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_gsis, sub_gsis, gsis_rows(gsis_f))
    add_blank()
    pm = doc.add_paragraph('MALE')
    pm.paragraph_format.space_before = Pt(0)
    pm.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_gsis, sub_gsis, gsis_rows(gsis_m))
    add_blank()

    # — TES (TDP) ———————————————————————————————
    tes_all = Application.objects.filter(
        status='Approved', scholarship__type='TDP'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')

    add_heading('TERTIARY EDUCATION SUBSIDY -TES (*)', bold=True, size=11)
    add_heading('SCHOLARSHIP GRANT', bold=False, size=11)
    add_heading(f'{semester} SY: {ay}', bold=False, size=11)
    add_blank()

    tes_f = [a for a in tes_all if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    tes_m = [a for a in tes_all if a not in tes_f]
    pf = doc.add_paragraph('FEMALE')
    pf.paragraph_format.space_before = Pt(0)
    pf.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_ched, sub_ched, ched_rows(tes_f))
    add_blank()
    pm = doc.add_paragraph('MALE')
    pm.paragraph_format.space_before = Pt(0)
    pm.paragraph_format.space_after = Pt(0)
    add_scholar_table(headers_ched, sub_ched, ched_rows(tes_m))
    add_blank()

    # — Page footer with signatories (appears on every page) —————————
    footer = section.footer
    footer.is_linked_to_previous = False
    # Clear default empty paragraph
    for para in footer.paragraphs:
        p_elem = para._p
        p_elem.getparent().remove(p_elem)

    footer_table = footer.add_table(rows=2, cols=4, width=Inches(12.0))
    footer_table.style = 'Table Grid'

    sig_labels = [
        ('Prepared by:', 'Noted:', 'Recommending approval:', 'Approved:'),
        (
            'MARICEL S. SAULAN\nScholarship in charge',
            'NORMA M. DUALLO, Ph.D.TM\nSDSO Director',
            'ERWIN G. SALVATIERRA, Ph. D.\nVP for Extension Services,\nStudent and External Affairs',
            'VICTOR C. CAÃ‘EZO, JR., Ed. D.\nUniversity President',
        ),
    ]
    for ri, row_vals in enumerate(sig_labels):
        for ci, val in enumerate(row_vals):
            cell = footer_table.rows[ri].cells[ci]
            para = cell.paragraphs[0]
            para.alignment = WD_ALIGN_PARAGRAPH.CENTER
            # remove cell borders
            tc = cell._tc
            tcPr = tc.get_or_add_tcPr()
            for edge in ('top', 'left', 'bottom', 'right'):
                tag = OxmlElement(f'w:{edge}')
                tag.set(qn('w:val'), 'none')
                tcPr.append(tag)
            if ri == 0:
                run = para.add_run(val)
                run.bold = True
                run.font.size = Pt(8)
            else:
                lines = val.split('\n')
                r1 = para.add_run(lines[0])
                r1.bold = True
                r1.font.size = Pt(8)
                for line in lines[1:]:
                    para.add_run('\n' + line).font.size = Pt(8)

    # — Save & return —————————————————————————————
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    filename = f'Scholarship_Report_{ay.replace("-","_")}_{semester.replace(" ","_")}.docx'
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@_vpsea_required
def vpsea_report_download_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from django.http import HttpResponse
    from .models import Application, AffirmativeStaffApplication, SystemSettings

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    semester = settings_obj.active_semester
    ay = settings_obj.academic_year

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Scholars'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=True, size=9)
    header_fill = PatternFill('solid', fgColor='D9E1F2')
    section_fill = PatternFill('solid', fgColor='BDD7EE')
    title_fill = PatternFill('solid', fgColor='1F4E79')
    title_font = Font(bold=True, size=11, color='FFFFFF')

    current_row = [1]  # mutable so nested helpers can update it

    def write_title(text, ncols):
        r = current_row[0]
        ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center
        cell.border = border
        ws.row_dimensions[r].height = 18
        current_row[0] += 1

    def write_section(text, ncols):
        r = current_row[0]
        ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, size=10)
        cell.fill = section_fill
        cell.alignment = center
        cell.border = border
        current_row[0] += 1

    def write_gender_label(text, ncols):
        r = current_row[0]
        ws.cell(row=r, column=1, value=text)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row[0] += 1

    def write_headers(headers):
        r = current_row[0]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
        current_row[0] += 1

    def write_rows(rows_data):
        for row_vals in rows_data:
            r = current_row[0]
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.font = Font(size=9)
            current_row[0] += 1

    def blank_row():
        current_row[0] += 1

    def _split_name(full_name):
        parts = full_name.strip().split()
        if len(parts) == 0: return ('', '', '')
        if len(parts) == 1: return (parts[0], '', '')
        if len(parts) == 2: return (parts[-1], parts[0], '')
        last = parts[-1]; first = parts[0]
        middle = ' '.join(parts[1:-1])
        return (last, first, middle[0] + '.' if middle else '')

    def _name_parts(user):
        return (user.last_name or '', user.first_name or '', '')

    def addr_parts(addr):
        parts = [x.strip() for x in (addr or '').split(',')]
        return (parts[0] if len(parts) > 0 else '',
                parts[1] if len(parts) > 1 else '',
                parts[2] if len(parts) > 2 else '')

    MAX_COLS = 13  # widest table

    # — Document title ————————————————————————————
    write_title('Republic of the Philippines', MAX_COLS)
    write_title('BILIRAN PROVINCE STATE UNIVERSITY — Naval, Biliran', MAX_COLS)
    write_title(f'LIST OF SCHOLARS FOR {semester} SY: {ay}', MAX_COLS)
    blank_row()

    # — ACADEMIC ———————————————————————————————
    academic = list(Application.objects.filter(
        status='Approved', scholarship__type='Academic'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name'))
    females_a = [a for a in academic if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    males_a   = [a for a in academic if a not in females_a]

    headers_acad = ['NO.', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'COURSE', 'YR.', 'GWA', '%', 'SCHOLARSHIP PROGRAM']
    write_section(f'ACADEMIC (@) SCHOLARSHIP GRANT — {semester} SY: {ay}', len(headers_acad))

    def acad_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            p = app.student; u = p.user
            last, first, mi = _name_parts(u)
            brgy, mun, prov = addr_parts(p.address)
            pct = 'University Scholar' if p.gwa <= 1.29 else ('College Scholars' if p.gwa <= 1.50 else '')
            rows.append([i, last, first, mi, p.gender or '', brgy, mun, prov, p.course, p.year_level, p.gwa, pct, 'ACADEMIC'])
        return rows

    write_gender_label('FEMALE', len(headers_acad))
    write_headers(headers_acad)
    write_rows(acad_rows(females_a))
    write_gender_label('MALE', len(headers_acad))
    write_headers(headers_acad)
    write_rows(acad_rows(males_a))
    blank_row()

    # — BiPSU STAFF ———————————————————————————————
    staff = list(AffirmativeStaffApplication.objects.filter(
        status='Approved', qualified_for='Staff'
    ).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name'))
    headers_staff = ['NO.', 'LAST NAME', 'FIRST NAME', 'M.I.', 'SEX', 'COURSE', 'YEAR LEVEL', 'STUDENT NUMBER', '%', 'SCHOLARSHIP PROGRAM']
    write_section(f'BiPSU STAFF (@) SCHOLARSHIP GRANT — {semester} SY: {ay}', len(headers_staff))
    write_headers(headers_staff)
    staff_rows = []
    for i, app in enumerate(staff, 1):
        last, first, mi = _split_name(app.full_name)
        pct = '100' if app.is_nsu_staff else '75'
        staff_rows.append([i, last, first, mi, app.gender or '', app.course, app.year_level, app.student_id or '', pct, 'BiPSU STAFF SCHOLARSHIP'])
    write_rows(staff_rows)
    blank_row()

    # — AFFIRMATIVE ——————————————————————————————
    affirmative = list(AffirmativeStaffApplication.objects.filter(
        status='Approved', qualified_for='Affirmative'
    ).select_related(*STAFF_APPLICATION_DETAILS).order_by('full_name'))
    aff_females = [a for a in affirmative if a.gender and a.gender.upper() in ('F', 'FEMALE')]
    aff_males   = [a for a in affirmative if a not in aff_females]
    headers_aff = ['NO.', 'AWARD NUMBER', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']
    write_section(f'AFFIRMATIVE ACTION (*) SCHOLARSHIP GRANT — {semester} SY: {ay}',
                  len(headers_aff))

    def aff_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            last, first, mi = _split_name(app.full_name)
            brgy, mun, prov = addr_parts(app.address)
            rows.append([i, '', last, first, mi, app.gender or '', brgy, mun, prov, '', app.course, app.year_level, 'Affirmative Action Scholarship'])
        return rows

    write_gender_label('FEMALE', len(headers_aff))
    write_headers(headers_aff)
    write_rows(aff_rows(aff_females))
    write_gender_label('MALE', len(headers_aff))
    write_headers(headers_aff)
    write_rows(aff_rows(aff_males))
    blank_row()

    # — CHED —————————————————————————————————
    ched_all = list(Application.objects.filter(
        status='Approved', scholarship__type='CHED'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name'))
    ched_full, ched_half = split_ched(ched_all)
    headers_ched = ['NO.', 'AWARD NUMBER', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']

    def ched_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            p = app.student; u = p.user
            last, first, mi = _name_parts(u)
            brgy, mun, prov = addr_parts(p.address)
            award = app.award_number
            cong  = app.congress_district
            rows.append([i, award, last, first, mi, p.gender or '', brgy, mun, prov, cong, p.course, p.year_level, app.scholarship.name])
        return rows

    for block_title, block_apps in [('FULL MERIT/ FULL SCHOLAR (*)', ched_full), ('HALF MERIT/ PARTIAL SCHOLAR (*)', ched_half)]:
        write_section(f'{block_title} SCHOLARSHIP GRANT — {semester} SY: {ay}', len(headers_ched))
        bf = [a for a in block_apps if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
        bm = [a for a in block_apps if a not in bf]
        write_gender_label('FEMALE', len(headers_ched))
        write_headers(headers_ched)
        write_rows(ched_rows(bf))
        write_gender_label('MALE', len(headers_ched))
        write_headers(headers_ched)
        write_rows(ched_rows(bm))
        blank_row()

    # — DOST —————————————————————————————————
    dost_all = list(Application.objects.filter(
        status='Approved', scholarship__type='DOST'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name'))
    write_section(f'DOST (*) SCHOLARSHIP GRANT — {semester} SY: {ay}', len(headers_ched))
    dost_f = [a for a in dost_all if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    dost_m = [a for a in dost_all if a not in dost_f]
    write_gender_label('FEMALE', len(headers_ched))
    write_headers(headers_ched)
    write_rows(ched_rows(dost_f))
    write_gender_label('MALE', len(headers_ched))
    write_headers(headers_ched)
    write_rows(ched_rows(dost_m))
    blank_row()

    # — GSIS —————————————————————————————————
    gsis_all = list(Application.objects.filter(
        status='Approved', scholarship__type='GSIS'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name'))
    headers_gsis = ['NO.', 'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'SEX', 'BRGY./ST.', 'MUN.', 'PROV.', 'CONG. DIST.', 'COURSE', 'YR.', 'SCHOLARSHIP PROGRAM']
    write_section(f'GSIS (*) SCHOLARSHIP GRANT — {semester} SY: {ay}', len(headers_gsis))

    def gsis_rows(apps):
        rows = []
        for i, app in enumerate(apps, 1):
            p = app.student; u = p.user
            last, first, mi = _name_parts(u)
            brgy, mun, prov = addr_parts(p.address)
            cong = app.congress_district
            rows.append([i, last, first, mi, p.gender or '', brgy, mun, prov, cong, p.course, p.year_level, app.scholarship.name])
        return rows

    gsis_f = [a for a in gsis_all if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    gsis_m = [a for a in gsis_all if a not in gsis_f]
    write_gender_label('FEMALE', len(headers_gsis))
    write_headers(headers_gsis)
    write_rows(gsis_rows(gsis_f))
    write_gender_label('MALE', len(headers_gsis))
    write_headers(headers_gsis)
    write_rows(gsis_rows(gsis_m))
    blank_row()

    # — TES (TDP) ———————————————————————————————
    tes_all = list(Application.objects.filter(
        status='Approved', scholarship__type='TDP'
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name'))
    write_section(f'TERTIARY EDUCATION SUBSIDY -TES (*) SCHOLARSHIP GRANT — {semester} SY: {ay}', len(headers_ched))
    tes_f = [a for a in tes_all if a.student.gender and a.student.gender.upper() in ('F', 'FEMALE')]
    tes_m = [a for a in tes_all if a not in tes_f]
    write_gender_label('FEMALE', len(headers_ched))
    write_headers(headers_ched)
    write_rows(ched_rows(tes_f))
    write_gender_label('MALE', len(headers_ched))
    write_headers(headers_ched)
    write_rows(ched_rows(tes_m))
    blank_row()
    blank_row()

    # — Page footer with signatories (appears on every printed page) —————
    footer_text = (
        'Prepared by:\t\t\t\t\tNoted:\t\t\t\t\t\tRecommending approval:\t\t\t\t\t\tApproved:\n'
        'MARICEL S. SAULAN\t\t\t\tNORMA M. DUALLO, Ph.D.TM\t\t\tERWIN G. SALVATIERRA, Ph. D.\t\t\tVICTOR C. CAÃ‘EZO, JR., Ed. D.\n'
        'Scholarship in charge\t\t\t\tSDSO Director\t\t\t\t\tVP for Extension Services, Student and External Affairs\t\tUniversity President'
    )
    ws.oddFooter.center.text = footer_text
    ws.oddFooter.center.size = 8
    ws.evenFooter.center.text = footer_text
    ws.evenFooter.center.size = 8

    # — Auto-fit columns (skip MergedCell objects) ——————————————
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

    # — Save & return —————————————————————————————
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f'Scholarship_Report_{ay.replace("-","_")}_{semester.replace(" ","_")}.xlsx'
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@_vpsea_required
def vpsea_accounts(request):
    """Verification queue for accounts that registered themselves.

    Nobody who signs up on the public form can sign in until an officer decides
    here. The decision writes the message the person reads on the login page,
    which is the only channel that reaches someone who cannot get in yet — so
    a rejection has to say why.
    """
    from .models import ActivityLog, Notification, StaffProfile, StudentProfile

    if request.method == 'POST':
        account = User.objects.filter(
            id=request.POST.get('user_id'), role__in=('student', 'nsu_staff'),
        ).first()
        if not account:
            return redirect('/vpsea/accounts/?error=Account+not+found')

        action = request.POST.get('action')
        message = request.POST.get('message', '').strip()
        if action == 'reject' and not message:
            return redirect('/vpsea/accounts/?error=A+reason+is+required+when+rejecting'
                            '+—+it+is+what+the+person+reads+when+they+try+to+sign+in')
        if action not in ('approve', 'reject'):
            return redirect('/vpsea/accounts/?error=Unknown+action')

        # The scholarship the student declared on the registration form is
        # decided with the account, not on a queue of its own: it is part of
        # what the officer is checking, and it arrived with this registration.
        declared = declared_scholarship(getattr(account, 'profile', None))
        if declared and action == 'approve':
            archive = None
            archive_id = request.POST.get('archive_id', '').strip()
            if archive_id:
                from .models import ImportedScholar, SystemSettings
                settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
                archive = ImportedScholar.objects.filter(
                    id=archive_id, scholarship_type=declared.scholarship_type,
                    term_label=settings_obj.academic_year, claimed_by__isnull=True,
                ).first()
                if not archive:
                    return redirect('/vpsea/accounts/?error=That+archive+row+is+no+'
                                    'longer+available')
            _award, problem = approve_declared_scholarship(
                declared, request.user, archive=archive, remarks=message,
                tier=request.POST.get('award_tier', ''))
            if problem:
                from urllib.parse import quote
                return redirect(f'/vpsea/accounts/?error={quote(problem)}')
        elif declared:
            reject_declared_scholarship(declared, request.user, message)

        status = 'approved' if action == 'approve' else 'rejected'
        account.decide_verification(status, message, request.user)

        # A verified student finds this waiting in their portal. Staff
        # notifications hang off StudentProfile, which staff do not have, so
        # for them the login page is the whole of it.
        # Emailed either way: a rejected applicant has no portal to read a
        # notification in, and used to be told nothing at all.
        _in_app, emailed = notify.account_decision(
            account, status, account.verification_note)
        ActivityLog.objects.create(
            user=request.user,
            action=(f'Account {status}: {account.get_full_name() or account.email} '
                    f'({account.get_role_display()}) — {account.verification_note}'),
        )
        # Whether the message actually left the building. The office was told
        # 'verified' either way, so a mail server that was down, or never
        # configured, looked exactly like one that had delivered.
        return redirect(f'/vpsea/accounts/?{action}d=1&emailed={1 if emailed else 0}')

    pending = list(User.objects.filter(
        verification_status='pending', role__in=('student', 'nsu_staff'),
    ).order_by('date_joined'))
    decided = list(User.objects.filter(
        verification_status__in=('approved', 'rejected'),
        role__in=('student', 'nsu_staff'), verified_at__isnull=False,
    ).select_related('verified_by').order_by('-verified_at')[:25])

    # What each account claimed about itself, so the officer can check it
    # against their own records without opening another page.
    students = {p.user_id: p for p in StudentProfile.with_details().filter(
        user__in=pending + decided)}
    staff = {p.user_id: p for p in StaffProfile.objects.filter(
        user__in=pending + decided)}
    for account in pending + decided:
        account.student_profile = students.get(account.id)
        account.employee_profile = staff.get(account.id)

    # The scholarship each waiting registration declared, with the imported rows
    # it could be. Only for the queue: a decided account's award is on the
    # archives page, where every other award is.
    from .models import SystemSettings
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active_label = settings_obj.academic_year
    for account in pending:
        req = declared_scholarship(account.student_profile)
        account.declared = req
        account.archive_candidates = (
            list(_archive_candidates(req, active_label)) if req else [])
        account.other_semester_rows = (
            list(_archive_candidates(req).exclude(term_label=active_label)[:5])
            if req else [])

    from django.conf import settings as django_settings
    from .models import CHED_TIER_CHOICES

    return render(request, 'vpsea/accounts.html', {
        'active': 'accounts',
        'pending': pending,
        'decided': decided,
        'ched_tiers': CHED_TIER_CHOICES,
        'active_label': active_label,
        'error': request.GET.get('error', ''),
        'approved': request.GET.get('approved'),
        'rejected': request.GET.get('rejected'),
        # Whether the decision just made was actually emailed, and whether this
        # deployment can send mail at all. Without the second, an office with no
        # SMTP configured would read every 'not emailed' as a broken server.
        'emailed': request.GET.get('emailed'),
        'email_enabled': django_settings.EMAIL_ENABLED,
    })


@_vpsea_required
def vpsea_students(request):
    from .models import StudentProfile, Application, Scholarship, SystemSettings
    from django.db.models import Q

    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    # academic_year holds the '<yy>-<sem>' label, e.g. '26-1'. Expanding it is
    # what the page means to show: it was printing 'A.Y. 26-1' in the subtitle
    # and, worse, int('26-1'.split('-')[0]) == 26 was being used as a calendar
    # year, so the approved-scholar filter matched nothing at all.
    active_label = settings_obj.academic_year
    parsed = SystemSettings.parse_label(active_label)
    current_sy = parsed['sy']                    # e.g. '2026-2027'
    current_sem = parsed['semester']

    q = request.GET.get('q', '').strip()
    stype = request.GET.get('stype', '').strip()

    scholarship_types = ['Academic', 'TDP', 'DOST', 'CHED', 'CoScho', 'Sports', 'Affirmative', 'Staff', 'GSIS']

    # Base: only students with an Approved award for the active term, matched
    # on the term column rather than guessed from the submission date.
    students = (StudentProfile.objects
                .select_related('user', *StudentProfile.DETAIL_RELATIONS)
                .order_by('user__last_name'))
    if active_label:
        students = students.filter(
            applications__status='Approved',
            applications__term_label=active_label,
        ).distinct()
    else:
        students = students.filter(applications__status='Approved').distinct()

    if q:
        students = students.filter(
            Q(user__last_name__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(student_id__icontains=q)
        )

    if stype:
        if stype in ('Affirmative', 'Staff'):
            from .models import AffirmativeStaffApplication
            aff_emails = AffirmativeStaffApplication.objects.filter(
                qualified_for=stype, status='Approved'
            ).values_list('email', flat=True)
            students = students.filter(user__email__in=aff_emails)
        else:
            students = students.filter(
                applications__scholarship__type=stype
            ).distinct()

    # Students with NO approved application at all — shown in the second tab.
    # This is a simple anti-join: every StudentProfile whose pk is not in the
    # set of profiles that have at least one Approved application. Accounts the
    # office rejected at registration are left out for the same reason as on the
    # archives tab: they cannot sign in to apply, so they are not students the
    # system failed to serve.
    approved_pks = Application.objects.filter(
        status='Approved'
    ).values_list('student_id', flat=True).distinct()

    no_scholarship_qs = (
        StudentProfile.objects
        .select_related('user', *StudentProfile.DETAIL_RELATIONS)
        .exclude(pk__in=approved_pks)
        .exclude(user__verification_status='rejected')
        .order_by('user__last_name', 'user__first_name')
    )
    if q:
        no_scholarship_qs = no_scholarship_qs.filter(
            Q(user__last_name__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(student_id__icontains=q)
        )

    tab = request.GET.get('tab', 'scholars')   # 'scholars' | 'no_scholarship'

    return render(request, 'vpsea/students.html', {
        'students': students,
        'no_scholarship_students': no_scholarship_qs,
        'q': q,
        'stype': stype,
        'tab': tab,
        'scholarship_types': scholarship_types,
        'current_sy': current_sy,
        'current_sem': current_sem,
    })


@_vpsea_required
def vpsea_student_add(request):
    from .models import StudentProfile, Scholarship, Application, ApplicationDocument
    from .constants import CIVIL_STATUSES, SEMESTERS, STUDENT_LEVELS
    errors = []
    if request.method == 'POST':
        p = request.POST
        f = request.FILES
        email = p.get('email', '').strip()
        student_id = p.get('student_id', '').strip()
        if User.objects.filter(email=email).exists():
            errors.append('Email already registered.')
        if StudentProfile.objects.filter(student_id=student_id).exists():
            errors.append('Student ID already registered.')
        if not errors:
            user = User.objects.create_user(
                username=email, email=email,
                password=p.get('password') or student_id,
                first_name=p.get('first_name', ''),
                last_name=p.get('last_name', ''),
                role='student',
            )
            profile = StudentProfile.objects.create(
                user=user,
                student_id=student_id,
                middle_name=p.get('middle_name', '').strip(),
                suffix=p.get('suffix', '').strip(),
                school=p.get('school', ''),
                course=p.get('course', ''),
                year_level=int(p.get('year_level', 1)),
                gwa=float(p.get('gwa', 0) or 0),
                contact_number=p.get('contact_number', ''),
                barangay=p.get('barangay', ''),
                municipality=p.get('municipality', ''),
                province=p.get('province', ''),
                date_of_birth=p.get('date_of_birth') or None,
                gender=p.get('gender', ''),
                family_income=float(p.get('family_income', 0) or 0),
                **_enrollment_fields(p),
            )
            # Build form_data from all extra fields
            form_data = {
                k: v for k, v in p.items()
                if k not in ('csrfmiddlewaretoken',) + STUDENT_RECORD_FIELDS
            }
            scholarship = Scholarship.objects.filter(type='Academic').first()
            if scholarship:
                app = Application.objects.create(
                    student=profile, scholarship=scholarship,
                    status='Approved', form_data=form_data,
                )
                doc_fields = ['doc_certificate_of_grades', 'doc_certificate_of_enrollment',
                              'doc_prospectus', 'doc_id_photo', 'doc_application_form']
                for field in doc_fields:
                    uploaded = f.get(field)
                    if uploaded:
                        ApplicationDocument.objects.create(
                            application=app, name=field.replace('doc_', '').replace('_', ' ').title(),
                            file=uploaded,
                        )
            return redirect('/vpsea/students/?added=1')
    import json

    return render(request, 'vpsea/student_form.html', {'errors': errors, 'action': 'Add', 'form_data': request.POST, 'doc_list': _doc_list(),
        # Adding a student creates an application, so the application sections apply.
        'show_application_fields': True,
        'cancel_url': request.GET.get('next') or '/vpsea/students/',

        'bipsu_schools': BIPSU_SCHOOLS, 'bipsu_courses_json': json.dumps(BIPSU_COURSES),
        'v_first_name': request.POST.get('first_name', ''),
        'v_last_name': request.POST.get('last_name', ''),
        'v_middle_name': request.POST.get('middle_name', ''),
        'v_suffix': request.POST.get('suffix', ''),
        'v_school': request.POST.get('school', ''),
        'v_email': request.POST.get('email', ''),
        'v_student_id': request.POST.get('student_id', ''),
        'v_course': request.POST.get('course', ''),
        'v_year_level': request.POST.get('year_level', '1'),
        'v_gwa': request.POST.get('gwa', ''),
        'v_gender': request.POST.get('gender', ''),
        'v_date_of_birth': request.POST.get('date_of_birth', ''),
        'v_contact_number': request.POST.get('contact_number', ''),
        'v_barangay': request.POST.get('barangay', ''),
        'v_municipality': request.POST.get('municipality', ''),
        'v_province': request.POST.get('province', ''),
        'v_family_income': request.POST.get('family_income', ''),
        'v_birth_place': request.POST.get('birth_place', ''),
        'v_civil_status': request.POST.get('civil_status', ''),
        'v_level': request.POST.get('level', ''),
        'v_department': request.POST.get('department', ''),
        'v_curriculum': request.POST.get('curriculum', ''),
        'v_learner_ref_no': request.POST.get('learner_ref_no', ''),
        'v_entry_period': request.POST.get('entry_period', ''),
        'v_entry_date': request.POST.get('entry_date', ''),
        'v_exam_score': request.POST.get('exam_score', ''),
        'civil_statuses': CIVIL_STATUSES, 'student_levels': STUDENT_LEVELS,
        'semesters': SEMESTERS,
        'v_elementary': request.POST.get('elementary', ''),
        'v_highschool': request.POST.get('highschool', ''),
        'v_last_school': request.POST.get('last_school', ''),
        'v_father_name': request.POST.get('father_name', ''),
        'v_father_occupation': request.POST.get('father_occupation', ''),
        'v_mother_name': request.POST.get('mother_name', ''),
        'v_mother_occupation': request.POST.get('mother_occupation', ''),
        'v_semester': request.POST.get('semester', '1st Semester'),
        'v_school_year': request.POST.get('school_year', '2025-2026'),
    })


@_vpsea_required
def vpsea_student_edit(request, pk):
    from .models import StudentProfile, Application, ApplicationDocument
    from .constants import CIVIL_STATUSES, SEMESTERS, STUDENT_LEVELS
    try:
        profile = StudentProfile.objects.select_related('user').get(pk=pk)
    except StudentProfile.DoesNotExist:
        return redirect('/vpsea/students/')

    # Get the latest Academic application for this student (for form_data + documents)
    app = Application.objects.filter(
        student=profile, scholarship__type='Academic'
    ).prefetch_related('documents').order_by('-submitted_at').first()

    errors = []
    if request.method == 'POST':
        p = request.POST
        f = request.FILES
        u = profile.user
        new_email = p.get('email', '').strip()
        if new_email != u.email and User.objects.filter(email=new_email).exists():
            errors.append('Email already in use by another account.')
        new_sid = p.get('student_id', '').strip()
        if new_sid != profile.student_id and StudentProfile.objects.filter(student_id=new_sid).exists():
            errors.append('Student ID already in use.')
        if not errors:
            u.first_name = p.get('first_name', '')
            u.last_name = p.get('last_name', '')
            u.email = new_email
            u.username = new_email
            if p.get('password'):
                u.set_password(p.get('password'))
            u.save()
            profile.student_id = new_sid
            profile.middle_name = p.get('middle_name', profile.middle_name).strip()
            profile.suffix = p.get('suffix', profile.suffix).strip()
            profile.school = p.get('school', profile.school)
            profile.course = p.get('course', '')
            profile.year_level = int(p.get('year_level', 1))
            profile.gwa = float(p.get('gwa', 0) or 0)
            profile.contact_number = p.get('contact_number', '')
            profile.barangay = p.get('barangay', '')
            profile.municipality = p.get('municipality', '')
            profile.province = p.get('province', '')
            profile.date_of_birth = p.get('date_of_birth') or None
            profile.gender = p.get('gender', '')
            profile.family_income = float(p.get('family_income', 0) or 0)
            for field, value in _enrollment_fields(p, profile).items():
                setattr(profile, field, value)
            profile.save()
            # Update form_data on the application
            if app:
                form_data = dict(app.form_data)
                for k, v in p.items():
                    if k not in ('csrfmiddlewaretoken',) + STUDENT_RECORD_FIELDS:
                        form_data[k] = v
                app.form_data = form_data
                app.save()
                # Replace documents if new files uploaded
                doc_fields = ['doc_certificate_of_grades', 'doc_certificate_of_enrollment',
                              'doc_prospectus', 'doc_id_photo', 'doc_application_form']
                for field in doc_fields:
                    uploaded = f.get(field)
                    if uploaded:
                        label = field.replace('doc_', '').replace('_', ' ').title()
                        app.documents.filter(name=label).delete()
                        ApplicationDocument.objects.create(
                            application=app, name=label, file=uploaded,
                        )
            return redirect('/vpsea/students/?edited=1')
    fd = request.POST if request.method == 'POST' else {}
    afd = (app.form_data or {}) if app else {}
    import json
    return render(request, 'vpsea/student_form.html', {
        'errors': errors, 'action': 'Edit',
        'profile': profile, 'app': app,
        # A student who never applied has no GWA to validate, no term to record
        # and no documents on file. Showing those sections asks the office to
        # fill in an application that does not exist.
        'show_application_fields': app is not None,
        'cancel_url': request.GET.get('next') or '/vpsea/students/',
        'form_data': fd,
        'doc_list': _doc_list(),
        'bipsu_schools': BIPSU_SCHOOLS, 'bipsu_courses_json': json.dumps(BIPSU_COURSES),
        'v_first_name': fd.get('first_name', profile.user.first_name),
        'v_last_name': fd.get('last_name', profile.user.last_name),
        'v_middle_name': fd.get('middle_name', profile.middle_name),
        'v_suffix': fd.get('suffix', profile.suffix),
        'v_school': fd.get('school', profile.school),
        'v_email': fd.get('email', profile.user.email),
        'v_student_id': fd.get('student_id', profile.student_id),
        'v_course': fd.get('course', profile.course),
        'v_year_level': fd.get('year_level', str(profile.year_level)),
        'v_gwa': fd.get('gwa', str(profile.gwa)),
        'v_gender': fd.get('gender', profile.gender),
        'v_date_of_birth': fd.get('date_of_birth', profile.date_of_birth.strftime('%Y-%m-%d') if profile.date_of_birth else ''),
        'v_contact_number': fd.get('contact_number', profile.contact_number),
        'v_barangay': fd.get('barangay', profile.barangay),

        'v_municipality': fd.get('municipality', profile.municipality),

        'v_province': fd.get('province', profile.province),
        'v_family_income': fd.get('family_income', str(profile.family_income)),
        'v_birth_place': fd.get('birth_place', profile.birth_place),
        'v_level': fd.get('level', profile.level),
        'v_department': fd.get('department', profile.department),
        'v_curriculum': fd.get('curriculum', profile.curriculum),
        'v_learner_ref_no': fd.get('learner_ref_no', profile.learner_ref_no),
        'v_entry_period': fd.get('entry_period', profile.entry_period),
        'v_entry_date': fd.get('entry_date', profile.entry_date.strftime('%Y-%m-%d') if profile.entry_date else ''),
        'v_exam_score': fd.get('exam_score', '' if profile.exam_score is None else profile.exam_score),
        'civil_statuses': CIVIL_STATUSES, 'student_levels': STUDENT_LEVELS,
        'semesters': SEMESTERS,
        'v_elementary': fd.get('elementary', afd.get('elementary', '')),
        'v_highschool': fd.get('highschool', afd.get('highschool', '')),
        'v_last_school': fd.get('last_school', afd.get('last_school', '')),
        'v_father_name': fd.get('father_name', afd.get('father_name', '')),
        'v_elementary': fd.get('elementary', profile.elementary or afd.get('elementary', '')),
        'v_highschool': fd.get('highschool', profile.highschool or afd.get('highschool', '')),
        'v_last_school': fd.get('last_school', profile.last_school or afd.get('last_school', '')),
        'v_father_name': fd.get('father_name', profile.father_name or afd.get('father_name', '')),
        'v_father_occupation': fd.get('father_occupation', profile.father_occupation or afd.get('father_occupation', '')),
        'v_mother_name': fd.get('mother_name', profile.mother_name or afd.get('mother_name', '')),
        'v_mother_occupation': fd.get('mother_occupation', profile.mother_occupation or afd.get('mother_occupation', '')),
        'v_semester': fd.get('semester', afd.get('semester', '1st Semester')),
        'v_school_year': fd.get('school_year', afd.get('school_year', '2025-2026')),
        # Extra profile fields for the read-only personal info panel (Edit only)
        'v_civil_status': fd.get('civil_status', profile.civil_status),
        'v_citizenship': profile.citizenship,
        'v_household_size': profile.household_size,
        'v_year_first_enrolled': profile.year_first_enrolled,
        'v_is_listahanan': profile.is_listahanan_household,
        'v_is_4ps': profile.is_4ps_beneficiary,
        'v_has_previous_degree': profile.has_previous_degree,
        'v_disability_type': profile.disability_type,
        'v_is_pwd': profile.is_pwd,
        'v_indigenous_group': profile.indigenous_group,
        'v_shs_gpa': profile.shs_gpa,
        'v_suc_exam_score': profile.suc_exam_display or profile.suc_exam_score,
        'v_is_tes_beneficiary': profile.is_tes_beneficiary,
    })


# The posted keys the office student form writes onto the profile itself.
# Everything else it posts is application form_data, so this list is what keeps
# a profile column from being copied into the award as well.
STUDENT_RECORD_FIELDS = (
    'first_name', 'last_name', 'email', 'password', 'student_id',
    'middle_name', 'suffix', 'birth_place', 'civil_status',
    'date_of_birth', 'gender', 'contact_number',
    'barangay', 'municipality', 'province',
    'school', 'course', 'level', 'department', 'curriculum', 'year_level',
    'learner_ref_no', 'entry_period', 'entry_date', 'exam_score', 'gwa',
    'family_income',
)


def _enrollment_fields(p, profile=None):
    """The Enrollment Data and personal columns the office form posts.

    Returned as a dict so the add and edit views set the same things the same
    way; a blank number stays blank rather than becoming a zero the office never
    typed. ``profile`` supplies the current value for a field the form omitted.
    """
    def current(name, default=''):
        return getattr(profile, name, default) if profile is not None else default

    def number(name):
        raw = (p.get(name) or '').strip()
        if raw == '':
            return current(name, None)
        try:
            return float(raw)
        except ValueError:
            return current(name, None)

    return {
        'birth_place': p.get('birth_place', current('birth_place')),
        'civil_status': p.get('civil_status', current('civil_status')),
        'level': p.get('level', current('level')),
        'department': p.get('department', current('department')),
        'curriculum': p.get('curriculum', current('curriculum')),
        'learner_ref_no': p.get('learner_ref_no', current('learner_ref_no')),
        'entry_period': p.get('entry_period', current('entry_period')),
        'entry_date': p.get('entry_date') or current('entry_date', None),
        'exam_score': number('exam_score'),
    }


def _save_column_values(request, base_url):
    """Save the values typed into a programme's custom columns.

    One post for the whole table: the office fills a column down the page, not a
    cell at a time. Field names carry which record each box belongs to —
    ``extra__<kind>__<pk>__<column key>`` — because the three record shapes have
    separate id spaces and a bare pk would not say which table to look in.

    Nothing here is taken on trust: a field naming a column that is not a custom
    one, or a record shape that is not one of the three, is dropped rather than
    written. The two portals wrap this in their own permission check.
    """
    from . import scholar_columns
    from .models import AffirmativeStaffApplication, Application, ImportedScholar
    from urllib.parse import quote

    stype = request.POST.get('type', '')
    back = f'{base_url}?type={quote(stype)}' if stype else base_url
    if request.method != 'POST':
        return redirect(back)

    models_by_kind = {
        'award': Application,
        'imported': ImportedScholar,
        'staff': AffirmativeStaffApplication,
    }
    # Grouped by record first, so a row with three custom columns is written once.
    edits = {}
    for field, value in request.POST.items():
        parts = field.split('__')
        if len(parts) != 4 or parts[0] != 'extra':
            continue
        _, kind, pk, key = parts
        if kind not in models_by_kind or not pk.isdigit():
            continue
        if not key.startswith(scholar_columns.CUSTOM_PREFIX):
            continue
        edits.setdefault((kind, int(pk)), {})[key] = value.strip()

    saved = 0
    for (kind, pk), values in edits.items():
        record = models_by_kind[kind].objects.filter(pk=pk).first()
        if record is None:
            continue
        current = scholar_columns.extra_values(record)
        if all(current.get(key, '') == value for key, value in values.items()):
            continue          # nothing typed changed — no write, no updated_at bump
        scholar_columns.set_extra_values(record, values)
        saved += 1

    sy = request.POST.get('sy', '')
    if sy:
        back += f'&sy={quote(sy)}'
    return redirect(f'{back}&columns_saved={saved}')


@_vpsea_required
def vpsea_archive_columns(request):
    """The SDSO portal's save for custom column values."""
    return _save_column_values(request, '/vpsea/archives/')


def _doc_list():
    return [
        ('doc_certificate_of_grades',   'Certificate Of Grades',   'Official COG from the Registrar for the previous semester.'),
        ('doc_certificate_of_enrollment', 'Certificate Of Enrollment', 'Official COE from the Registrar for the current semester.'),
        ('doc_prospectus',              'Prospectus',              'Program prospectus or subject checklist showing enrolled subjects.'),
        ('doc_id_photo',                'Id Photo',                'Recent 2Ã—2 ID photo with white background.'),
        ('doc_application_form',        'Application Form',        'Signed and accomplished scholarship application form.'),
    ]


@_vpsea_required
def vpsea_student_delete(request, pk):
    from .models import StudentProfile
    if request.method == 'POST':
        try:
            profile = StudentProfile.objects.select_related('user').get(pk=pk)
            profile.user.delete()  # cascades to profile
        except StudentProfile.DoesNotExist:
            pass
        return redirect('/vpsea/students/?deleted=1')
    return redirect('/vpsea/students/')


@_vpsea_required
def vpsea_scholarships(request):
    scholarships = Scholarship.objects.all().order_by('type')
    return render(request, 'vpsea/scholarships.html', {
        'scholarships': scholarships,
        'added': request.GET.get('added'),
        'saved': request.GET.get('saved'),
    })


def _column_picker_context(posted=None, scholarship=None):
    """What the archive-table column picker needs to draw itself.

    Reads a rejected submission back off ``posted`` rather than the saved
    programme, so a form that comes back with an error still shows the boxes as
    the officer left them.
    """
    from . import scholar_columns

    if posted is not None:
        chosen = scholar_columns.clean_choice(posted.getlist('table_columns'))
        custom = scholar_columns.clean_custom(posted.getlist('extra_columns'))
    else:
        chosen = scholar_columns.clean_choice(getattr(scholarship, 'table_columns', None))
        custom = list(getattr(scholarship, 'extra_columns', None) or [])
    return {
        'column_catalogue': [{'key': k, 'label': l} for k, l in scholar_columns.COLUMNS],
        'chosen_columns': chosen,
        'custom_columns': custom,
    }


@_vpsea_required
def vpsea_scholarship_add(request):
    errors = []
    if request.method == 'POST':
        p = request.POST
        name = p.get('name','').strip()
        stype = p.get('type','').strip()
        description = p.get('description','').strip()
        background = p.get('background','').strip()
        eligibility_list = [l.strip() for l in p.get('eligibility_list','').splitlines() if l.strip()]
        benefits = [l.strip() for l in p.get('benefits','').splitlines() if l.strip()]
        if not name: errors.append('Name is required.')
        if not stype: errors.append('Type is required.')
        if not errors:
            Scholarship.objects.create(
                name=name, type=stype, category='application',
                description=description, eligibility='',
                background=background, eligibility_list=eligibility_list,
                benefits=benefits, group=p.get('group','internal'), is_active=True,
                table_columns=scholar_columns.clean_choice(p.getlist('table_columns')),
                extra_columns=scholar_columns.clean_custom(p.getlist('extra_columns')),
            )
            return redirect('/vpsea/scholarships/?added=1')
    return render(request, 'vpsea/scholarship_form.html', {
        'action': 'Add', 'errors': errors, 'form': request.POST,
        **_column_picker_context(request.POST if request.method == 'POST' else None),
    })


@_vpsea_required
def vpsea_scholarship_edit(request, pk):
    try:
        s = Scholarship.objects.get(pk=pk)
    except Scholarship.DoesNotExist:
        return redirect('/vpsea/scholarships/')
    errors = []
    if request.method == 'POST':
        p = request.POST
        s.name = p.get('name','').strip()
        s.type = p.get('type','').strip()
        s.description = p.get('description','').strip()
        s.background = p.get('background','').strip()
        s.group = p.get('group', 'internal')
        s.eligibility_list = [l.strip() for l in p.get('eligibility_list','').splitlines() if l.strip()]
        s.benefits = [l.strip() for l in p.get('benefits','').splitlines() if l.strip()]
        s.table_columns = scholar_columns.clean_choice(p.getlist('table_columns'))
        s.extra_columns = scholar_columns.clean_custom(p.getlist('extra_columns'))
        if not s.name: errors.append('Name is required.')
        if not s.type: errors.append('Type is required.')
        if not errors:
            s.save()
            return redirect('/vpsea/scholarships/?saved=1')
    return render(request, 'vpsea/scholarship_form.html', {
        'action': 'Edit', 'errors': errors, 's': s,
        **_column_picker_context(
            request.POST if request.method == 'POST' else None, s),
    })


@_vpsea_required
def vpsea_scholarship_toggle(request, pk):
    if request.method == 'POST':
        Scholarship.objects.filter(pk=pk).update(
            is_active=not Scholarship.objects.get(pk=pk).is_active
        )
    return redirect('/vpsea/scholarships/')


@_vpsea_required
def vpsea_ranking(request):
    from .models import AffirmativeRecommendation, SystemSettings

    # TES is UniFAST's programme to award, so its recommender lives in their
    # portal — see unifast_tes_ranking.
    # Affirmative Action only. The BiPSU Staff scholarship has no merit test —
    # a regular appointment qualifies and nothing is scored — so ranking it
    # sorted a list by a constant. Staff applications are reviewed on the
    # Applications page instead.
    #
    # There is no applicant list here either. Nobody applies for Affirmative
    # Action: eligibility is decided from the student's own profile by
    # evaluate_and_sync below, and this page is where that is read. The tab that
    # ranked AffirmativeStaffApplication rows scored every one of them zero --
    # no form has ever written the SHS GPA or exam score it read -- so it was a
    # permanently empty table for a submission that cannot be made.
    scholarship_type = 'Affirmative'

    # ── Passing threshold (can be tweaked via GET param for VPSEA, default 75) ──
    try:
        passing_threshold = float(request.GET.get('passing', 75.0))
    except (TypeError, ValueError):
        passing_threshold = 75.0

    # ── Handle POST: re-evaluate every recommendation ───────────────────────
    # The only thing this page can be told to do. Endorsing and disqualifying by
    # hand are both gone: the award is recorded on the Archives page like every
    # other programme's, so a status set here was a second, private answer to a
    # question already written down somewhere the reports read. What a
    # recommendation says is now decided only by the rules -- evaluate_and_sync
    # still writes 'Disqualified' itself when a student stops passing them.
    if request.method == 'POST':
        if request.POST.get('action') == 'resync':
            AffirmativeRecommendation.evaluate_and_sync(passing_threshold)
        return redirect(f'/vpsea/ranking/?type={scholarship_type}&passing={passing_threshold}')

    # ── Enrolled students, evaluated by the rule-based engine ───────────────
    # Sync first so the table is always current
    AffirmativeRecommendation.evaluate_and_sync(passing_threshold)

    recommendations = (
        AffirmativeRecommendation.objects
        .select_related('student__user', *STUDENT_DETAILS)
        .order_by('-fit_score', 'student__user__last_name')
    )

    # Build per-row rule breakdown from live profile data
    rec_rows = []
    for i, rec in enumerate(recommendations):
        p = rec.student
        gpa_pass   = p.shs_gpa is not None and p.shs_gpa >= passing_threshold
        exam_pass  = p.suc_exam_percent is not None and p.suc_exam_percent >= 50.0
        not_tes    = not p.is_tes_beneficiary
        eligible   = gpa_pass and exam_pass and not_tes
        rec_rows.append({
            'rank': i + 1 if eligible else None,
            'rec': rec,
            'profile': p,
            'gpa_pass': gpa_pass,
            'exam_pass': exam_pass,
            'not_tes': not_tes,
            'eligible': eligible,
        })

    # Sort: eligible first (by fit_score desc), then ineligible
    rec_rows.sort(key=lambda r: (0 if r['eligible'] else 1, -r['rec'].fit_score))
    # Re-assign ranks only to eligible rows
    rank_counter = 1
    for row in rec_rows:
        if row['eligible']:
            row['rank'] = rank_counter
            rank_counter += 1
        else:
            row['rank'] = None

    eligible_count   = sum(1 for r in rec_rows if r['eligible'])
    ineligible_count = sum(1 for r in rec_rows if not r['eligible'])

    return render(request, 'vpsea/ranking.html', {
        'rec_rows': rec_rows,
        'passing_threshold': passing_threshold,
        'eligible_count': eligible_count,
        'ineligible_count': ineligible_count,
    })


# The profile fields the TES form fills itself in from. They are editable here
# and saved back to the profile when the application is — one record of each
# fact, corrected wherever the student happens to be looking at it.
#
# On the profile page these lock after the first save, because an edit there
# would silently change a record the office has already reviewed. That reason
# does not hold here: this form only opens while the application is undecided,
# so nothing has been reviewed yet. Same rule, applied where it means something.
TES_PROFILE_FIELDS = [
    'student_id', 'middle_name', 'suffix', 'gender', 'year_level',
    'father_last_name', 'father_first_name', 'father_middle_name',
    'mother_last_name', 'mother_first_name', 'mother_middle_name',
]

# Of those, the ones CHED marks Required on the Annex 1. The father's names are
# deliberately absent: CHED marks them optional, and a student raised by one
# parent should not be stopped by a box they cannot honestly fill.
TES_REQUIRED_FIELDS = [
    ('student_id', 'Student ID'),
    ('last_name', 'Last name'),
    ('first_name', 'Given name'),
    ('gender', 'Sex'),
    ('year_level', 'Year level'),
    ('mother_last_name', "Mother's last name"),
    ('mother_first_name', "Mother's given name"),
]


def _tes_profile_errors(posted, profile):
    """What the form is missing, and anything it cannot be given."""
    errors = []
    blank = [label for field, label in TES_REQUIRED_FIELDS
             if not (posted.get(field) or '').strip()]
    if blank:
        errors.append('CHED requires ' + ', '.join(blank) + ' on the TES form. '
                      'Fill them in above — they are saved to your profile too.')

    year = (posted.get('year_level') or '').strip()
    if year and (not year.isdigit() or not 1 <= int(year) <= 6):
        errors.append('Year level must be a number from 1 to 6.')

    # The student number is the key the office matches against its enrolment
    # list, so two students cannot share one.
    student_id = (posted.get('student_id') or '').strip()
    if student_id and profile is not None and StudentProfile.objects.filter(
            student_id=student_id).exclude(pk=profile.pk).exists():
        errors.append(f'Student ID {student_id} belongs to another account. '
                      'Check it for a typo, or ask the SDSO office to sort it out.')
    return errors


def _tes_profile_values(profile):
    """What the form's profile half shows when it is first opened."""
    if profile is None:
        return {}
    values = {field: str(getattr(profile, field, '') or '')
              for field in TES_PROFILE_FIELDS}
    values['first_name'] = profile.user.first_name or ''
    values['last_name'] = profile.user.last_name or ''
    return values


def _save_tes_profile_fields(profile, posted):
    """Write the form's profile half back, so there is still one copy of each."""
    user = profile.user
    user.first_name = (posted.get('first_name') or '').strip()
    user.last_name = (posted.get('last_name') or '').strip()
    user.save(update_fields=['first_name', 'last_name'])

    for field in TES_PROFILE_FIELDS:
        value = (posted.get(field) or '').strip()
        if field == 'year_level':
            profile.year_level = int(value) if value.isdigit() else profile.year_level
        else:
            setattr(profile, field, value)
    profile.save()


@login_required(login_url='/login/')
def student_apply_tes(request):
    profile = StudentProfile.objects.filter(user=request.user).first()
    from .constants import EDITABLE_REVIEW_STATUSES
    from . import annex1_report
    existing = TESApplication.objects.filter(student=profile).first() if profile else None
    # Undecided means still the applicant's to correct; a decided one is final.
    editing = existing if existing and existing.status in EDITABLE_REVIEW_STATUSES else None
    errors = []

    # TES sits alongside an Academic scholarship and nothing else. A student who
    # already holds TDP, DOST, CHED or any other award cannot add it, and the
    # form says so rather than taking a submission UniFAST would have to refuse.
    # Their own approved TES is not a block here — it is what `existing` shows.
    if existing and existing.status == 'Approved':
        blocked_reason = ''
    else:
        blocked_reason = scholarship_block_reason(
            profile, 'TES', 'Tertiary Education Subsidy')
    if blocked_reason:
        return render(request, 'student/apply_tes.html', {
            'profile': profile, 'blocked': True, 'blocked_reason': blocked_reason,
            'existing': existing, 'errors': [], 'post': {},
        })

    if request.method == 'POST' and profile and (editing or not existing):
        p = request.POST
        errors.extend(_tes_profile_errors(p, profile))

        # The dropdown is the registry, so anything else reaching here came from
        # a hand-made post or a template that could not be read. Checked only
        # when there is a list to check against — with the workbook missing the
        # form falls back to a text box and this would reject everything.
        registry = annex1_report.registry_programs()
        program = p.get('complete_program', '').strip()
        if registry and program not in registry:
            errors.append(
                'Pick your programme from the list — CHED reads the Annex 1 '
                'against its own registry of programme names.')

        # 'Other' is this system's word for a condition CHED's list does not
        # name; what reaches the workbook is what the student typed instead.
        disability = p.get('disability_type', '').strip()
        if disability == annex1_report.OTHER:
            disability = p.get('disability_type_other', '').strip()
            if not disability:
                errors.append('Name the disability you chose "Other" for.')

        if not errors:
            # The profile half is saved to the profile, not copied onto the
            # application: the names live in exactly one place, and this form is
            # another window onto them rather than a second record of them.
            _save_tes_profile_fields(profile, p)
            # The same question My Profile asks, off the same CHED list, so the
            # answer is stored once. Without this the two records could — and
            # did — disagree about the same student.
            profile.disability_type = disability
            profile.save(update_fields=['disability_type'])
            TESApplication.objects.update_or_create(
                student=profile,
                defaults=dict(
                lrn=p.get('lrn', ''),
                philsys_id=p.get('philsys_id', '').strip(),
                four_ps_id=p.get('four_ps_id', '').strip(),
                birthdate=p.get('birthdate') or None,
                complete_program=program,
                street_barangay=p.get('street_barangay', ''),
                city_municipality=p.get('city_municipality', ''),
                province=p.get('province', ''),
                region=p.get('region', ''),
                zip_code=p.get('zip_code', ''),
                contact_number=p.get('contact_number', ''),
                email_address=p.get('email_address', ''),
                disability_type=disability or 'N/A',
                is_solo_parent_dependent=p.get('is_solo_parent_dependent') == '1',
                is_first_gen_college=p.get('is_first_gen_college') == '1',
                indigenous_people_group=p.get('indigenous_people_group', 'Not Applicable'),
                ),
            )
            return redirect('/student/apply/tes/?submitted=1')

    # Grouped under the school that offers them, in the schools' own order, and
    # alphabetical inside each. Forty entries that all open with the same four
    # words are unscannable as one list; a student knows their own school.
    from .constants import GENDERS, group_programs_by_school
    programs = annex1_report.registry_programs()
    program_groups = group_programs_by_school(programs)
    disabilities = annex1_report.disability_types()
    # A saved value that is not on CHED's list is one somebody typed under
    # 'Other'. The form has to come back showing it that way, or editing an
    # application would silently drop what they wrote.
    saved_disability = ((existing.disability_type if existing
                         else (profile.disability_type if profile else '')) or '').strip()
    custom_disability = (saved_disability
                         if saved_disability and saved_disability not in disabilities
                         else '')

    return render(request, 'student/apply_tes.html', {
        'profile': profile,
        'existing': existing,
        'editing': editing,
        'errors': errors,
        'programs': programs,
        'program_groups': program_groups,
        'program_count': len(programs),
        'genders': GENDERS,
        'disabilities': disabilities,
        'other_option': annex1_report.OTHER,
        'submitted': request.GET.get('submitted'),
        'enrolled': _is_enrolled(profile),
        # On a redisplay the boxes hold what was typed, not what is stored —
        # otherwise a field the student deliberately cleared would come back
        # filled in from the profile they were trying to correct.
        'post': request.POST if request.method == 'POST' else dict(
            _tes_profile_values(profile),
            # Outside the `existing` block: a first-time applicant's disability
            # comes off their profile, where registration put it.
            disability_type=annex1_report.OTHER if custom_disability else saved_disability,
            disability_type_other=custom_disability,
            **(
            {
                'lrn': existing.lrn,
                'philsys_id': existing.philsys_id,
                'four_ps_id': existing.four_ps_id,
                'birthdate': existing.birthdate.strftime('%Y-%m-%d') if existing.birthdate else '',
                'complete_program': existing.complete_program,
                'street_barangay': existing.street_barangay,
                'city_municipality': existing.city_municipality,
                'province': existing.province,
                'region': existing.region,
                'zip_code': existing.zip_code,
                'contact_number': existing.contact_number,
                'email_address': existing.email_address,
                'is_solo_parent_dependent': '1' if existing.is_solo_parent_dependent else '0',
                'is_first_gen_college': '1' if existing.is_first_gen_college else '0',
                'indigenous_people_group': existing.indigenous_people_group,
            } if existing else {})
        ),
    })


@_unifast_required
def unifast_tes_ranking(request):
    """Rule-based TES ranking and recommendation, for the office that awards it.

    Reads the student records UniFAST already holds and concludes nothing it
    cannot evidence. api/tes_ranking.py carries the rules, the field each one
    reads, and the reason it reports back.
    """
    from . import tes_ranking

    # Students still waiting on a decision — the ones this list exists to
    # choose between. Ranking every profile on file put people who never asked
    # for TES onto an award list; ranking approved and rejected applicants put
    # people there whose case is already closed.
    applicants = (
        StudentProfile.objects
        .filter(tes_applications__status='Pending')
        .select_related('user', *StudentProfile.DETAIL_RELATIONS)
        .prefetch_related('tes_applications')
        .distinct()
    )
    evaluations = tes_ranking.rank(applicants)

    # The ranked list holds only applicants whose record is complete. A rank is
    # a position against other students, and an incomplete record cannot support
    # one — so those applicants are held in their own list instead of being
    # given a number, and instead of being dropped, which would lose exactly the
    # people who need chasing. Not Eligible stays out of both lists and is
    # reported in the summary cards.
    ranked = [e for e in evaluations if e.status == tes_ranking.ELIGIBLE]
    for position, evaluation in enumerate(ranked, start=1):
        evaluation.rank = position
    needs_info = [e for e in evaluations if e.status == tes_ranking.FOR_VERIFICATION]

    return render(request, 'unifast/tes_ranking.html', {
        'active': 'tes_ranking',
        'tes_rows': ranked,
        'tes_needs_info': needs_info,
        'tes_applicant_total': len(evaluations),
        'tes_counts': {
            'eligible': sum(1 for e in evaluations if e.status == tes_ranking.ELIGIBLE),
            'verification': sum(1 for e in evaluations if e.status == tes_ranking.FOR_VERIFICATION),
            'not_eligible': sum(1 for e in evaluations if e.status == tes_ranking.NOT_ELIGIBLE),
            'priority_1': sum(1 for e in evaluations if e.priority == tes_ranking.PRIORITY_1),
            'priority_2': sum(1 for e in evaluations if e.priority == tes_ranking.PRIORITY_2),
        },
    })


@_unifast_required
def unifast_tes_applications(request):
    from .constants import DECIDED_REVIEW_STATUSES
    from urllib.parse import quote
    if request.method == 'POST':
        app_id = request.POST.get('app_id')
        new_status = request.POST.get('status')
        remarks = request.POST.get('remarks', '')
        # Same rule as the review screen below: a decision already recorded is
        # not overwritten from here either.
        posted = TESApplication.objects.filter(id=app_id).first()
        if posted and posted.status in DECIDED_REVIEW_STATUSES:
            return redirect('/unifast/tes-applications/?error=' + quote(
                f'That application was already {posted.status.lower()}. '
                'A decision is made once.'))
        TESApplication.objects.filter(id=app_id).update(status=new_status, remarks=remarks)
        return redirect('/unifast/tes-applications/?saved=1')

    import os
    from . import annex1_report, tes_report

    # The school year scopes both the review list and the Annex 1 report built
    # from it, so what an officer generates is exactly what they are looking at.
    # Blank means every year: this queue is read across terms, and a legacy row
    # whose term was never stamped would otherwise be invisible here.
    school_year = request.GET.get('sy', '').strip()
    apps = TESApplication.objects.select_related('student__user', *STUDENT_DETAILS).order_by('-submitted_at')
    if school_year:
        apps = apps.filter(school_year=school_year)
    return render(request, 'unifast/tes_applications.html', {
        'applications': apps,
        'total': apps.count(),
        'pending_count': apps.filter(status='Pending').count(),
        'approved_count': apps.filter(status='Approved').count(),
        'school_year': school_year,
        'school_years': tes_report.school_year_options(),
        'annex1_headers': annex1_report.ANNEX1_HEADERS,
        'annex1_rows': annex1_report.applicant_rows(school_year=school_year),
    })


@_unifast_required
def unifast_tes_applicants_report(request):
    """The list of TES applicants, as a plain table.

    Scoped to the school year picked on the TES Applications page. The columns
    are the ones CHED's Annex 1 asks for, but the file is generated here rather
    than being that form filled in — see api/annex1_report.py.
    """
    from . import annex1_report

    school_year = request.GET.get('sy', '').strip()
    buf, _written = annex1_report.build_workbook(school_year)

    label = (school_year or 'all_years').replace('-', '_')
    return _xlsx_response(buf, f'TES_Applicants_{label}.xlsx')


@_unifast_required
def unifast_tes_review(request, pk):
    """Record UniFAST's decision on one TES application.

    The decision is made once: an application already approved or rejected
    keeps that status, and posting a different one is refused. A later save
    can still correct the award number and remarks, because CHED issues the
    award number after the decision and the billing report is built on it —
    locking that away with the status would strand a typo the office has no
    other way to fix.
    """
    from .constants import DECIDED_REVIEW_STATUSES
    from urllib.parse import quote
    if request.method == 'POST':
        tes_app = TESApplication.objects.select_related('student__user', *STUDENT_DETAILS).filter(pk=pk).first()
        if not tes_app:
            return redirect('/unifast/tes-applications/?error=' + quote(
                'That application was not found.'))
        new_status = request.POST.get('status', 'Pending')
        remarks = request.POST.get('remarks', '')
        award_number = request.POST.get('award_number', '')
        already_decided = tes_app.status in DECIDED_REVIEW_STATUSES
        if already_decided and new_status != tes_app.status:
            return redirect('/unifast/tes-applications/?error=' + quote(
                f'That application was already {tes_app.status.lower()}. '
                'A decision is made once.'))
        TESApplication.objects.filter(pk=pk).update(
            status=new_status,
            remarks=remarks,
            award_number=award_number,
        )
        # Only the first decision is announced. A later award-number correction
        # is bookkeeping, not news the applicant should be told a second time.
        if not already_decided:
            notify.decision(
                tes_app.student, 'Your Tertiary Education Subsidy application',
                new_status, remarks,
                detail=f'Award number: {award_number}' if award_number else '',
                link='/student/applications/',
            )
        if new_status == 'Approved':
            from .models import SystemSettings
            try:
                tes_app = TESApplication.objects.select_related('student').get(pk=pk)
                scholarship = Scholarship.objects.filter(type='TES').first()
                if scholarship:
                    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
                    parsed = SystemSettings.parse_label(settings_obj.academic_year)
                    already = Application.objects.filter(
                        student=tes_app.student, scholarship=scholarship,
                        school_year=parsed['sy'], semester=parsed['semester'],
                    ).exists()
                    if not already:
                        Application.objects.create(
                            student=tes_app.student,
                            scholarship=scholarship,
                            status='Approved',
                            remarks=remarks,
                            source='tes_application',
                            tes_application_id=pk,
                            school_year=parsed['sy'],
                            semester=parsed['semester'],
                            award_number=award_number,
                        )
                    else:
                        # Update award_number on existing Application row if already present
                        Application.objects.filter(
                            student=tes_app.student, scholarship=scholarship,
                            school_year=parsed['sy'], semester=parsed['semester'],
                        ).update(
                            source='tes_application',
                            tes_application_id=pk,
                            award_number=award_number,
                        )
            except TESApplication.DoesNotExist:
                pass
    return redirect('/unifast/tes-applications/?saved=1')


@_unifast_required
def unifast_archive_columns(request):
    """The UniFAST portal's save for custom column values.

    The same work as the SDSO one behind this office's own permission check —
    one shared view would have let either office write the other's records.
    """
    return _save_column_values(request, '/unifast/archives/')


@_unifast_required
def unifast_archives(request):
    from .models import ScholarListImport, SystemSettings, ImportedScholar
    db_types = list(Scholarship.objects.values_list('type', flat=True).distinct())
    archive_types = UNIFAST_TYPES + [t for t in db_types if t not in UNIFAST_TYPES and t not in ['Academic','DOST','CHED','CoScho','Sports','Affirmative','Staff','GSIS']]
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    active_label = settings_obj.academic_year
    parsed = SystemSettings.parse_label(active_label)
    active_sy = parsed['sy']
    active_semester = parsed['semester']
    stype = request.GET.get('type', 'TDP')
    if stype not in archive_types:
        stype = archive_types[0]
    history = ScholarListImport.objects.filter(scholarship_type=stype).order_by('-created_at')
    all_labels = list(ScholarListImport.objects.filter(scholarship_type=stype).values_list('term_label', flat=True).distinct().order_by('-term_label'))
    ar_labels = list(ImportedScholar.objects.exclude(term_label='').filter(scholarship_type=stype).values_list('term_label', flat=True).distinct())
    for lbl in ar_labels:
        if lbl not in all_labels:
            all_labels.append(lbl)
    all_labels = sorted(set(all_labels), reverse=True)
    if active_label not in all_labels:
        all_labels.insert(0, active_label)
    all_sy_display = [(lbl, SystemSettings.parse_label(lbl)['sy'] + ' - ' + SystemSettings.parse_label(lbl)['semester']) for lbl in all_labels]
    selected_label = request.GET.get('sy', active_label)
    if selected_label not in all_labels:
        selected_label = active_label
    yy, s = active_label.split('-')
    prev_label = f'{yy}-1' if s == '2' else f'{int(yy)-1}-2'
    prev_parsed = SystemSettings.parse_label(prev_label)
    next_label = settings_obj.next_label()
    imported_rows = ImportedScholar.objects.filter(
        scholarship_type=stype, term_label=selected_label, claimed_by__isnull=True,
    ).order_by('last_name', 'first_name')

    def sy_filter(qs):
        # See the matching helper in vpsea_archives — the term is a column now.
        return qs.filter(term_label=selected_label)

    scholars = sy_filter(
        Application.objects.filter(status='Approved', scholarship__type=stype)
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name').distinct()
    return render(request, 'unifast/archives.html', {
        'archive_types': archive_types,
        'active_type': stype,
        **_scholar_groups(stype, [
            (None, list(scholars) + list(imported_rows),
             f'No approved {stype} scholars yet.'),
        ], portal='unifast'),
        'total': scholars.count() + imported_rows.count(),
        'history': history,
        'all_sy_display': all_sy_display,
        'selected_sy': selected_label,
        'active_sy': active_label,
        'active_sy_display': active_sy + ' - ' + active_semester,
        'active_semester': active_semester,
        'next_sy': next_label,
        'prev_sy': prev_label,
        'prev_sy_display': prev_parsed['sy'] + ' - ' + prev_parsed['semester'],
    })


@_unifast_required
def unifast_archive_add(request):
    from .models import Application, Scholarship, StudentProfile, User, SystemSettings, ApplicationDocument
    if request.method != 'POST':
        return redirect('/unifast/archives/')
    p = request.POST
    f = request.FILES
    stype = p.get('scholarship_type', 'TDP')
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    active_sy = parsed['sy']
    active_semester = parsed['semester']
    email = p.get('email', '').strip()
    student_id = p.get('student_id', '').strip()
    if not email:
        email = f"{student_id}@bipsu.edu.ph"
    user = User.objects.filter(email=email).first()
    if not user:
        user = User.objects.create_user(
            username=email, email=email,
            password=student_id or 'bipsu1234',
            first_name=p.get('first_name', ''),
            last_name=p.get('last_name', ''),
            role='student',
        )
    profile = StudentProfile.objects.filter(user=user).first()
    if not profile:
        from .constants import school_for_course
        profile = StudentProfile.objects.create(
            user=user,
            student_id=student_id,
            middle_name=p.get('middle_name', '').strip(),
            school=school_for_course(p.get('course', '')),
            course=p.get('course', ''),
            year_level=int(p.get('year_level', 1) or 1),
            gwa=float(p.get('gwa', 0) or 0),
            gender=p.get('gender', ''),
            barangay=p.get('barangay', ''),
            municipality=p.get('municipality', ''),
            province=p.get('province', ''),
            contact_number=p.get('contact_number', ''),
            date_of_birth=p.get('date_of_birth') or None,
        )
    else:
        profile.course = p.get('course', profile.course)
        profile.year_level = int(p.get('year_level', profile.year_level) or profile.year_level)
        profile.gender = p.get('gender', profile.gender)
        profile.barangay = p.get('barangay', profile.barangay)
        profile.municipality = p.get('municipality', profile.municipality)
        profile.province = p.get('province', profile.province)
        profile.save()
    scholarship = Scholarship.objects.filter(type=stype).first()
    if scholarship:
        award_fields = {
            'source': 'import',
            'school_year': active_sy,
            'semester': active_semester,
            'award_number': p.get('award_number', ''),
            'congress_district': p.get('congress_district', ''),
        }
        already = Application.objects.filter(
            student=profile, scholarship=scholarship,
            school_year=active_sy, semester=active_semester,
        ).exists()
        if not already:
            app = Application.objects.create(
                student=profile, scholarship=scholarship,
                status='Approved', **award_fields,
            )
        else:
            app = Application.objects.filter(
                student=profile, scholarship=scholarship,
                school_year=active_sy, semester=active_semester,
            ).first()
        for field, label in [
            ('doc_certificate_of_grades', 'Certificate Of Grades'),
            ('doc_certificate_of_enrollment', 'Certificate Of Enrollment'),
            ('doc_prospectus', 'Prospectus'),
            ('doc_id_photo', 'Id Photo'),
            ('doc_application_form', 'Application Form'),
            ('proof_document', 'Proof Document'),
        ]:
            uploaded = f.get(field)
            if uploaded:
                ApplicationDocument.objects.create(application=app, name=label, file=uploaded)
    return redirect(f'/unifast/archives/?type={stype}&added=1')


@_unifast_required
def unifast_archive_edit(request, pk):
    from .models import Application
    if request.method != 'POST':
        return redirect('/unifast/archives/')
    p = request.POST
    stype = p.get('scholarship_type', 'TDP')
    try:
        app = Application.objects.select_related('student__user', *STUDENT_DETAILS).get(pk=pk)
    except Application.DoesNotExist:
        return redirect(f'/unifast/archives/?type={stype}')
    profile = app.student
    user = profile.user
    user.first_name = p.get('first_name', user.first_name)
    user.last_name = p.get('last_name', user.last_name)
    user.save()
    profile.course = p.get('course', profile.course)
    profile.year_level = int(p.get('year_level', profile.year_level) or profile.year_level)
    profile.gender = p.get('gender', profile.gender)
    profile.barangay = p.get('barangay', profile.barangay)
    profile.municipality = p.get('municipality', profile.municipality)
    profile.province = p.get('province', profile.province)
    if p.get('student_id'):
        profile.student_id = p.get('student_id')
    profile.save()
    if p.get('award_number') is not None:
        app.award_number = p.get('award_number')
    if p.get('congress_district') is not None:
        app.congress_district = p.get('congress_district')
    app.save()
    return redirect(f'/unifast/archives/?type={stype}&edited=1')


@_unifast_required
def unifast_archive_delete(request, pk):
    from .models import Application
    if request.method != 'POST':
        return redirect('/unifast/archives/')
    stype = request.POST.get('scholarship_type', 'TDP')
    Application.objects.filter(pk=pk).delete()
    return redirect(f'/unifast/archives/?type={stype}&deleted=1')


@_unifast_required
def unifast_archive_import(request):
    from django.core.files.base import ContentFile
    from .models import ScholarListImport, ActivityLog, SystemSettings, ImportedScholar
    if request.method != 'POST':
        return redirect('/unifast/archives/')
    stype = request.POST.get('type', 'TDP')
    rollover_label = request.POST.get('rollover_label', '').strip()
    file = request.FILES.get('file')
    if not file:
        return redirect(f'/unifast/archives/?type={stype}&import_error=No+file+provided')
    if not rollover_label:
        return redirect(f'/unifast/archives/?type={stype}&import_error=Rollover+name+is+required')
    try:
        import openpyxl
        settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
        parsed = SystemSettings.parse_label(settings_obj.academic_year)
        active_semester = parsed['semester']
        rollover_parsed = SystemSettings.parse_label(rollover_label) if '-' in rollover_label else {'sy': rollover_label, 'semester': active_semester}
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        col_map = COLUMN_MAPS.get(stype, COLUMN_MAPS['CoScho'])
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            try:
                int(row[0])
            except (ValueError, TypeError):
                continue
            extra = {}
            for idx, field in col_map:
                val = row[idx] if idx < len(row) else None
                extra[field] = str(val).strip() if val is not None else ''
            last = extra.get('last_name', '').strip()
            first = extra.get('first_name', '').strip()
            if not last and not first:
                continue
            try:
                year_level = int(extra.get('year', 0) or 0)
            except (ValueError, TypeError):
                year_level = 0
            try:
                gwa = float(extra.get('gwa', 0) or 0)
            except (ValueError, TypeError):
                gwa = 0.0
            records.append(ImportedScholar(
                scholarship_type=stype,
                term_label=rollover_label,
                last_name=last,
                first_name=first,
                middle_name=extra.get('middle_name', extra.get('middle_initial', '')),
                gender=extra.get('sex', ''),
                course=extra.get('course', ''),
                year_level=year_level,
                gwa=gwa,
                barangay=extra.get('barangay', ''),
                municipality=extra.get('municipality', ''),
                province=extra.get('province', ''),
                student_id=extra.get('student_number', extra.get('student_id', '')),
                award_number=extra.get('award_number', ''),
                congress_district=extra.get('congress_district', ''),
                imported_from=file.name,
            ))
        # Same replace-in-one-transaction rule as vpsea_archive_import above.
        with transaction.atomic():
            ImportedScholar.objects.filter(
                scholarship_type=stype, term_label=rollover_label).delete()
            ImportedScholar.objects.bulk_create(records)
        created = len(records)
        file.seek(0)
        if not ScholarListImport.objects.filter(term_label=rollover_label, scholarship_type=stype).exists():
            rollover = ScholarListImport(
                scholarship_type=stype,
                school_year=rollover_parsed['sy'],
                semester=rollover_parsed.get('semester', active_semester),
                term_label=rollover_label,
                scholar_count=created,
                imported_by=request.user,
            )
            rollover.excel_file.save(f'{stype}_{rollover_label}.xlsx', ContentFile(file.read()), save=True)
        else:
            ScholarListImport.objects.filter(term_label=rollover_label, scholarship_type=stype).update(scholar_count=created)
        ActivityLog.objects.create(
            user=request.user,
            action=f'Imported {file.name} ({created} rows) for {stype} as "{rollover_label}"'
        )
    except Exception as e:
        return redirect(f'/unifast/archives/?type={stype}&import_error={e}')
    return redirect(f'/unifast/archives/?type={stype}&import_ok={created}')


@_unifast_required
def unifast_rollover_delete(request, pk):
    from .models import ScholarListImport, ActivityLog
    if request.method != 'POST':
        return redirect('/unifast/archives/')
    stype = request.POST.get('type', 'TDP')
    try:
        r = ScholarListImport.objects.get(pk=pk)
    except ScholarListImport.DoesNotExist:
        return redirect(f'/unifast/archives/?type={stype}')
    removed, imported_type, label = _delete_import_with_scholars(r)
    ActivityLog.objects.create(
        user=request.user,
        action=f'Deleted the {imported_type} import for "{label}" and the '
               f'{removed} scholar row(s) it had created.'
    )
    return redirect(f'/unifast/archives/?type={stype}')


@_unifast_required
def unifast_archive_download(request):
    from django.http import HttpResponse
    from io import BytesIO
    from .models import SystemSettings
    from django.db.models import Q
    stype = request.GET.get('type', 'TDP')
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    active_sy = parsed['sy']
    semester = settings_obj.active_semester
    scholars = Application.objects.filter(
        status='Approved', scholarship__type=stype
    ).select_related('student__user', 'scholarship', *STUDENT_DETAILS).order_by('student__user__last_name')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{stype} Scholars'
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfont = Font(bold=True)
    hfill = PatternFill('solid', fgColor='D9E1F2')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    headers = ['No.', 'Last Name', 'First Name', 'Gender', 'Course', 'Year Level', 'Student ID', 'Award No.', 'Congress District', 'Barangay', 'Municipality', 'Province']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = hfont; cell.fill = hfill; cell.border = border; cell.alignment = center
    for i, app in enumerate(scholars, 1):
        p = app.student
        ws.append([i, p.user.last_name, p.user.first_name, p.gender, p.course, p.year_level,
                   p.student_id, app.award_number,
                   app.congress_district,
                   p.barangay, p.municipality, p.province])
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    filename = f'{stype}_scholars_{settings_obj.academic_year}_{semester.replace(" ", "_")}.xlsx'
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── BiPSU Staff portal ─────────────────────────────────────────────────────────

def _nsu_staff_required(view_fn):
    from functools import wraps
    @wraps(view_fn)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role != 'nsu_staff':
            return redirect('/login/')
        return view_fn(request, *args, **kwargs)
    return wrapper


def _staff_profile(user):
    """The staff member's own record, created on first visit if it is missing.

    Migration 0032 backfilled a profile for every staff account that existed
    when it ran, but accounts created since — and any created outside the
    registration form — have none, so the portal cannot assume a row is there.
    """
    from .models import StaffProfile
    profile, _ = StaffProfile.objects.get_or_create(user=user)
    return profile


def _parse_date(raw):
    """('2016-06-01' | '' | 'nonsense') -> (date | None, ok?).

    Blank clears the field, which is a valid edit; anything unparseable is
    reported back instead of being handed to the DB, which errors out on save.
    """
    from datetime import datetime
    raw = (raw or '').strip()
    if not raw:
        return None, True
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date(), True
    except ValueError:
        return None, False


def _pick(application, staff, field):
    """The application's value if it has one, else the staff profile's."""
    value = getattr(application, field, '') if application else ''
    return value or getattr(staff, field, '') or ''


def _pick_date(application, staff, field):
    """:func:`_pick` for a date field, rendered for a date input."""
    value = getattr(application, field, None) if application else None
    value = value or getattr(staff, field, None)
    return value.strftime('%Y-%m-%d') if value else ''


def _nsu_staff_enrolled(user):
    """True when the staff member has any non-rejected application — the Apply link is hidden."""
    from .models import AffirmativeStaffApplication
    return AffirmativeStaffApplication.objects.filter(
        email=user.email,
    ).exclude(status='Rejected').exists()


@_nsu_staff_required
def nsu_staff_dashboard(request):
    from .models import StaffRenewal, Notification, Announcement, AffirmativeStaffApplication
    user = request.user
    # Try to find the matching AffirmativeStaffApplication record for this staff member
    aff_app = AffirmativeStaffApplication.objects.filter(
        email=user.email, qualified_for='Staff'
    ).exclude(status='Rejected').first()
    renewals = StaffRenewal.objects.filter(staff_user=user).order_by('-submitted_at')
    announcements = Announcement.objects.order_by('-created_at')[:3]
    unread_count = Notification.objects.filter(
        student__user=user, is_read=False
    ).count() if hasattr(user, 'profile') else 0
    return render(request, 'nsu_staff/dashboard.html', {
        'aff_app': aff_app,
        'renewals': renewals,
        'announcements': announcements,
        'unread_count': unread_count,
        'pending_renewals': renewals.filter(status='Pending').count(),
        'approved_renewals': renewals.filter(status='Approved').count(),
        'enrolled': _nsu_staff_enrolled(user),
    })


@_nsu_staff_required
def nsu_staff_profile(request):
    """The staff member's own record. Employment details are edited here.

    They used to be written straight into the AffirmativeStaffApplication the
    office had reviewed. They live on StaffProfile now, so correcting a
    department or an employee ID no longer edits an approved award — the
    application is shown below, read-only, as the snapshot it is.
    """
    from .models import AffirmativeStaffApplication, StaffProfile
    from .constants import CIVIL_STATUSES, DESIGNATIONS, EMPLOYMENT_STATUSES
    user = request.user
    staff = _staff_profile(user)
    # Any status, not only Approved — a staff member whose application is still
    # Pending Validation or Needs Revision has to be able to see where it stands.
    # Latest wins if they re-applied after a rejection.
    aff_app = AffirmativeStaffApplication.objects.filter(
        email=user.email, qualified_for='Staff'
    ).order_by('-submitted_at').first()
    saved = False
    errors = []
    if request.method == 'POST':
        p = request.POST
        # Only the given and family names live on the User row; every other
        # detail belongs to the profile.
        user.first_name = p.get('first_name', user.first_name).strip()
        user.last_name  = p.get('last_name',  user.last_name).strip()
        user.save()

        # Personal
        staff.middle_name    = p.get('middle_name', staff.middle_name).strip()
        staff.suffix         = p.get('suffix', staff.suffix).strip()
        staff.gender         = p.get('gender', staff.gender)
        staff.civil_status   = p.get('civil_status', staff.civil_status)
        staff.contact_number = p.get('contact_number', staff.contact_number).strip()
        staff.barangay       = p.get('barangay', staff.barangay)
        staff.municipality   = p.get('municipality', staff.municipality)
        staff.province       = p.get('province', staff.province)

        # Employment
        employee_id = p.get('employee_id', staff.employee_id).strip()
        clash = StaffProfile.objects.filter(employee_id=employee_id).exclude(pk=staff.pk)
        if employee_id and clash.exists():
            errors.append(f'Employee ID {employee_id} is already on another staff record. '
                          'Contact the VPSEA office if that is not right.')
        else:
            staff.employee_id = employee_id
        staff.school            = p.get('school', staff.school)
        staff.department        = p.get('department', staff.department).strip()
        staff.position          = p.get('position', staff.position).strip()
        staff.employment_status = p.get('employment_status', staff.employment_status)
        staff.designation       = p.get('designation', staff.designation)
        staff.highest_education = p.get('highest_education', staff.highest_education).strip()
        staff.has_baccalaureate = 'has_baccalaureate' in p

        for field, label in (
            ('date_of_birth', 'Date of birth'),
            ('date_hired', 'Date hired'),
            ('date_of_regularization', 'Date of regularization'),
        ):
            parsed, ok = _parse_date(p.get(field, ''))
            if ok:
                setattr(staff, field, parsed)
            else:
                errors.append(f'{label} must be a valid date.')

        # Only read when the hiring date is unknown — see the model property.
        yos = p.get('years_of_service', '').strip()
        if yos == '':
            staff.declared_years_of_service = None
        elif yos.isdigit():
            staff.declared_years_of_service = int(yos)
        else:
            errors.append('Years of service must be a whole number of years.')

        if request.FILES.get('appointment_paper'):
            staff.appointment_paper = request.FILES['appointment_paper']

        # The valid fields in a submission go through even when another one was
        # rejected, so a single typo does not throw the whole form away.
        staff.save()
        saved = not errors
    return render(request, 'nsu_staff/profile.html', {
        'staff': staff,
        'aff_app': aff_app,
        'saved': saved,
        'errors': errors,
        'bipsu_schools': BIPSU_SCHOOLS,
        'civil_statuses': CIVIL_STATUSES,
        'employment_statuses': EMPLOYMENT_STATUSES,
        'designations': DESIGNATIONS,
        'enrolled': _nsu_staff_enrolled(user),
    })


@_nsu_staff_required
def nsu_staff_notifications(request):
    from .models import Notification
    user = request.user
    # Notifications are tied to StudentProfile; staff may not have one — guard gracefully
    notifications = []
    if hasattr(user, 'profile'):
        notifications = Notification.objects.filter(
            student=user.profile
        ).order_by('-created_at')
        if request.method == 'POST' and request.POST.get('mark_all_read'):
            notifications.update(is_read=True)
            return redirect('/nsu-staff/notifications/')
    return render(request, 'nsu_staff/notifications.html', {
        'notifications': notifications,
        'enrolled': _nsu_staff_enrolled(user),
    })


@_nsu_staff_required
def nsu_staff_applications(request):
    from .models import AffirmativeStaffApplication, StaffRenewal
    user = request.user
    applications = AffirmativeStaffApplication.objects.filter(
        email=user.email
    ).select_related(*STAFF_APPLICATION_DETAILS).order_by('-submitted_at')
    renewals = StaffRenewal.objects.filter(staff_user=user).order_by('-submitted_at')
    return render(request, 'nsu_staff/applications.html', {
        'applications': applications,
        'renewals': renewals,
        'enrolled': _nsu_staff_enrolled(user),
    })


@_nsu_staff_required
def nsu_staff_renewal(request):
    from .models import StaffRenewal, SystemSettings
    user = request.user
    settings_obj, _ = SystemSettings.objects.get_or_create(pk=1)
    parsed = SystemSettings.parse_label(settings_obj.academic_year)
    renewals = StaffRenewal.objects.filter(staff_user=user).order_by('-submitted_at')
    errors = []
    submitted = False

    if request.method == 'POST':
        sup = request.FILES.get('supporting_document')
        if not errors:
            StaffRenewal.objects.create(
                staff_user=user,
                supporting_document=sup or None,
            )
            return redirect('/nsu-staff/renewal/?submitted=1')
        return render(request, 'nsu_staff/renewal.html', {
            'renewals': renewals, 'errors': errors,
            'semester': parsed['semester'], 'academic_year': parsed['sy'],
            'enrolled': _nsu_staff_enrolled(user),
        })

    return render(request, 'nsu_staff/renewal.html', {
        'renewals': renewals,
        'submitted': request.GET.get('submitted'),
        'semester': parsed['semester'],
        'academic_year': parsed['sy'],
        'errors': errors,
        'enrolled': _nsu_staff_enrolled(user),
    })


@_nsu_staff_required
def nsu_staff_apply(request):
    from .models import AffirmativeStaffApplication, SystemSettings
    user = request.user
    staff = _staff_profile(user)

    # An application already decided is not applied for again.
    existing = AffirmativeStaffApplication.objects.filter(
        email=user.email, qualified_for='Staff'
    ).exclude(status='Rejected').first()

    if existing and existing.status in ('Approved', 'Pending Validation'):
        return render(request, 'nsu_staff/apply.html', {
            'blocked': True,
            'blocked_reason': f'You already have a Staff Scholarship application with status: {existing.status}.',
            'existing': existing,
            'enrolled': True,
        })

    errors = []

    if request.method == 'POST':
        p = request.POST
        f = request.FILES

        # Required field validation — skip for draft
        if True:
            required = {
                'first_name': 'First name',
                'last_name': 'Last name',
                'date_of_birth': 'Date of birth',
                'gender': 'Gender',
                'course': 'Course',
                'student_number': 'Student / Employee number',
                'employment_status': 'Employment status',
                'designation': 'Designation',
                'years_of_service': 'Years of service',
                'date_of_regularization': 'Date of regularization',
            }
            for field, label in required.items():
                if not p.get(field, '').strip():
                    errors.append(f'{label} is required.')

            # A regular appointment is the whole eligibility rule for this
            # programme — there is nothing else to qualify on.
            employment = p.get('employment_status', '').strip()
            if employment and employment != 'Regular':
                errors.append(
                    'The BiPSU Staff Scholarship is open to regular employees. '
                    f'Your appointment is recorded as {employment} — contact the '
                    'VPSEA office if that is out of date.'
                )
            if not f.get('appointment_paper') and not (existing and existing.appointment_paper):
                errors.append('Appointment paper document is required.')

        if not errors:
            full_name = f"{p.get('first_name','').strip()} {p.get('last_name','').strip()}".strip()
            try:
                yos = int(p.get('years_of_service', 0) or 0)
            except (ValueError, TypeError):
                yos = 0

            if existing:
                existing.full_name = full_name or existing.full_name
                existing.contact_number = p.get('contact_number', existing.contact_number)
                existing.barangay = p.get('barangay', existing.barangay)
                existing.municipality = p.get('municipality', existing.municipality)
                existing.province = p.get('province', existing.province)
                if p.get('date_of_birth'):
                    existing.date_of_birth = p.get('date_of_birth')
                if p.get('gender'):
                    existing.gender = p.get('gender')
                if p.get('course'):
                    existing.course = p.get('course')
                if p.get('student_number'):
                    existing.student_id = p.get('student_number')
                existing.employment_status = p.get('employment_status', existing.employment_status)
                existing.designation = p.get('designation', existing.designation)
                if yos:
                    existing.years_of_service = yos
                if p.get('date_of_regularization'):
                    existing.date_of_regularization = p.get('date_of_regularization')
                existing.is_nsu_staff = True
                existing.status = 'Pending Validation'
                existing.remarks = ''
                if f.get('appointment_paper'):
                    existing.appointment_paper = f.get('appointment_paper')
                existing.save()
            else:
                existing = AffirmativeStaffApplication.objects.create(
                    full_name=full_name,
                    email=user.email,
                    contact_number=p.get('contact_number', ''),
                    barangay=p.get('barangay', ''),
                    municipality=p.get('municipality', ''),
                    province=p.get('province', ''),
                    date_of_birth=p.get('date_of_birth') or '2000-01-01',
                    gender=p.get('gender', ''),
                    course=p.get('course', ''),
                    year_level=int(p.get('year_level', 1) or 1),
                    student_id=p.get('student_number', ''),
                    is_nsu_staff=True,
                    employment_status=p.get('employment_status', ''),
                    designation=p.get('designation', ''),
                    years_of_service=yos or None,
                    date_of_regularization=p.get('date_of_regularization') or None,
                    appointment_paper=f.get('appointment_paper') or None,
                    qualified_for='Staff',
                    status=new_status,
                )
            # The application is the snapshot the office reviews; the profile
            # is the live record. What the staff member just entered about
            # themselves belongs on both.
            staff.employee_id       = p.get('student_number', '').strip() or staff.employee_id
            staff.contact_number    = p.get('contact_number', '').strip() or staff.contact_number
            staff.gender            = p.get('gender', '') or staff.gender
            staff.barangay          = p.get('barangay', '') or staff.barangay
            staff.municipality      = p.get('municipality', '') or staff.municipality
            staff.province          = p.get('province', '') or staff.province
            staff.employment_status = p.get('employment_status', '') or staff.employment_status
            staff.designation       = p.get('designation', '') or staff.designation
            if yos:
                staff.declared_years_of_service = yos
            dob, ok = _parse_date(p.get('date_of_birth', ''))
            if ok and dob:
                staff.date_of_birth = dob
            dor, ok = _parse_date(p.get('date_of_regularization', ''))
            if ok and dor:
                staff.date_of_regularization = dor
            # Point at the file the application just stored rather than
            # uploading a second copy of the same document.
            if existing.appointment_paper:
                staff.appointment_paper = existing.appointment_paper.name
            staff.save()

            param = 'submitted'
            return redirect(f'/nsu-staff/apply/?{param}=1')

    return render(request, 'nsu_staff/apply.html', {
        'blocked': False,
        'existing': existing,
        'submitted': request.GET.get('submitted'),
        'errors': errors,
        'user': user,
        'bipsu_courses': BIPSU_COURSES,
        'bipsu_schools': BIPSU_SCHOOLS,
        'enrolled': _nsu_staff_enrolled(user),
        # Pre-fill from existing draft/rejected record, or fall back to the
        # User's account fields so the staff member doesn't retype their own info.
        # Filled from the draft/rejected application first, then from the staff
        # member's own profile, then from their account — so nothing already on
        # file has to be retyped.
        'prefill': {
            'first_name':   existing.full_name.split()[0] if existing and existing.full_name else user.first_name,
            'last_name':    existing.full_name.split()[-1] if existing and existing.full_name and len(existing.full_name.split()) > 1 else user.last_name,
            'date_of_birth': _pick_date(existing, staff, 'date_of_birth'),
            'gender':        _pick(existing, staff, 'gender'),
            'contact_number': _pick(existing, staff, 'contact_number'),
            'barangay':      _pick(existing, staff, 'barangay'),
            'municipality':  _pick(existing, staff, 'municipality'),
            'province':      _pick(existing, staff, 'province'),
            'student_number': (existing.student_id if existing and existing.student_id
                               else staff.employee_id),
            'year_level':    existing.year_level if existing else 1,
            'course':        existing.course if existing else '',
            'employment_status':      _pick(existing, staff, 'employment_status'),
            'designation':            _pick(existing, staff, 'designation'),
            'years_of_service':       (existing.years_of_service if existing and existing.years_of_service is not None
                                       else staff.years_of_service if staff.years_of_service is not None else ''),
            'date_of_regularization': _pick_date(existing, staff, 'date_of_regularization'),
        },
    })
