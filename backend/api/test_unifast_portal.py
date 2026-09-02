"""Announcements, analytics and reports on the UniFAST portal.

Also guards the VPSEA analytics page, which now shares the same template body
and context builder.
"""
import openpyxl
from io import BytesIO
from unittest import mock

from django.test import TestCase, Client

from api import doc_convert
from api.models import (
    User, StudentProfile, Scholarship, Application, Announcement,
    TESApplication, SystemSettings,
)
from api.test_support import pdf_text


class UniFASTPortalTest(TestCase):
    def setUp(self):
        SystemSettings.objects.create(pk=1, academic_year='26-1', active_semester='1st Semester')
        self.unifast = User.objects.create_user(
            username='unifast@bipsu.edu.ph', email='unifast@bipsu.edu.ph',
            password='pw', role='unifast', first_name='Uni', last_name='Fast',
        )
        self.vpsea = User.objects.create_user(
            username='vpsea@bipsu.edu.ph', email='vpsea@bipsu.edu.ph',
            password='pw', role='vpsea',
        )
        for stype, name in (('TDP', 'TDP Scholarship'), ('TES', 'TES'),
                            ('Academic', 'Academic Scholarship')):
            Scholarship.objects.create(
                name=name, type=stype, category='application',
                description='x', eligibility='x', requirements=[],
            )
        self.c = Client()
        self.assertTrue(self.c.login(email='unifast@bipsu.edu.ph', password='pw'))

    def _scholar(self, stype, last, gender, sid):
        u = User.objects.create_user(
            username=f'{sid}@bipsu.edu.ph', email=f'{sid}@bipsu.edu.ph',
            password='pw', first_name='Test', last_name=last, role='student',
        )
        p = StudentProfile.objects.create(
            user=u, student_id=sid, course='BSCS', year_level=2, gender=gender,
            barangay='Poblacion', municipality='Naval', province='Biliran',
        )
        return Application.objects.create(
            student=p, scholarship=Scholarship.objects.get(type=stype),
            status='Approved',
            school_year='2026-2027', semester='1st Semester',
            award_number=f'AW-{sid}',
        )

    # ── Announcements ───────────────────────────────────────────────────────
    def test_unifast_can_publish_an_announcement(self):
        r = self.c.post('/unifast/announcements/', {
            'title': 'TES payout schedule', 'body': 'Releasing on the 15th.',
        })
        self.assertEqual(r.status_code, 302)
        a = Announcement.objects.get()
        self.assertEqual(a.title, 'TES payout schedule')
        self.assertEqual(a.published_by, self.unifast)

    def test_empty_announcement_is_rejected(self):
        r = self.c.post('/unifast/announcements/', {'title': '  ', 'body': ''})
        self.assertContains(r, 'Both a title and a body are required')
        self.assertEqual(Announcement.objects.count(), 0)

    def test_announcement_list_marks_who_published(self):
        Announcement.objects.create(title='Ours', body='b', published_by=self.unifast)
        Announcement.objects.create(title='Theirs', body='b', published_by=self.vpsea)
        r = self.c.get('/unifast/announcements/')
        self.assertContains(r, 'Ours')
        self.assertContains(r, 'Theirs')
        self.assertContains(r, 'Posted by UniFAST')

    # ── Analytics ───────────────────────────────────────────────────────────
    def test_analytics_is_scoped_to_tes_and_tdp_only(self):
        self._scholar('TDP', 'Cruz', 'F', '2024-0001')
        self._scholar('TES', 'Reyes', 'M', '2024-0002')
        self._scholar('Academic', 'Santos', 'F', '2024-0003')

        r = self.c.get('/unifast/analytics/')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.context['all_types'], ['TDP', 'TES'])
        self.assertEqual(r.context['rollover_counts'], {'TDP': 1, 'TES': 1})
        self.assertNotIn('Academic', r.context['rollover_counts'])
        # Needs-based programmes are not banded by GWA.
        self.assertEqual(r.context['gpa_ranges'], [])
        self.assertNotContains(r, 'GWA Distribution')

    def test_vpsea_analytics_still_works_after_the_refactor(self):
        self._scholar('Academic', 'Santos', 'F', '2024-0003')
        self._scholar('TDP', 'Cruz', 'F', '2024-0001')
        a = Client()
        self.assertTrue(a.login(email='vpsea@bipsu.edu.ph', password='pw'))
        r = a.get('/vpsea/analytics/')
        self.assertEqual(r.status_code, 200)
        self.assertIn('Academic', r.context['all_types'])
        self.assertIn('TES', r.context['all_types'])
        self.assertEqual(r.context['rollover_counts']['Academic'], 1)
        # VPSEA keeps the GWA panel.
        self.assertEqual(len(r.context['gpa_ranges']), 5)
        self.assertContains(r, 'GWA Distribution')

    # ── Reports ─────────────────────────────────────────────────────────────
    def test_reports_lists_only_tdp_and_tes_split_by_gender(self):
        self._scholar('TDP', 'Cruz', 'F', '2024-0001')
        self._scholar('TDP', 'Bautista', 'M', '2024-0004')
        self._scholar('TES', 'Reyes', 'M', '2024-0002')
        self._scholar('Academic', 'Santos', 'F', '2024-0003')

        r = self.c.get('/unifast/reports/')
        self.assertEqual(r.status_code, 200)
        sections = r.context['sections']
        self.assertEqual([s['key'] for s in sections], ['tdp', 'tes'])
        self.assertEqual(r.context['grand_total'], 3)

        tdp = sections[0]
        self.assertEqual(len(tdp['female_rows']), 1)
        self.assertEqual(len(tdp['male_rows']), 1)
        self.assertEqual(tdp['female_rows'][0]['last'], 'Cruz')
        self.assertEqual(tdp['female_rows'][0]['award'], 'AW-2024-0001')

    def test_reports_render_with_no_scholars_at_all(self):
        r = self.c.get('/unifast/reports/')
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.context['grand_total'], 0)
        self.assertEqual(r.context['tes_total'], 0)
        self.assertContains(r, 'No approved TES grantees yet')

    def test_excel_download_contains_both_programmes(self):
        self._scholar('TDP', 'Cruz', 'F', '2024-0001')
        self._scholar('TES', 'Reyes', 'M', '2024-0002')

        r = self.c.get('/unifast/reports/download/excel/')
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        self.assertIn('UniFAST_Scholars', r['Content-Disposition'])

        ws = openpyxl.load_workbook(BytesIO(r.content)).active
        text = '\n'.join(
            str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
        )
        self.assertIn('TULONG DUNONG', text)
        self.assertIn('TERTIARY EDUCATION SUBSIDY', text)
        self.assertIn('Cruz', text)
        self.assertIn('Reyes', text)
        self.assertIn('FEMALE', text)
        self.assertIn('MALE', text)

    # ── Access control ──────────────────────────────────────────────────────
    def test_the_new_pages_are_closed_to_other_roles(self):
        a = Client()
        a.login(email='vpsea@bipsu.edu.ph', password='pw')
        for url in ('/unifast/announcements/', '/unifast/analytics/',
                    '/unifast/reports/', '/unifast/reports/preview/',
                    '/unifast/reports/download/excel/',
                    '/unifast/reports/download/tes/'):
            r = a.get(url)
            self.assertEqual(r.status_code, 302, url)
            self.assertIn('/login/', r['Location'], url)


class TESValidationWorkbookTest(TestCase):
    """The CHED Annex 2 workbook download is the office template, filled in."""

    def setUp(self):
        SystemSettings.objects.create(pk=1, academic_year='26-1', active_semester='1st Semester')
        User.objects.create_user(
            username='unifast@bipsu.edu.ph', email='unifast@bipsu.edu.ph',
            password='pw', role='unifast',
        )
        Scholarship.objects.create(
            name='TES', type='TES', category='application',
            description='x', eligibility='x', requirements=[],
        )
        self.c = Client()
        self.assertTrue(self.c.login(email='unifast@bipsu.edu.ph', password='pw'))

    def _grantee(self, last, first, sid, gender='F', pwd=False, approved=True):
        from datetime import date
        u = User.objects.create_user(
            username=f'{sid}@bipsu.edu.ph', email=f'{sid}@bipsu.edu.ph', password='pw',
            first_name=first, last_name=last, role='student',
        )
        p = StudentProfile.objects.create(
            user=u, student_id=sid, course='BSCS', year_level=3, gender=gender,
            contact_number='09171234567',
            # The middle name lives on the profile; the TES export reads it there.
            middle_name='Santos',
        )
        if approved:
            Application.objects.create(
                student=p, scholarship=Scholarship.objects.get(type='TES'),
                status='Approved', award_number=f'AW-{sid}',
            )
        return TESApplication.objects.create(
            student=p, birthdate=date(2004, 5, 9),
            complete_program='Bachelor of Science in Computer Science',
            contact_number='09171234567',
            disability_type='Visual impairment' if pwd else 'N/A',
            award_number=f'AW-{sid}' if approved else '',
            status='Approved' if approved else 'Pending',
        )

    def test_rows_map_onto_the_ched_columns(self):
        from api import tes_report
        self._grantee('Cruz', 'Ana', '2024-0001', pwd=True)
        rows = tes_report.grantee_rows(batch='On-going')
        self.assertEqual(len(rows), 1)
        r = rows[0]
        self.assertEqual(r['seq'], 1)
        self.assertEqual(r['student_no'], '2024-0001')
        self.assertEqual(r['award_no'], 'AW-2024-0001')
        self.assertEqual(r['last_name'], 'Cruz')
        self.assertEqual(r['first_name'], 'Ana')
        self.assertEqual(r['middle_name'], 'Santos')
        self.assertEqual(r['sex'], 'F')
        self.assertEqual(r['course'], 'Bachelor of Science in Computer Science')
        self.assertEqual(r['year_level'], 3)
        self.assertEqual(r['batch'], 'On-going')
        self.assertIsNone(r['tes_amount'])      # CHED sets it, the office enters it
        self.assertIsNone(r['tes_3a_amount'])
        self.assertEqual(r['remarks'], 'Enrolled')
        self.assertTrue(r['is_pwd'])

    def test_only_approved_tes_applications_are_listed(self):
        from api import tes_report
        self._grantee('Cruz', 'Ana', '2024-0001')
        self._grantee('Reyes', 'Ben', '2024-0002', approved=False)
        self.assertEqual([r['last_name'] for r in tes_report.grantee_rows()], ['Cruz'])

    def test_download_preserves_every_sheet_and_writes_the_data(self):
        self._grantee('Cruz', 'Ana', '2024-0001')
        self._grantee('Bautista', 'Ben', '2024-0002', gender='M')

        r = self.c.get('/unifast/reports/download/tes/', {'batch': 'On-going (10,11)'})
        self.assertEqual(r.status_code, 200)
        self.assertIn('spreadsheetml', r['Content-Type'])
        self.assertIn('TES_Validation_Billing', r['Content-Disposition'])

        wb = openpyxl.load_workbook(BytesIO(r.content))
        self.assertEqual(
            wb.sheetnames,
            ['Official List', 'Annex 2-Form 2', 'Annex 2-Form 1', 'Annex 2-Form 4'],
        )

        ws = wb['Official List']
        # The caption carries the batch, semester and AY.
        self.assertIn('On-going (10,11)', ws['A3'].value)
        self.assertIn('AY 2026-2027', ws['A3'].value)
        # Header row untouched; data starts on row 20, alphabetical by last name.
        self.assertEqual(ws['A19'].value, 'SEQ')
        self.assertEqual(ws['A20'].value, 1)
        self.assertEqual(ws['B20'].value, '2024-0002')
        self.assertEqual(ws['D20'].value, 'Bautista')
        self.assertEqual(ws['D21'].value, 'Cruz')
        self.assertEqual(ws['P20'].value, 'Enrolled')
        # Unused pre-formatted rows are cleared and hidden, not left numbered.
        self.assertIsNone(ws['A22'].value)
        self.assertTrue(ws.row_dimensions[22].hidden)
        # Signatory block survives.
        self.assertEqual(ws['B1113'].value, 'Prepared by:')

    def test_form2_control_numbers_and_totals_formulas_survive(self):
        self._grantee('Cruz', 'Ana', '2024-0001')
        r = self.c.get('/unifast/reports/download/tes/')
        ws = openpyxl.load_workbook(BytesIO(r.content))['Annex 2-Form 2']

        self.assertEqual(ws['A31'].value, '00001')
        self.assertEqual(ws['B31'].value, '2024-0001')
        self.assertEqual(ws['D31'].value, 'Cruz')
        # Both amount columns go out blank for the office to complete.
        self.assertIsNone(ws['N31'].value)
        self.assertEqual(ws['Q31'].value, 0)
        # The template's own totals still sum the data range, so Excel computes
        # the billing statement on open.
        self.assertEqual(ws['N1101'].value, '=SUM(N31:N1100)')
        self.assertEqual(ws['Q1103'].value, '=SUM(Q1101:Q1102)')
        self.assertEqual(ws['A1108'].value, 'Prepared by:')

    def test_billing_statement_gets_the_headcount(self):
        self._grantee('Cruz', 'Ana', '2024-0001')
        self._grantee('Bautista', 'Ben', '2024-0002', gender='M')
        r = self.c.get('/unifast/reports/download/tes/')
        wb = openpyxl.load_workbook(BytesIO(r.content))
        self.assertEqual(wb['Annex 2-Form 1']['Q24'].value, 2)
        self.assertEqual(wb['Annex 2-Form 1']['V25'].value, "='Annex 2-Form 2'!N1101")

    def test_reports_page_frames_the_official_list_as_a_pdf(self):
        self._grantee('Cruz', 'Ana', '2024-0001', pwd=True)
        r = self.c.get('/unifast/reports/', {'sy': '2026-2027'})
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r.context['tes_total'], 1)
        self.assertEqual(r.context['pwd_count'], 1)
        self.assertTrue(r.context['template_available'])
        self.assertNotIn('tes_amount_total', r.context)
        body = r.content.decode()
        self.assertIn('/unifast/reports/preview/', body)
        self.assertIn('report-preview-frame', body)
        # The school year follows the page into the frame. The TES Batch was
        # dropped from this toolbar — CHED's bookkeeping label is not something
        # the office filters its own reports by.
        self.assertIn('/unifast/reports/preview/?sy=2026-2027', body)
        self.assertNotIn('TES Batch', body)

    def test_the_frame_serves_the_converted_ched_workbook(self):
        self._grantee('Cruz', 'Ana', '2024-0001')
        with mock.patch('api.doc_convert.available', return_value=True), \
                mock.patch('api.doc_convert.to_pdf',
                           return_value=b'%PDF-1.7 converted') as convert:
            r = self.c.get('/unifast/reports/preview/', {'batch': 'On-going'})

        self.assertEqual(r.content, b'%PDF-1.7 converted')
        source, suffix = convert.call_args[0]
        self.assertEqual(suffix, '.xlsx')
        # What went to the converter is the filled workbook itself, all sheets.
        wb = openpyxl.load_workbook(BytesIO(source))
        self.assertEqual(
            wb.sheetnames,
            ['Official List', 'Annex 2-Form 2', 'Annex 2-Form 1', 'Annex 2-Form 4'])
        self.assertEqual(wb['Official List']['D20'].value, 'Cruz')

    def test_the_stand_in_is_used_when_no_converter_is_installed(self):
        self._grantee('Cruz', 'Ana', '2024-0001', pwd=True)
        with mock.patch('api.doc_convert.available', return_value=False):
            r = self.c.get('/unifast/reports/preview/', {'batch': 'On-going'})

        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        self.assertTrue(r.content.startswith(b'%PDF'))

        # Column headings wrap inside their cells, so they land in the page as
        # separate words — assert on the words, not the joined heading.
        text = pdf_text(r.content)
        self.assertIn('LIST OF CONTINUING TES GRANTEES SUBJECT FOR VALIDATION', text)
        self.assertIn('BILIRAN PROVINCE STATE UNIVERSITY', text)
        for word in ('STUDENT', 'AWARD', 'BIRTHDATE', 'VALIDATION', 'REMARKS'):
            self.assertIn(word, text, word)
        self.assertIn('Cruz', text)
        self.assertIn('2024-0001', text)
        self.assertIn('AW-2024-0001', text)
        # And it says on its face that it is not the office workbook.
        self.assertIn('STAND-IN LAYOUT', text)

    def test_a_failed_conversion_falls_back_instead_of_erroring(self):
        with mock.patch('api.doc_convert.available', return_value=True), \
                mock.patch('api.doc_convert.to_pdf',
                           side_effect=doc_convert.ConversionFailed('no pdf')):
            r = self.c.get('/unifast/reports/preview/')

        self.assertEqual(r.status_code, 200)
        self.assertIn('STAND-IN LAYOUT', pdf_text(r.content))

    def test_the_preview_renders_with_no_grantees_at_all(self):
        with mock.patch('api.doc_convert.available', return_value=False):
            r = self.c.get('/unifast/reports/preview/')
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.content.startswith(b'%PDF'))

    def test_the_preview_may_be_framed_by_our_own_pages(self):
        with mock.patch('api.doc_convert.available', return_value=False):
            r = self.c.get('/unifast/reports/preview/')
        self.assertNotEqual(r.headers.get('X-Frame-Options'), 'DENY')
